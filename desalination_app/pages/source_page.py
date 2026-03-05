"""
WT-Desal Lab — Source Page. Light "Frost" theme.
"""

import customtkinter as ctk
import requests
import threading
import csv
import os
from tkinter import filedialog
from datetime import datetime, timezone
from theme import Colors, Fonts, Spacing

# Turbine specifications database with Vd (cut-in), Vn (rated speed), Pn (rated power), rotor diameter
# Based on manufacturer datasheets and typical values
TURBINE_DATABASE = {
    # Vestas (cut_out_speed added for Model 2 power curve)
    "Vestas V27 225": {"rated_power_kw": 225, "rotor_diameter_m": 27, "cut_in_speed": 3.5, "rated_speed": 13.0, "cut_out_speed": 25.0},
    "Vestas V29 225": {"rated_power_kw": 225, "rotor_diameter_m": 29, "cut_in_speed": 3.5, "rated_speed": 13.0, "cut_out_speed": 25.0},
    "Vestas V39 500": {"rated_power_kw": 500, "rotor_diameter_m": 39, "cut_in_speed": 3.5, "rated_speed": 13.0, "cut_out_speed": 25.0},
    "Vestas V42 600": {"rated_power_kw": 600, "rotor_diameter_m": 42, "cut_in_speed": 3.5, "rated_speed": 13.0, "cut_out_speed": 25.0},
    "Vestas V44 600": {"rated_power_kw": 600, "rotor_diameter_m": 44, "cut_in_speed": 3.5, "rated_speed": 13.0, "cut_out_speed": 25.0},
    "Vestas V47 660": {"rated_power_kw": 660, "rotor_diameter_m": 47, "cut_in_speed": 3.5, "rated_speed": 14.0, "cut_out_speed": 25.0},
    "Vestas V52 850": {"rated_power_kw": 850, "rotor_diameter_m": 52, "cut_in_speed": 3.5, "rated_speed": 14.0, "cut_out_speed": 25.0},
    "Vestas V66 1650": {"rated_power_kw": 1650, "rotor_diameter_m": 66, "cut_in_speed": 3.5, "rated_speed": 13.0, "cut_out_speed": 25.0},
    "Vestas V66 1750": {"rated_power_kw": 1750, "rotor_diameter_m": 66, "cut_in_speed": 3.5, "rated_speed": 13.0},
    "Vestas V66 2000": {"rated_power_kw": 2000, "rotor_diameter_m": 66, "cut_in_speed": 3.5, "rated_speed": 13.0},
    "Vestas V80 1800": {"rated_power_kw": 1800, "rotor_diameter_m": 80, "cut_in_speed": 4.0, "rated_speed": 14.0},
    "Vestas V80 2000": {"rated_power_kw": 2000, "rotor_diameter_m": 80, "cut_in_speed": 4.0, "rated_speed": 14.0},
    "Vestas V82 1650": {"rated_power_kw": 1650, "rotor_diameter_m": 82, "cut_in_speed": 3.5, "rated_speed": 13.0},
    "Vestas V90 1800": {"rated_power_kw": 1800, "rotor_diameter_m": 90, "cut_in_speed": 3.5, "rated_speed": 12.0},
    "Vestas V90 2000": {"rated_power_kw": 2000, "rotor_diameter_m": 90, "cut_in_speed": 3.5, "rated_speed": 12.0},
    "Vestas V90 3000": {"rated_power_kw": 3000, "rotor_diameter_m": 90, "cut_in_speed": 3.0, "rated_speed": 12.0},
    "Vestas V100 1800": {"rated_power_kw": 1800, "rotor_diameter_m": 100, "cut_in_speed": 3.0, "rated_speed": 11.0},
    "Vestas V100 2000": {"rated_power_kw": 2000, "rotor_diameter_m": 100, "cut_in_speed": 3.0, "rated_speed": 11.0},
    "Vestas V100 2600": {"rated_power_kw": 2600, "rotor_diameter_m": 100, "cut_in_speed": 3.0, "rated_speed": 11.0},
    "Vestas V105 3300": {"rated_power_kw": 3300, "rotor_diameter_m": 105, "cut_in_speed": 3.0, "rated_speed": 11.5},
    "Vestas V105 3450": {"rated_power_kw": 3450, "rotor_diameter_m": 105, "cut_in_speed": 3.0, "rated_speed": 11.5},
    "Vestas V110 2000": {"rated_power_kw": 2000, "rotor_diameter_m": 110, "cut_in_speed": 3.0, "rated_speed": 11.0},
    "Vestas V112 3000": {"rated_power_kw": 3000, "rotor_diameter_m": 112, "cut_in_speed": 3.0, "rated_speed": 11.0},
    "Vestas V112 3300": {"rated_power_kw": 3300, "rotor_diameter_m": 112, "cut_in_speed": 3.0, "rated_speed": 11.0},
    "Vestas V112 3450": {"rated_power_kw": 3450, "rotor_diameter_m": 112, "cut_in_speed": 3.0, "rated_speed": 11.0},
    "Vestas V117 3300": {"rated_power_kw": 3300, "rotor_diameter_m": 117, "cut_in_speed": 3.0, "rated_speed": 11.0},
    "Vestas V117 3450": {"rated_power_kw": 3450, "rotor_diameter_m": 117, "cut_in_speed": 3.0, "rated_speed": 11.0},
    "Vestas V117 3600": {"rated_power_kw": 3600, "rotor_diameter_m": 117, "cut_in_speed": 3.0, "rated_speed": 11.0},
    "Vestas V117 4000": {"rated_power_kw": 4000, "rotor_diameter_m": 117, "cut_in_speed": 3.0, "rated_speed": 11.0},
    "Vestas V120 2200": {"rated_power_kw": 2200, "rotor_diameter_m": 120, "cut_in_speed": 3.0, "rated_speed": 11.0},
    "Vestas V126 3000": {"rated_power_kw": 3000, "rotor_diameter_m": 126, "cut_in_speed": 3.0, "rated_speed": 11.0},
    "Vestas V126 3300": {"rated_power_kw": 3300, "rotor_diameter_m": 126, "cut_in_speed": 3.0, "rated_speed": 11.0},
    "Vestas V126 3450": {"rated_power_kw": 3450, "rotor_diameter_m": 126, "cut_in_speed": 3.0, "rated_speed": 11.0},
    "Vestas V136 3450": {"rated_power_kw": 3450, "rotor_diameter_m": 136, "cut_in_speed": 3.0, "rated_speed": 10.5},
    "Vestas V136 4000": {"rated_power_kw": 4000, "rotor_diameter_m": 136, "cut_in_speed": 3.0, "rated_speed": 10.5},
    "Vestas V150 4000": {"rated_power_kw": 4000, "rotor_diameter_m": 150, "cut_in_speed": 3.0, "rated_speed": 10.0},
    "Vestas V150 4200": {"rated_power_kw": 4200, "rotor_diameter_m": 150, "cut_in_speed": 3.0, "rated_speed": 10.0},
    "Vestas V150 4500": {"rated_power_kw": 4500, "rotor_diameter_m": 150, "cut_in_speed": 3.0, "rated_speed": 10.0},
    "Vestas V150 5600": {"rated_power_kw": 5600, "rotor_diameter_m": 150, "cut_in_speed": 3.0, "rated_speed": 10.0},
    "Vestas V150 6000": {"rated_power_kw": 6000, "rotor_diameter_m": 150, "cut_in_speed": 3.0, "rated_speed": 10.0},
    "Vestas V162 5600": {"rated_power_kw": 5600, "rotor_diameter_m": 162, "cut_in_speed": 3.0, "rated_speed": 10.0},
    "Vestas V162 6000": {"rated_power_kw": 6000, "rotor_diameter_m": 162, "cut_in_speed": 3.0, "rated_speed": 10.0},
    "Vestas V162 6200": {"rated_power_kw": 6200, "rotor_diameter_m": 162, "cut_in_speed": 3.0, "rated_speed": 10.0},
    "Vestas V162 7200": {"rated_power_kw": 7200, "rotor_diameter_m": 162, "cut_in_speed": 3.0, "rated_speed": 10.0},
    "Vestas V164 7000": {"rated_power_kw": 7000, "rotor_diameter_m": 164, "cut_in_speed": 3.0, "rated_speed": 10.0},
    "Vestas V164 8000": {"rated_power_kw": 8000, "rotor_diameter_m": 164, "cut_in_speed": 3.0, "rated_speed": 10.0},
    "Vestas V164 9500": {"rated_power_kw": 9500, "rotor_diameter_m": 164, "cut_in_speed": 3.0, "rated_speed": 10.0},
    "Vestas V172 7200": {"rated_power_kw": 7200, "rotor_diameter_m": 172, "cut_in_speed": 3.0, "rated_speed": 10.0},
    # Enercon
    "Enercon E40 500": {"rated_power_kw": 500, "rotor_diameter_m": 40, "cut_in_speed": 2.5, "rated_speed": 12.0},
    "Enercon E40 600": {"rated_power_kw": 600, "rotor_diameter_m": 40, "cut_in_speed": 2.5, "rated_speed": 12.0},
    "Enercon E44 900": {"rated_power_kw": 900, "rotor_diameter_m": 44, "cut_in_speed": 2.5, "rated_speed": 12.0},
    "Enercon E48 800": {"rated_power_kw": 800, "rotor_diameter_m": 48, "cut_in_speed": 2.5, "rated_speed": 12.0},
    "Enercon E53 800": {"rated_power_kw": 800, "rotor_diameter_m": 53, "cut_in_speed": 2.5, "rated_speed": 12.0},
    "Enercon E66 1500": {"rated_power_kw": 1500, "rotor_diameter_m": 66, "cut_in_speed": 2.5, "rated_speed": 12.0},
    "Enercon E66 1800": {"rated_power_kw": 1800, "rotor_diameter_m": 66, "cut_in_speed": 2.5, "rated_speed": 12.0},
    "Enercon E66 2000": {"rated_power_kw": 2000, "rotor_diameter_m": 66, "cut_in_speed": 2.5, "rated_speed": 12.0},
    "Enercon E70 2000": {"rated_power_kw": 2000, "rotor_diameter_m": 70, "cut_in_speed": 2.5, "rated_speed": 12.0},
    "Enercon E70 2300": {"rated_power_kw": 2300, "rotor_diameter_m": 70, "cut_in_speed": 2.5, "rated_speed": 12.0},
    "Enercon E82 1800": {"rated_power_kw": 1800, "rotor_diameter_m": 82, "cut_in_speed": 2.5, "rated_speed": 12.0},
    "Enercon E82 2000": {"rated_power_kw": 2000, "rotor_diameter_m": 82, "cut_in_speed": 2.5, "rated_speed": 12.0},
    "Enercon E82 2300": {"rated_power_kw": 2300, "rotor_diameter_m": 82, "cut_in_speed": 2.5, "rated_speed": 12.0},
    "Enercon E82 3000": {"rated_power_kw": 3000, "rotor_diameter_m": 82, "cut_in_speed": 2.5, "rated_speed": 11.0},
    "Enercon E92 2300": {"rated_power_kw": 2300, "rotor_diameter_m": 92, "cut_in_speed": 2.5, "rated_speed": 11.0},
    "Enercon E92 2350": {"rated_power_kw": 2350, "rotor_diameter_m": 92, "cut_in_speed": 2.5, "rated_speed": 11.0},
    "Enercon E101 3000": {"rated_power_kw": 3000, "rotor_diameter_m": 101, "cut_in_speed": 2.5, "rated_speed": 11.0},
    "Enercon E101 3500": {"rated_power_kw": 3500, "rotor_diameter_m": 101, "cut_in_speed": 2.5, "rated_speed": 11.0},
    "Enercon E103 2350": {"rated_power_kw": 2350, "rotor_diameter_m": 103, "cut_in_speed": 2.5, "rated_speed": 11.0},
    "Enercon E112 4500": {"rated_power_kw": 4500, "rotor_diameter_m": 112, "cut_in_speed": 2.5, "rated_speed": 11.0},
    "Enercon E115 2500": {"rated_power_kw": 2500, "rotor_diameter_m": 115, "cut_in_speed": 2.5, "rated_speed": 11.0},
    "Enercon E115 3000": {"rated_power_kw": 3000, "rotor_diameter_m": 115, "cut_in_speed": 2.5, "rated_speed": 11.0},
    "Enercon E115 3200": {"rated_power_kw": 3200, "rotor_diameter_m": 115, "cut_in_speed": 2.5, "rated_speed": 11.0},
    "Enercon E126 3500": {"rated_power_kw": 3500, "rotor_diameter_m": 126, "cut_in_speed": 2.5, "rated_speed": 11.0},
    "Enercon E126 4000": {"rated_power_kw": 4000, "rotor_diameter_m": 126, "cut_in_speed": 2.5, "rated_speed": 11.0},
    "Enercon E126 4200": {"rated_power_kw": 4200, "rotor_diameter_m": 126, "cut_in_speed": 2.5, "rated_speed": 11.0},
    "Enercon E126 6500": {"rated_power_kw": 6500, "rotor_diameter_m": 126, "cut_in_speed": 2.5, "rated_speed": 11.0},
    "Enercon E126 7000": {"rated_power_kw": 7000, "rotor_diameter_m": 126, "cut_in_speed": 2.5, "rated_speed": 11.0},
    "Enercon E126 7500": {"rated_power_kw": 7500, "rotor_diameter_m": 126, "cut_in_speed": 2.5, "rated_speed": 11.0},
    "Enercon E138 3500": {"rated_power_kw": 3500, "rotor_diameter_m": 138, "cut_in_speed": 2.5, "rated_speed": 11.0},
    "Enercon E138 4260": {"rated_power_kw": 4260, "rotor_diameter_m": 138, "cut_in_speed": 2.5, "rated_speed": 11.0},
    "Enercon E141 4200": {"rated_power_kw": 4200, "rotor_diameter_m": 141, "cut_in_speed": 2.5, "rated_speed": 11.0},
    "Enercon E160 4600": {"rated_power_kw": 4600, "rotor_diameter_m": 160, "cut_in_speed": 2.5, "rated_speed": 11.0},
    "Enercon E160 5560": {"rated_power_kw": 5560, "rotor_diameter_m": 160, "cut_in_speed": 2.5, "rated_speed": 11.0},
    "Enercon E175 6000": {"rated_power_kw": 6000, "rotor_diameter_m": 175, "cut_in_speed": 2.5, "rated_speed": 11.0},
    # GE
    "GE 1.5s": {"rated_power_kw": 1500, "rotor_diameter_m": 70.5, "cut_in_speed": 3.5, "rated_speed": 11.5},
    "GE 1.5se": {"rated_power_kw": 1500, "rotor_diameter_m": 77, "cut_in_speed": 3.5, "rated_speed": 11.5},
    "GE 1.5sl": {"rated_power_kw": 1500, "rotor_diameter_m": 77, "cut_in_speed": 3.5, "rated_speed": 11.5},
    "GE 1.5sle": {"rated_power_kw": 1500, "rotor_diameter_m": 77, "cut_in_speed": 3.5, "rated_speed": 11.5},
    "GE 1.5xle": {"rated_power_kw": 1500, "rotor_diameter_m": 82.5, "cut_in_speed": 3.5, "rated_speed": 11.5},
    "GE 1.6": {"rated_power_kw": 1600, "rotor_diameter_m": 82.5, "cut_in_speed": 3.5, "rated_speed": 11.5},
    "GE 1.6-82.5": {"rated_power_kw": 1600, "rotor_diameter_m": 82.5, "cut_in_speed": 3.5, "rated_speed": 11.5},
    "GE 1.7": {"rated_power_kw": 1700, "rotor_diameter_m": 103, "cut_in_speed": 3.0, "rated_speed": 11.0},
    "GE 1.85-82.5": {"rated_power_kw": 1850, "rotor_diameter_m": 82.5, "cut_in_speed": 3.0, "rated_speed": 11.0},
    "GE 1.85-87": {"rated_power_kw": 1850, "rotor_diameter_m": 87, "cut_in_speed": 3.0, "rated_speed": 11.0},
    "GE 2.5-100": {"rated_power_kw": 2500, "rotor_diameter_m": 100, "cut_in_speed": 3.0, "rated_speed": 11.0},
    "GE 2.5-103": {"rated_power_kw": 2500, "rotor_diameter_m": 103, "cut_in_speed": 3.0, "rated_speed": 11.0},
    "GE 2.5-120": {"rated_power_kw": 2500, "rotor_diameter_m": 120, "cut_in_speed": 3.0, "rated_speed": 10.0},
    "GE 2.5-88": {"rated_power_kw": 2500, "rotor_diameter_m": 88, "cut_in_speed": 3.0, "rated_speed": 11.0},
    "GE 2.5xl": {"rated_power_kw": 2500, "rotor_diameter_m": 100, "cut_in_speed": 3.0, "rated_speed": 11.0},
    "GE 2.75-103": {"rated_power_kw": 2750, "rotor_diameter_m": 103, "cut_in_speed": 3.0, "rated_speed": 11.0},
    "GE 2.75-120": {"rated_power_kw": 2750, "rotor_diameter_m": 120, "cut_in_speed": 3.0, "rated_speed": 10.0},
    "GE 2.75 103": {"rated_power_kw": 2750, "rotor_diameter_m": 103, "cut_in_speed": 3.0, "rated_speed": 11.0},
    "GE 2.85-103": {"rated_power_kw": 2850, "rotor_diameter_m": 103, "cut_in_speed": 3.0, "rated_speed": 11.0},
    "GE 3.2-103": {"rated_power_kw": 3200, "rotor_diameter_m": 103, "cut_in_speed": 3.0, "rated_speed": 11.0},
    "GE 3.2-130": {"rated_power_kw": 3200, "rotor_diameter_m": 130, "cut_in_speed": 3.0, "rated_speed": 10.0},
    "GE 3.2 103": {"rated_power_kw": 3200, "rotor_diameter_m": 103, "cut_in_speed": 3.0, "rated_speed": 11.0},
    "GE 3.2 130": {"rated_power_kw": 3200, "rotor_diameter_m": 130, "cut_in_speed": 3.0, "rated_speed": 10.0},
    "GE 3.4-130": {"rated_power_kw": 3400, "rotor_diameter_m": 130, "cut_in_speed": 3.0, "rated_speed": 10.0},
    "GE 3.4-137": {"rated_power_kw": 3400, "rotor_diameter_m": 137, "cut_in_speed": 3.0, "rated_speed": 10.0},
    "GE 3.4 130": {"rated_power_kw": 3400, "rotor_diameter_m": 130, "cut_in_speed": 3.0, "rated_speed": 10.0},
    "GE 3.6sl": {"rated_power_kw": 3600, "rotor_diameter_m": 130, "cut_in_speed": 3.0, "rated_speed": 10.0},
    "GE 3.8-130": {"rated_power_kw": 3800, "rotor_diameter_m": 130, "cut_in_speed": 3.0, "rated_speed": 10.0},
    "GE 3.8 130": {"rated_power_kw": 3800, "rotor_diameter_m": 130, "cut_in_speed": 3.0, "rated_speed": 10.0},
    "GE 5.3-158": {"rated_power_kw": 5300, "rotor_diameter_m": 158, "cut_in_speed": 3.0, "rated_speed": 10.0},
    "GE 5.3 158": {"rated_power_kw": 5300, "rotor_diameter_m": 158, "cut_in_speed": 3.0, "rated_speed": 10.0},
    "GE 5.5-158": {"rated_power_kw": 5500, "rotor_diameter_m": 158, "cut_in_speed": 3.0, "rated_speed": 10.0},
    "GE 5.5 158": {"rated_power_kw": 5500, "rotor_diameter_m": 158, "cut_in_speed": 3.0, "rated_speed": 10.0},
    "GE 6.0-164": {"rated_power_kw": 6000, "rotor_diameter_m": 164, "cut_in_speed": 3.0, "rated_speed": 10.0},
    "GE 900S": {"rated_power_kw": 900, "rotor_diameter_m": 56, "cut_in_speed": 3.5, "rated_speed": 13.0},
    "GE Haliade 6-150": {"rated_power_kw": 6000, "rotor_diameter_m": 150, "cut_in_speed": 3.0, "rated_speed": 10.0},
    "GE Haliade-X 12-220": {"rated_power_kw": 12000, "rotor_diameter_m": 220, "cut_in_speed": 3.0, "rated_speed": 10.0},
    # Gamesa
    "Gamesa G47 660": {"rated_power_kw": 660, "rotor_diameter_m": 47, "cut_in_speed": 3.5, "rated_speed": 13.0},
    "Gamesa G52 850": {"rated_power_kw": 850, "rotor_diameter_m": 52, "cut_in_speed": 3.5, "rated_speed": 13.0},
    "Gamesa G58 850": {"rated_power_kw": 850, "rotor_diameter_m": 58, "cut_in_speed": 3.5, "rated_speed": 13.0},
    "Gamesa G80 2000": {"rated_power_kw": 2000, "rotor_diameter_m": 80, "cut_in_speed": 4.0, "rated_speed": 14.0},
    "Gamesa G87 2000": {"rated_power_kw": 2000, "rotor_diameter_m": 87, "cut_in_speed": 3.5, "rated_speed": 12.0},
    "Gamesa G90 2000": {"rated_power_kw": 2000, "rotor_diameter_m": 90, "cut_in_speed": 3.5, "rated_speed": 12.0},
    "Gamesa G97 2000": {"rated_power_kw": 2000, "rotor_diameter_m": 97, "cut_in_speed": 3.5, "rated_speed": 11.5},
    "Gamesa G114 2000": {"rated_power_kw": 2000, "rotor_diameter_m": 114, "cut_in_speed": 3.0, "rated_speed": 11.0},
    "Gamesa G114 2100": {"rated_power_kw": 2100, "rotor_diameter_m": 114, "cut_in_speed": 3.0, "rated_speed": 11.0},
    "Gamesa G114 2500": {"rated_power_kw": 2500, "rotor_diameter_m": 114, "cut_in_speed": 3.0, "rated_speed": 11.0},
    "Gamesa G114 2625": {"rated_power_kw": 2625, "rotor_diameter_m": 114, "cut_in_speed": 3.0, "rated_speed": 11.0},
    "Gamesa G128 4500": {"rated_power_kw": 4500, "rotor_diameter_m": 128, "cut_in_speed": 3.0, "rated_speed": 11.0},
    "Gamesa G128 5000": {"rated_power_kw": 5000, "rotor_diameter_m": 128, "cut_in_speed": 3.0, "rated_speed": 11.0},
    # Siemens
    "Siemens SWT-2.3-93": {"rated_power_kw": 2300, "rotor_diameter_m": 93, "cut_in_speed": 3.0, "rated_speed": 12.0},
    "Siemens SWT-2.3-82": {"rated_power_kw": 2300, "rotor_diameter_m": 82, "cut_in_speed": 3.5, "rated_speed": 13.0},
    "Siemens SWT-2.3-101": {"rated_power_kw": 2300, "rotor_diameter_m": 101, "cut_in_speed": 3.0, "rated_speed": 12.0},
    "Siemens SWT-2.3-108": {"rated_power_kw": 2300, "rotor_diameter_m": 108, "cut_in_speed": 3.0, "rated_speed": 11.0},
    "Siemens SWT-3.6-107": {"rated_power_kw": 3600, "rotor_diameter_m": 107, "cut_in_speed": 3.0, "rated_speed": 11.0},
    "Siemens SWT-3.6-120": {"rated_power_kw": 3600, "rotor_diameter_m": 120, "cut_in_speed": 3.0, "rated_speed": 11.0},
    "Siemens SWT-3.6-130": {"rated_power_kw": 3600, "rotor_diameter_m": 130, "cut_in_speed": 3.0, "rated_speed": 11.0},
    # Bonus
    "Bonus B23 150": {"rated_power_kw": 150, "rotor_diameter_m": 23, "cut_in_speed": 3.5, "rated_speed": 12.0},
    "Bonus B33 300": {"rated_power_kw": 300, "rotor_diameter_m": 33, "cut_in_speed": 3.5, "rated_speed": 12.0},
    "Bonus B37 450": {"rated_power_kw": 450, "rotor_diameter_m": 37, "cut_in_speed": 3.5, "rated_speed": 12.0},
    "Bonus B41 500": {"rated_power_kw": 500, "rotor_diameter_m": 41, "cut_in_speed": 3.5, "rated_speed": 12.0},
    "Bonus B44 600": {"rated_power_kw": 600, "rotor_diameter_m": 44, "cut_in_speed": 3.5, "rated_speed": 12.0},
    "Bonus B54 1000": {"rated_power_kw": 1000, "rotor_diameter_m": 54, "cut_in_speed": 3.5, "rated_speed": 12.0},
    "Bonus B62 1300": {"rated_power_kw": 1300, "rotor_diameter_m": 62, "cut_in_speed": 3.5, "rated_speed": 12.0},
    "Bonus B82 2300": {"rated_power_kw": 2300, "rotor_diameter_m": 82, "cut_in_speed": 3.5, "rated_speed": 12.0},
    # Nordex
    "Nordex N27 150": {"rated_power_kw": 150, "rotor_diameter_m": 27, "cut_in_speed": 3.5, "rated_speed": 13.0},
    "Nordex N29 250": {"rated_power_kw": 250, "rotor_diameter_m": 29, "cut_in_speed": 3.5, "rated_speed": 13.0},
    "Nordex N43 600": {"rated_power_kw": 600, "rotor_diameter_m": 43, "cut_in_speed": 3.5, "rated_speed": 13.0},
    "Nordex N50 800": {"rated_power_kw": 800, "rotor_diameter_m": 50, "cut_in_speed": 3.5, "rated_speed": 13.0},
    "Nordex N60 1300": {"rated_power_kw": 1300, "rotor_diameter_m": 60, "cut_in_speed": 3.5, "rated_speed": 13.0},
    "Nordex N80 2500": {"rated_power_kw": 2500, "rotor_diameter_m": 80, "cut_in_speed": 3.5, "rated_speed": 13.0},
    "Nordex N90 2300": {"rated_power_kw": 2300, "rotor_diameter_m": 90, "cut_in_speed": 3.0, "rated_speed": 12.0},
    "Nordex N90 2500": {"rated_power_kw": 2500, "rotor_diameter_m": 90, "cut_in_speed": 3.0, "rated_speed": 12.0},
    "Nordex N100 2500": {"rated_power_kw": 2500, "rotor_diameter_m": 100, "cut_in_speed": 3.0, "rated_speed": 11.0},
    "Nordex N100 3300": {"rated_power_kw": 3300, "rotor_diameter_m": 100, "cut_in_speed": 3.0, "rated_speed": 11.0},
    # MHI Vestas
    "MHI Vestas V164 8000": {"rated_power_kw": 8000, "rotor_diameter_m": 164, "cut_in_speed": 3.0, "rated_speed": 12.0},
    "MHI Vestas V164 9500": {"rated_power_kw": 9500, "rotor_diameter_m": 164, "cut_in_speed": 3.0, "rated_speed": 12.0},
    "MHI Vestas V164 10000": {"rated_power_kw": 10000, "rotor_diameter_m": 164, "cut_in_speed": 3.0, "rated_speed": 12.0},
    # Default for unknown turbines
    "_default": {"rated_power_kw": 2000, "rotor_diameter_m": 80, "cut_in_speed": 3.0, "rated_speed": 12.0, "cut_out_speed": 25.0},
}

# Full list of turbine models (used by both Ninja and Energy tabs)
TURBINE_MODELS = [
    "Acciona AW77 1500",
    "Alstom Eco 110",
    "Alstom Eco 74",
    "Alstom Eco 80",
    "Bonus B23 150",
    "Bonus B33 300",
    "Bonus B37 450",
    "Bonus B41 500",
    "Bonus B44 600",
    "Bonus B54 1000",
    "Bonus B62 1300",
    "Bonus B82 2300",
    "Dewind D4 41 500",
    "Dewind D6 62 1000",
    "Dewind D6 1000",
    "Enercon E101 3000",
    "Enercon E101 3500",
    "Enercon E103 2350",
    "Enercon E112 4500",
    "Enercon E115 2500",
    "Enercon E115 3000",
    "Enercon E115 3200",
    "Enercon E126 3500",
    "Enercon E126 4000",
    "Enercon E126 4200",
    "Enercon E126 6500",
    "Enercon E126 7000",
    "Enercon E126 7500",
    "Enercon E138 3500",
    "Enercon E138 4260",
    "Enercon E141 4200",
    "Enercon E160 4600",
    "Enercon E160 5560",
    "Enercon E175 6000",
    "Enercon E40 500",
    "Enercon E40 600",
    "Enercon E44 900",
    "Enercon E48 800",
    "Enercon E53 800",
    "Enercon E66 1500",
    "Enercon E66 1800",
    "Enercon E66 2000",
    "Enercon E70 2000",
    "Enercon E70 2300",
    "Enercon E82 1800",
    "Enercon E82 2000",
    "Enercon E82 2300",
    "Enercon E82 3000",
    "Enercon E92 2300",
    "Enercon E92 2350",
    "EWT DirectWind 52 900",
    "Gamesa G114 2000",
    "Gamesa G114 2100",
    "Gamesa G114 2500",
    "Gamesa G114 2625",
    "Gamesa G128 4500",
    "Gamesa G128 5000",
    "Gamesa G47 660",
    "Gamesa G52 850",
    "Gamesa G58 850",
    "Gamesa G80 2000",
    "Gamesa G87 2000",
    "Gamesa G90 2000",
    "Gamesa G97 2000",
    "GE 1.5s",
    "GE 1.5se",
    "GE 1.5sl",
    "GE 1.5sle",
    "GE 1.5xle",
    "GE 1.6",
    "GE 1.6-82.5",
    "GE 1.7",
    "GE 1.85-82.5",
    "GE 1.85-87",
    "GE 2.5-100",
    "GE 2.5-103",
    "GE 2.5-120",
    "GE 2.5-88",
    "GE 2.5xl",
    "GE 2.75-103",
    "GE 2.75-120",
    "GE 2.75 103",
    "GE 2.85-103",
    "GE 3.2-103",
    "GE 3.2-130",
    "GE 3.2 103",
    "GE 3.2 130",
    "GE 3.4-130",
    "GE 3.4-137",
    "GE 3.4 130",
    "GE 3.6sl",
    "GE 3.8-130",
    "GE 3.8 130",
    "GE 5.3-158",
    "GE 5.3 158",
    "GE 5.5-158",
    "GE 5.5 158",
    "GE 6.0-164",
    "GE 900S",
    "GE Haliade 6-150",
    "GE Haliade-X 12-220",
    "Goldwind GW109 2500",
    "Goldwind GW121 2500",
    "Goldwind GW140 3400",
    "Goldwind GW140 3000",
    "Goldwind GW154 6700",
    "Goldwind GW82 1500",
    "Hitachi HTW5.2-127",
    "Hitachi HTW5.2-136",
    "MHI Vestas V117 4200",
    "MHI Vestas V164 10000",
    "MHI Vestas V164 8400",
    "MHI Vestas V164 9500",
    "MHI Vestas V174 9500",
    "Mingyang SCD 3000 100",
    "Mingyang SCD 3000 108",
    "Mingyang SCD 3000 92",
    "NEG Micon M1500 500",
    "NEG Micon M1500 750",
    "NEG Micon NM48 750",
    "NEG Micon NM52 900",
    "NEG Micon NM60 1000",
    "NEG Micon NM64c 1500",
    "NEG Micon NM80 2750",
    "Nordex N100 2500",
    "Nordex N100 3300",
    "Nordex N117 2400",
    "Nordex N117 3000",
    "Nordex N117 3600",
    "Nordex N131 3000",
    "Nordex N131 3300",
    "Nordex N131 3600",
    "Nordex N131 3900",
    "Nordex N149 4500",
    "Nordex N27 150",
    "Nordex N29 250",
    "Nordex N43 600",
    "Nordex N50 800",
    "Nordex N60 1300",
    "Nordex N80 2500",
    "Nordex N90 2300",
    "Nordex N90 2500",
    "Nordtank NTK500",
    "Nordtank NTK600",
    "PowerWind 56 900",
    "REpower 3.4M",
    "REpower 5M",
    "REpower 6M",
    "REpower MD70 1500",
    "REPower MD77 1500",
    "REpower MM70 2000",
    "REpower MM82 2000",
    "REpower MM92 2000",
    "Senvion 3.2M114",
    "Senvion 6.3M152",
    "Senvion MM82 2050",
    "Senvion MM92 2050",
    "Shanghai Electric W2000 105",
    "Shanghai Electric W2000 111",
    "Shanghai Electric W2000 116",
    "Shanghai Electric W2000 87",
    "Shanghai Electric W2000 93",
    "Shanghai Electric W2000 99",
    "Shanghai Electric W3600 116",
    "Shanghai Electric W3600 122",
    "Siemens Gamesa SG 10.0-193",
    "Siemens Gamesa SG 4.5-145",
    "Siemens Gamesa SG 5.0-132",
    "Siemens Gamesa SG 5.0-145",
    "Siemens Gamesa SG 6.0-154",
    "Siemens Gamesa SG 6.2-170",
    "Siemens Gamesa SG 6.6-155",
    "Siemens Gamesa SG 6.6-170",
    "Siemens Gamesa SG 7.0-154",
    "Siemens Gamesa SG 8.0-167",
    "Siemens Gamesa SG 8.5-167",
    "Siemens SWT-1.3-62",
    "Siemens SWT-2.3-101",
    "Siemens SWT-2.3-108",
    "Siemens SWT-2.3-82",
    "Siemens SWT-2.3-93",
    "Siemens SWT-2.5-108",
    "Siemens SWT-2.625-120",
    "Siemens SWT-3.0-101",
    "Siemens SWT-3.15-142",
    "Siemens SWT-3.2-101",
    "Siemens SWT-3.2-108",
    "Siemens SWT-3.2-113",
    "Siemens SWT-3.3-130",
    "Siemens SWT-3.6-107",
    "Siemens SWT-3.6-120",
    "Siemens SWT-3.6-130",
    "Siemens SWT-4.0-120",
    "Siemens SWT-4.0-130",
    "Siemens SWT-4.1-142",
    "Siemens SWT-4.3-120",
    "Siemens SWT-4.3-130",
    "Siemens SWT-6.0-154",
    "Siemens SWT-7.0-154",
    "Siemens SWT-8.0-154",
    "Siemens SWT 1.3 62",
    "Siemens SWT 2.3 82",
    "Siemens SWT 2.3 93",
    "Siemens SWT 3.0 101",
    "Siemens SWT 3.6 107",
    "Siemens SWT 3.6 120",
    "Siemens SWT 3.6 130",
    "Siemens SWT 4.0 130",
    "Siemens SWT 4.3 130",
    "Siemens SWT 4.1 142",
    "Siemens Gamesa SG 4.5 145",
    "Suzlon S88 2100",
    "Suzlon S97 2100",
    "Tacke TW600 43",
    "Vestas V100 1800",
    "Vestas V100 2000",
    "Vestas V100 2600",
    "Vestas V105 3300",
    "Vestas V105 3450",
    "Vestas V110 2000",
    "Vestas V112 3000",
    "Vestas V112 3300",
    "Vestas V112 3450",
    "Vestas V117 3300",
    "Vestas V117 3450",
    "Vestas V117 3600",
    "Vestas V117 4000",
    "Vestas V120 2200",
    "Vestas V126 3000",
    "Vestas V126 3300",
    "Vestas V126 3450",
    "Vestas V136 3450",
    "Vestas V136 4000",
    "Vestas V150 4000",
    "Vestas V150 4200",
    "Vestas V150 4500",
    "Vestas V150 5600",
    "Vestas V150 6000",
    "Vestas V162 5600",
    "Vestas V162 6000",
    "Vestas V162 6200",
    "Vestas V162 7200",
    "Vestas V164 7000",
    "Vestas V164 8000",
    "Vestas V164 9500",
    "Vestas V172 7200",
    "Vestas V27 225",
    "Vestas V29 225",
    "Vestas V39 500",
    "Vestas V42 600",
    "Vestas V44 600",
    "Vestas V47 660",
    "Vestas V52 850",
    "Vestas V66 1650",
    "Vestas V66 1750",
    "Vestas V66 2000",
    "Vestas V80 1800",
    "Vestas V80 2000",
    "Vestas V82 1650",
    "Vestas V90 1800",
    "Vestas V90 2000",
    "Vestas V90 3000",
    "Wind World W3700",
    "Wind World W4200",
    "Windflow 33 500",
    "Windflow 500",
    "Windmaster WM28 300",
    "Windmaster WM43 750",
    "XANT M21 100",
]


class SourcePage(ctk.CTkFrame):

    def __init__(self, master, **kwargs):
        super().__init__(master, fg_color="transparent", **kwargs)
        self._tmy_data = None
        self._hourly_data = None
        self._ninja_data = None
        self._ninja_temp_data = None  # Temperature data from Ninja PV API
        
        # Form data persistence storage
        self._form_data = {
            "ninja": {},
            "pvgis_hourly": {},
            "pvgis_tmy": {},
        }
        
        # Track active data provider for strict isolation
        self._active_provider = None  # "ninja" or "pvgis" or None
        
        self._configure_styles()
        self._build()

    def _configure_styles(self):
        """Configure ttk styles for Treeview."""
        import tkinter.ttk as ttk_mod
        style = ttk_mod.Style()
        style.theme_use("clam")
        style.configure("DataTable.Treeview",
                        background="#ffffff",
                        foreground="#1e293b",
                        rowheight=26,
                        fieldbackground="#ffffff",
                        font=("Segoe UI", 9))
        style.configure("DataTable.Treeview.Heading",
                        background="#4338ca",
                        foreground="#ffffff",
                        font=("Segoe UI", 10, "bold"),
                        relief="flat")
        style.map("DataTable.Treeview.Heading",
                  background=[("active", "#3730a3")])
        style.map("DataTable.Treeview",
                  background=[("selected", "#e0e7ff")],
                  foreground=[("selected", "#1e293b")])

    def _build(self):
        # Header
        header = ctk.CTkFrame(self, fg_color="transparent")
        header.pack(fill="x", padx=Spacing.PAD_XL, pady=(Spacing.PAD_LG, 0))

        title_row = ctk.CTkFrame(header, fg_color="transparent")
        title_row.pack(fill="x")

        bar = ctk.CTkFrame(title_row, width=4, height=24,
                           fg_color=Colors.ACCENT, corner_radius=2)
        bar.pack(side="left", padx=(0, Spacing.PAD_SM))
        bar.pack_propagate(False)

        ctk.CTkLabel(title_row, text="Source", font=Fonts.H1,
                     text_color=Colors.TEXT_DARK).pack(side="left")

        badge = ctk.CTkFrame(title_row, fg_color=Colors.ACCENT_BG,
                             corner_radius=Spacing.RADIUS_SM)
        badge.pack(side="left", padx=(Spacing.PAD_SM, 0))
        ctk.CTkLabel(badge, text="  Data Input  ", font=Fonts.MICRO,
                     text_color=Colors.ACCENT).pack(padx=4, pady=2)

        ctk.CTkLabel(header, text="Configure your data source and energy parameters.",
                     font=Fonts.CAPTION, text_color=Colors.TEXT_SECONDARY
                     ).pack(anchor="w", pady=(Spacing.PAD_XS, 0))

        # Tabs
        self.tabs = ctk.CTkTabview(
            self,
            fg_color=Colors.BG_CARD,
            segmented_button_fg_color="#334155",
            segmented_button_selected_color=Colors.ACCENT,
            segmented_button_selected_hover_color=Colors.ACCENT_DIM,
            segmented_button_unselected_color="#334155",
            segmented_button_unselected_hover_color="#475569",
            text_color="#ffffff",
            text_color_disabled=Colors.TEXT_MUTED,
            corner_radius=Spacing.RADIUS_LG,
            border_width=1,
            border_color=Colors.BORDER_SUBTLE,
        )
        self.tabs.pack(fill="both", expand=True, padx=Spacing.PAD_XL, pady=Spacing.PAD_MD)

        tab1 = self.tabs.add("  Import Data  ")
        tab2 = self.tabs.add("  Energy  ")

        self._build_import_data(tab1)
        self._build_energy(tab2)

    def _build_import_data(self, parent):
        # Wrapper
        wrap = ctk.CTkFrame(parent, fg_color="transparent")
        wrap.pack(fill="both", expand=True, padx=Spacing.PAD_SM, pady=Spacing.PAD_SM)

        # 1) Source Toggle - Custom styled buttons for better aesthetics
        self._current_source = "Renewable Ninja"
        
        toggle_container = ctk.CTkFrame(wrap, fg_color=Colors.BG_CARD,
                                        corner_radius=Spacing.RADIUS_LG,
                                        border_width=1, border_color=Colors.BORDER_SUBTLE)
        toggle_container.pack(fill="x", pady=(0, Spacing.PAD_MD))
        
        toggle_inner = ctk.CTkFrame(toggle_container, fg_color="transparent")
        toggle_inner.pack(padx=Spacing.PAD_LG, pady=Spacing.PAD_MD)


        # Button container with elevated background
        btn_group = ctk.CTkFrame(toggle_inner, fg_color=Colors.BG_ELEVATED,
                                 corner_radius=Spacing.RADIUS_MD)
        btn_group.pack(side="left")
        
        btn_inner = ctk.CTkFrame(btn_group, fg_color="transparent")
        btn_inner.pack(padx=3, pady=3)

        # Renewable Ninja Button
        self._ninja_btn = ctk.CTkButton(
            btn_inner,
            text="  Renewable Ninja  ",
            command=lambda: self._switch_source("Renewable Ninja"),
            font=Fonts.BODY_BOLD,
            fg_color=Colors.SUCCESS,
            hover_color="#059669",
            text_color=Colors.TEXT_ON_ACCENT,
            corner_radius=Spacing.RADIUS_SM,
            height=32,
        )
        self._ninja_btn.pack(side="left", padx=2)

        # PVGIS Button
        self._pvgis_btn = ctk.CTkButton(
            btn_inner,
            text="  PVGIS  ",
            command=lambda: self._switch_source("PVGIS"),
            font=Fonts.BODY,
            fg_color="transparent",
            hover_color=Colors.BG_HOVER,
            text_color=Colors.TEXT_SECONDARY,
            corner_radius=Spacing.RADIUS_SM,
            height=32,
        )
        self._pvgis_btn.pack(side="left", padx=2)

        # 2) Content Area
        self._import_content = ctk.CTkFrame(wrap, fg_color=Colors.BG_CARD,
                                            corner_radius=Spacing.RADIUS_LG,
                                            border_width=1, border_color=Colors.BORDER_SUBTLE)
        self._import_content.pack(fill="both", expand=True)

        # Initial render
        self._update_import_view("Renewable Ninja")

    def _switch_source(self, source):
        """Switch between data sources and update button styles."""
        # Save current form data before switching
        self._save_current_form_data()
        
        self._current_source = source
        
        # Set active provider based on selection
        if source == "PVGIS":
            self._active_provider = "pvgis"
            self._pvgis_btn.configure(
                font=Fonts.BODY_BOLD,
                fg_color=Colors.ACCENT,
                hover_color=Colors.ACCENT_DIM,
                text_color=Colors.TEXT_ON_ACCENT
            )
            self._ninja_btn.configure(
                font=Fonts.BODY,
                fg_color="transparent",
                hover_color=Colors.BG_HOVER,
                text_color=Colors.TEXT_SECONDARY
            )
        else:
            self._active_provider = "ninja"
            self._ninja_btn.configure(
                font=Fonts.BODY_BOLD,
                fg_color=Colors.SUCCESS,
                hover_color="#059669",
                text_color=Colors.TEXT_ON_ACCENT
            )
            self._pvgis_btn.configure(
                font=Fonts.BODY,
                fg_color="transparent",
                hover_color=Colors.BG_HOVER,
                text_color=Colors.TEXT_SECONDARY
            )
        
        # Update content
        self._update_import_view(source)
        
        # Update Energy tab UI to reflect active provider
        self._update_energy_provider_status()

    def _save_current_form_data(self):
        """Save current form data before switching views."""
        # Save Ninja form data
        if hasattr(self, '_ninja_lat_entry') and self._ninja_lat_entry.winfo_exists():
            self._form_data["ninja"]["lat"] = self._ninja_lat_entry.get()
        if hasattr(self, '_ninja_lon_entry') and self._ninja_lon_entry.winfo_exists():
            self._form_data["ninja"]["lon"] = self._ninja_lon_entry.get()
        if hasattr(self, '_ninja_dataset_menu') and self._ninja_dataset_menu.winfo_exists():
            self._form_data["ninja"]["dataset"] = self._ninja_dataset_menu.get()
        if hasattr(self, '_ninja_year_menu') and self._ninja_year_menu.winfo_exists():
            self._form_data["ninja"]["year"] = self._ninja_year_menu.get()
        if hasattr(self, '_ninja_capacity_entry') and self._ninja_capacity_entry.winfo_exists():
            self._form_data["ninja"]["capacity"] = self._ninja_capacity_entry.get()
        if hasattr(self, '_ninja_height_entry') and self._ninja_height_entry.winfo_exists():
            self._form_data["ninja"]["height"] = self._ninja_height_entry.get()
        if hasattr(self, '_ninja_turbine_menu') and self._ninja_turbine_menu.winfo_exists():
            self._form_data["ninja"]["turbine"] = self._ninja_turbine_menu.get()
        if hasattr(self, '_ninja_raw_check') and self._ninja_raw_check.winfo_exists():
            self._form_data["ninja"]["raw"] = self._ninja_raw_check.get()
        
        # Save PVGIS Hourly form data
        if hasattr(self, '_hourly_lat_entry') and self._hourly_lat_entry.winfo_exists():
            self._form_data["pvgis_hourly"]["lat"] = self._hourly_lat_entry.get()
        if hasattr(self, '_hourly_lon_entry') and self._hourly_lon_entry.winfo_exists():
            self._form_data["pvgis_hourly"]["lon"] = self._hourly_lon_entry.get()
        if hasattr(self, '_hourly_database_menu') and self._hourly_database_menu.winfo_exists():
            self._form_data["pvgis_hourly"]["database"] = self._hourly_database_menu.get()
        if hasattr(self, '_hourly_startyear_menu') and self._hourly_startyear_menu.winfo_exists():
            self._form_data["pvgis_hourly"]["start_year"] = self._hourly_startyear_menu.get()
        if hasattr(self, '_hourly_endyear_menu') and self._hourly_endyear_menu.winfo_exists():
            self._form_data["pvgis_hourly"]["end_year"] = self._hourly_endyear_menu.get()
        if hasattr(self, '_mount_var'):
            self._form_data["pvgis_hourly"]["mounting"] = self._mount_var.get()
        if hasattr(self, '_hourly_slope_entry') and self._hourly_slope_entry.winfo_exists():
            self._form_data["pvgis_hourly"]["slope"] = self._hourly_slope_entry.get()
        if hasattr(self, '_hourly_azimuth_entry') and self._hourly_azimuth_entry.winfo_exists():
            self._form_data["pvgis_hourly"]["azimuth"] = self._hourly_azimuth_entry.get()
        if hasattr(self, '_hourly_peakpower_entry') and self._hourly_peakpower_entry.winfo_exists():
            self._form_data["pvgis_hourly"]["peakpower"] = self._hourly_peakpower_entry.get()
        
        # Save PVGIS TMY form data
        if hasattr(self, '_tmy_lat_entry') and self._tmy_lat_entry.winfo_exists():
            self._form_data["pvgis_tmy"]["lat"] = self._tmy_lat_entry.get()
        if hasattr(self, '_tmy_lon_entry') and self._tmy_lon_entry.winfo_exists():
            self._form_data["pvgis_tmy"]["lon"] = self._tmy_lon_entry.get()
        if hasattr(self, '_tmy_database_menu') and self._tmy_database_menu.winfo_exists():
            self._form_data["pvgis_tmy"]["database"] = self._tmy_database_menu.get()

    def _restore_ninja_form_data(self):
        """Restore saved Ninja form data."""
        data = self._form_data.get("ninja", {})
        if not data:
            return
        
        if "lat" in data and hasattr(self, '_ninja_lat_entry'):
            self._ninja_lat_entry.delete(0, "end")
            self._ninja_lat_entry.insert(0, data["lat"])
        if "lon" in data and hasattr(self, '_ninja_lon_entry'):
            self._ninja_lon_entry.delete(0, "end")
            self._ninja_lon_entry.insert(0, data["lon"])
        if "dataset" in data and hasattr(self, '_ninja_dataset_menu'):
            self._ninja_dataset_menu.set(data["dataset"])
        if "year" in data and hasattr(self, '_ninja_year_menu'):
            self._ninja_year_menu.set(data["year"])
        if "capacity" in data and hasattr(self, '_ninja_capacity_entry'):
            self._ninja_capacity_entry.delete(0, "end")
            self._ninja_capacity_entry.insert(0, data["capacity"])
        if "height" in data and hasattr(self, '_ninja_height_entry'):
            self._ninja_height_entry.delete(0, "end")
            self._ninja_height_entry.insert(0, data["height"])
        if "turbine" in data and hasattr(self, '_ninja_turbine_menu'):
            self._ninja_turbine_menu.set(data["turbine"])
        if "raw" in data and hasattr(self, '_ninja_raw_check'):
            self._ninja_raw_check.select() if data["raw"] else self._ninja_raw_check.deselect()

    def _restore_tmy_form_data(self):
        """Restore saved PVGIS TMY form data."""
        data = self._form_data.get("pvgis_tmy", {})
        if not data:
            return
        
        if "lat" in data and hasattr(self, '_tmy_lat_entry'):
            self._tmy_lat_entry.delete(0, "end")
            self._tmy_lat_entry.insert(0, data["lat"])
        if "lon" in data and hasattr(self, '_tmy_lon_entry'):
            self._tmy_lon_entry.delete(0, "end")
            self._tmy_lon_entry.insert(0, data["lon"])
        if "database" in data and hasattr(self, '_tmy_database_menu'):
            self._tmy_database_menu.set(data["database"])

    def _restore_hourly_form_data(self):
        """Restore saved PVGIS Hourly form data."""
        data = self._form_data.get("pvgis_hourly", {})
        if not data:
            return
        
        if "lat" in data and hasattr(self, '_hourly_lat_entry'):
            self._hourly_lat_entry.delete(0, "end")
            self._hourly_lat_entry.insert(0, data["lat"])
        if "lon" in data and hasattr(self, '_hourly_lon_entry'):
            self._hourly_lon_entry.delete(0, "end")
            self._hourly_lon_entry.insert(0, data["lon"])
        if "database" in data and hasattr(self, '_hourly_database_menu'):
            self._hourly_database_menu.set(data["database"])
        if "start_year" in data and hasattr(self, '_hourly_startyear_menu'):
            self._hourly_startyear_menu.set(data["start_year"])
        if "end_year" in data and hasattr(self, '_hourly_endyear_menu'):
            self._hourly_endyear_menu.set(data["end_year"])
        if "mounting" in data and hasattr(self, '_mount_var'):
            self._mount_var.set(data["mounting"])
        if "slope" in data and hasattr(self, '_hourly_slope_entry'):
            self._hourly_slope_entry.delete(0, "end")
            self._hourly_slope_entry.insert(0, data["slope"])
        if "azimuth" in data and hasattr(self, '_hourly_azimuth_entry'):
            self._hourly_azimuth_entry.delete(0, "end")
            self._hourly_azimuth_entry.insert(0, data["azimuth"])
        if "peakpower" in data and hasattr(self, '_hourly_peakpower_entry'):
            self._hourly_peakpower_entry.delete(0, "end")
            self._hourly_peakpower_entry.insert(0, data["peakpower"])

    def _update_import_view(self, choice):
        # Save current form data before clearing
        self._save_current_form_data()
        
        # Clear existing
        for child in self._import_content.winfo_children():
            child.destroy()

        # Render based on choice
        if choice == "PVGIS":
            self._render_pvgis_form()
        else:
            self._render_ninja_form()

    def _render_pvgis_form(self):
        # Distinct Header for User Verification
        header = ctk.CTkFrame(self._import_content, fg_color=Colors.BG_ELEVATED, height=50, corner_radius=0)
        header.pack(fill="x")
        header.pack_propagate(False)
        
        # Indicator line
        ctk.CTkFrame(header, width=4, height=24, fg_color=Colors.ACCENT).pack(side="left", padx=(Spacing.PAD_LG, Spacing.PAD_SM))
        
        ctk.CTkLabel(header, text="PVGIS Configuration", font=Fonts.H3,
                     text_color=Colors.TEXT_DARK).pack(side="left")
        
        ctk.CTkLabel(header, text="(Photovoltaic Geographical Information System)", font=Fonts.MICRO,
                     text_color=Colors.TEXT_MUTED).pack(side="left", padx=Spacing.PAD_SM)

        # Empty body - ready for custom content
        body = ctk.CTkFrame(self._import_content, fg_color="transparent")
        body.pack(fill="both", expand=True, padx=Spacing.PAD_XL, pady=Spacing.PAD_LG)
        
        # PVGIS Sub-tabs (Hourly / TMY) - Centered
        self._pvgis_current = "Hourly"
        
        # Tab navigation container - centered
        tab_nav_container = ctk.CTkFrame(body, fg_color="transparent")
        tab_nav_container.pack(pady=(0, Spacing.PAD_LG))
        
        # Button group with elevated background (similar to main nav)
        tab_group = ctk.CTkFrame(tab_nav_container, fg_color=Colors.BG_ELEVATED,
                                 corner_radius=Spacing.RADIUS_MD)
        tab_group.pack()
        
        tab_inner = ctk.CTkFrame(tab_group, fg_color="transparent")
        tab_inner.pack(padx=3, pady=3)
        
        # Hourly Button
        self._hourly_btn = ctk.CTkButton(
            tab_inner,
            text="  Hourly  ",
            command=lambda: self._switch_pvgis_tab("Hourly"),
            font=Fonts.BODY_BOLD,
            fg_color=Colors.ACCENT,
            hover_color=Colors.ACCENT_DIM,
            text_color=Colors.TEXT_ON_ACCENT,
            corner_radius=Spacing.RADIUS_SM,
            height=32,
        )
        self._hourly_btn.pack(side="left", padx=2)
        
        # TMY Button
        self._tmy_btn = ctk.CTkButton(
            tab_inner,
            text="  TMY  ",
            command=lambda: self._switch_pvgis_tab("TMY"),
            font=Fonts.BODY,
            fg_color="transparent",
            hover_color=Colors.BG_HOVER,
            text_color=Colors.TEXT_SECONDARY,
            corner_radius=Spacing.RADIUS_SM,
            height=32,
        )
        self._tmy_btn.pack(side="left", padx=2)
        
        # Content area for the selected tab
        self._pvgis_tab_content = ctk.CTkFrame(body, fg_color="transparent")
        self._pvgis_tab_content.pack(fill="both", expand=True)
        
        # Initial render
        self._render_pvgis_tab("Hourly")
    
    def _switch_pvgis_tab(self, tab):
        """Switch between Hourly and TMY tabs in PVGIS."""
        # Save current PVGIS form data before switching tabs
        self._save_current_form_data()
        
        self._pvgis_current = tab
        
        # Update button styles
        if tab == "Hourly":
            self._hourly_btn.configure(
                font=Fonts.BODY_BOLD,
                fg_color=Colors.ACCENT,
                hover_color=Colors.ACCENT_DIM,
                text_color=Colors.TEXT_ON_ACCENT
            )
            self._tmy_btn.configure(
                font=Fonts.BODY,
                fg_color="transparent",
                hover_color=Colors.BG_HOVER,
                text_color=Colors.TEXT_SECONDARY
            )
        else:
            self._tmy_btn.configure(
                font=Fonts.BODY_BOLD,
                fg_color=Colors.ACCENT,
                hover_color=Colors.ACCENT_DIM,
                text_color=Colors.TEXT_ON_ACCENT
            )
            self._hourly_btn.configure(
                font=Fonts.BODY,
                fg_color="transparent",
                hover_color=Colors.BG_HOVER,
                text_color=Colors.TEXT_SECONDARY
            )
        
        # Update content
        self._render_pvgis_tab(tab)
    
    def _render_pvgis_tab(self, tab):
        """Render content for the selected PVGIS tab."""
        # Clear existing content
        for child in self._pvgis_tab_content.winfo_children():
            child.destroy()
        
        if tab == "Hourly":
            self._render_hourly_form()
        else:
            self._render_tmy_form()
    
    def _render_tmy_form(self):
        """Render the TMY (Typical Meteorological Year) data input form."""
        # Scrollable form container
        scroll_frame = ctk.CTkScrollableFrame(
            self._pvgis_tab_content,
            fg_color="transparent",
            scrollbar_button_color=Colors.BORDER_DEFAULT,
            scrollbar_button_hover_color=Colors.BORDER_SUBTLE
        )
        scroll_frame.pack(fill="both", expand=True, padx=Spacing.PAD_MD)
        
        # Main form container inside scrollable frame
        form = ctk.CTkFrame(scroll_frame, fg_color="transparent")
        form.pack(fill="both", expand=True)
        
        # Latitude and Longitude (side by side)
        coords_row = ctk.CTkFrame(form, fg_color="transparent")
        coords_row.pack(fill="x", pady=(0, Spacing.PAD_MD))
        coords_row.grid_columnconfigure((0, 1), weight=1)
        
        lat_f = ctk.CTkFrame(coords_row, fg_color="transparent")
        lat_f.grid(row=0, column=0, sticky="ew", padx=(0, Spacing.PAD_SM))
        ctk.CTkLabel(lat_f, text="Latitude*", font=Fonts.CAPTION,
                     text_color=Colors.TEXT_SECONDARY).pack(anchor="w")
        self._tmy_lat_entry = ctk.CTkEntry(lat_f, placeholder_text="e.g. 45.123", font=Fonts.BODY, height=38,
                     fg_color=Colors.BG_ELEVATED, border_color=Colors.BORDER_SUBTLE,
                     text_color=Colors.TEXT_PRIMARY)
        self._tmy_lat_entry.pack(fill="x", pady=(4, 0))
        
        lon_f = ctk.CTkFrame(coords_row, fg_color="transparent")
        lon_f.grid(row=0, column=1, sticky="ew", padx=(Spacing.PAD_SM, 0))
        ctk.CTkLabel(lon_f, text="Longitude*", font=Fonts.CAPTION,
                     text_color=Colors.TEXT_SECONDARY).pack(anchor="w")
        self._tmy_lon_entry = ctk.CTkEntry(lon_f, placeholder_text="e.g. -0.456", font=Fonts.BODY, height=38,
                     fg_color=Colors.BG_ELEVATED, border_color=Colors.BORDER_SUBTLE,
                     text_color=Colors.TEXT_PRIMARY)
        self._tmy_lon_entry.pack(fill="x", pady=(4, 0))
        
        # Select database
        period_frame = ctk.CTkFrame(form, fg_color="transparent")
        period_frame.pack(fill="x", pady=(0, Spacing.PAD_MD))
        ctk.CTkLabel(period_frame, text="Solar Radiation Database*", font=Fonts.CAPTION,
                     text_color=Colors.TEXT_SECONDARY).pack(anchor="w")
        self._tmy_database_menu = ctk.CTkOptionMenu(
            period_frame,
            values=[
                "PVGIS-SARAH3: 2005 - 2023",
                "PVGIS-ERA5: 2005 - 2023"
            ],
            font=Fonts.BODY, height=38,
            fg_color=Colors.BG_ELEVATED,
            button_color=Colors.BORDER_DEFAULT,
            dropdown_fg_color=Colors.BG_CARD,
            text_color=Colors.TEXT_PRIMARY
        )
        self._tmy_database_menu.pack(fill="x", pady=(4, 0))
        self._tmy_database_menu.set("PVGIS-SARAH3: 2005 - 2023")  # Set default
        
        # Fetch button
        # Status label for API feedback
        self._tmy_status_label = ctk.CTkLabel(
            form, text="", font=Fonts.CAPTION,
            text_color=Colors.TEXT_SECONDARY
        )
        self._tmy_status_label.pack(pady=(0, Spacing.PAD_SM))
        
        # Button row
        btn_row = ctk.CTkFrame(form, fg_color="transparent")
        btn_row.pack(fill="x", pady=Spacing.PAD_SM)
        
        ctk.CTkButton(
            btn_row, text="Fetch TMY Data",
            command=self._fetch_tmy_data,
            font=Fonts.BODY_BOLD, height=42,
            fg_color=Colors.ACCENT, hover_color=Colors.ACCENT_DIM,
            text_color=Colors.TEXT_ON_ACCENT,
            corner_radius=Spacing.RADIUS_MD,
            width=180
        ).pack(side="left", padx=(0, Spacing.PAD_SM))
        
        self._tmy_viz_btn = ctk.CTkButton(
            btn_row, text="Visualization & Export",
            command=lambda: self._show_data_popup("tmy"),
            font=Fonts.BODY_BOLD, height=42,
            fg_color="#6366f1", hover_color="#4f46e5",
            text_color="#ffffff",
            corner_radius=Spacing.RADIUS_MD,
            width=200, state="disabled"
        )
        self._tmy_viz_btn.pack(side="left")
        
        # Restore saved TMY form data
        self._restore_tmy_form_data()
        
        # Enable viz button if data exists
        if self._tmy_data is not None:
            self._tmy_viz_btn.configure(state="normal")

    
    def _render_hourly_form(self):
        """Render the hourly data input form."""
        # Scrollable form container
        scroll_frame = ctk.CTkScrollableFrame(
            self._pvgis_tab_content,
            fg_color="transparent",
            scrollbar_button_color=Colors.BORDER_DEFAULT,
            scrollbar_button_hover_color=Colors.BORDER_SUBTLE
        )
        scroll_frame.pack(fill="both", expand=True, padx=Spacing.PAD_MD)
        
        # Main form container inside scrollable frame
        form = ctk.CTkFrame(scroll_frame, fg_color="transparent")
        form.pack(fill="both", expand=True)

        
        # Latitude and Longitude (side by side)
        coords_row = ctk.CTkFrame(form, fg_color="transparent")
        coords_row.pack(fill="x", pady=(0, Spacing.PAD_MD))
        coords_row.grid_columnconfigure((0, 1), weight=1)
        
        lat_f = ctk.CTkFrame(coords_row, fg_color="transparent")
        lat_f.grid(row=0, column=0, sticky="ew", padx=(0, Spacing.PAD_SM))
        ctk.CTkLabel(lat_f, text="Latitude*", font=Fonts.CAPTION,
                     text_color=Colors.TEXT_SECONDARY).pack(anchor="w")
        self._hourly_lat_entry = ctk.CTkEntry(lat_f, placeholder_text="e.g. 45.123", font=Fonts.BODY, height=38,
                     fg_color=Colors.BG_ELEVATED, border_color=Colors.BORDER_SUBTLE,
                     text_color=Colors.TEXT_PRIMARY)
        self._hourly_lat_entry.pack(fill="x", pady=(4, 0))
        # Bind to auto-fetch altitude when value changes
        self._hourly_lat_entry.bind("<FocusOut>", lambda e: self._on_coordinates_changed())
        self._hourly_lat_entry.bind("<Return>", lambda e: self._on_coordinates_changed())
        
        lon_f = ctk.CTkFrame(coords_row, fg_color="transparent")
        lon_f.grid(row=0, column=1, sticky="ew", padx=(Spacing.PAD_SM, 0))
        ctk.CTkLabel(lon_f, text="Longitude*", font=Fonts.CAPTION,
                     text_color=Colors.TEXT_SECONDARY).pack(anchor="w")
        self._hourly_lon_entry = ctk.CTkEntry(lon_f, placeholder_text="e.g. -0.456", font=Fonts.BODY, height=38,
                     fg_color=Colors.BG_ELEVATED, border_color=Colors.BORDER_SUBTLE,
                     text_color=Colors.TEXT_PRIMARY)
        self._hourly_lon_entry.pack(fill="x", pady=(4, 0))
        # Bind to auto-fetch altitude when value changes
        self._hourly_lon_entry.bind("<FocusOut>", lambda e: self._on_coordinates_changed())
        self._hourly_lon_entry.bind("<Return>", lambda e: self._on_coordinates_changed())
        
        # Solar radiation database
        db_frame = ctk.CTkFrame(form, fg_color="transparent")
        db_frame.pack(fill="x", pady=(0, Spacing.PAD_MD))
        ctk.CTkLabel(db_frame, text="Solar radiation database*", font=Fonts.CAPTION,
                     text_color=Colors.TEXT_SECONDARY).pack(anchor="w")
        self._hourly_database_menu = ctk.CTkOptionMenu(db_frame, values=["PVGIS-SARAH3", "PVGIS-ERA5"],
                          font=Fonts.BODY, height=38, fg_color=Colors.BG_ELEVATED,
                          button_color=Colors.BORDER_DEFAULT, dropdown_fg_color=Colors.BG_CARD,
                          text_color=Colors.TEXT_PRIMARY)
        self._hourly_database_menu.pack(fill="x", pady=(4, 0))
        self._hourly_database_menu.set("PVGIS-SARAH3")
        
        # Start and End year (side by side) - 2005 to 2023 for both databases
        year_row = ctk.CTkFrame(form, fg_color="transparent")
        year_row.pack(fill="x", pady=(0, Spacing.PAD_MD))
        year_row.grid_columnconfigure((0, 1), weight=1)
        
        # Generate year lists (2005-2023)
        years_list = [str(year) for year in range(2005, 2024)]
        
        start_f = ctk.CTkFrame(year_row, fg_color="transparent")
        start_f.grid(row=0, column=0, sticky="ew", padx=(0, Spacing.PAD_SM))
        ctk.CTkLabel(start_f, text="Start year*", font=Fonts.CAPTION,
                     text_color=Colors.TEXT_SECONDARY).pack(anchor="w")
        self._hourly_startyear_menu = ctk.CTkOptionMenu(start_f, values=years_list,
                          font=Fonts.BODY, height=38, fg_color=Colors.BG_ELEVATED,
                          button_color=Colors.BORDER_DEFAULT, dropdown_fg_color=Colors.BG_CARD,
                          text_color=Colors.TEXT_PRIMARY)
        self._hourly_startyear_menu.pack(fill="x", pady=(4, 0))
        self._hourly_startyear_menu.set("2020")
        
        end_f = ctk.CTkFrame(year_row, fg_color="transparent")
        end_f.grid(row=0, column=1, sticky="ew", padx=(Spacing.PAD_SM, 0))
        ctk.CTkLabel(end_f, text="End year*", font=Fonts.CAPTION,
                     text_color=Colors.TEXT_SECONDARY).pack(anchor="w")
        self._hourly_endyear_menu = ctk.CTkOptionMenu(end_f, values=years_list,
                          font=Fonts.BODY, height=38, fg_color=Colors.BG_ELEVATED,
                          button_color=Colors.BORDER_DEFAULT, dropdown_fg_color=Colors.BG_CARD,
                          text_color=Colors.TEXT_PRIMARY)
        self._hourly_endyear_menu.pack(fill="x", pady=(4, 0))
        self._hourly_endyear_menu.set("2023")
        
        # Mounting type
        mount_frame = ctk.CTkFrame(form, fg_color="transparent")
        mount_frame.pack(fill="x", pady=(0, Spacing.PAD_MD))
        ctk.CTkLabel(mount_frame, text="Mounting type*", font=Fonts.CAPTION,
                     text_color=Colors.TEXT_SECONDARY).pack(anchor="w", pady=(0, 4))
        
        # Radio buttons for mounting type
        self._mount_var = ctk.StringVar(value="Fixed")
        radio_row = ctk.CTkFrame(mount_frame, fg_color="transparent")
        radio_row.pack(fill="x")
        
        for option in ["Fixed", "Vertical axis", "Inclined axis", "Two axis"]:
            ctk.CTkRadioButton(radio_row, text=option, variable=self._mount_var, value=option,
                               font=Fonts.BODY, text_color=Colors.TEXT_PRIMARY,
                               fg_color=Colors.ACCENT, hover_color=Colors.ACCENT_DIM,
                               command=self._on_mounting_change).pack(side="left", padx=(0, Spacing.PAD_LG))
        
        # Slope and Azimuth (side by side)
        slope_az_row = ctk.CTkFrame(form, fg_color="transparent")
        slope_az_row.pack(fill="x", pady=(0, Spacing.PAD_MD))
        slope_az_row.grid_columnconfigure((0, 1), weight=1)
        
        slope_f = ctk.CTkFrame(slope_az_row, fg_color="transparent")
        slope_f.grid(row=0, column=0, sticky="ew", padx=(0, Spacing.PAD_SM))
        ctk.CTkLabel(slope_f, text="Slope [°]", font=Fonts.CAPTION,
                     text_color=Colors.TEXT_SECONDARY).pack(anchor="w")
        self._hourly_slope_entry = ctk.CTkEntry(slope_f, placeholder_text="(0-90)", font=Fonts.BODY, height=38,
                     fg_color=Colors.BG_ELEVATED, border_color=Colors.BORDER_SUBTLE,
                     text_color=Colors.TEXT_PRIMARY)
        self._hourly_slope_entry.pack(fill="x", pady=(4, 0))
        
        az_f = ctk.CTkFrame(slope_az_row, fg_color="transparent")
        az_f.grid(row=0, column=1, sticky="ew", padx=(Spacing.PAD_SM, 0))
        ctk.CTkLabel(az_f, text="Azimuth [°]", font=Fonts.CAPTION,
                     text_color=Colors.TEXT_SECONDARY).pack(anchor="w")
        self._hourly_azimuth_entry = ctk.CTkEntry(az_f, placeholder_text="(-180-180)", font=Fonts.BODY, height=38,
                     fg_color=Colors.BG_ELEVATED, border_color=Colors.BORDER_SUBTLE,
                     text_color=Colors.TEXT_PRIMARY)
        self._hourly_azimuth_entry.pack(fill="x", pady=(4, 0))
        
        # Optimize checkboxes
        opt_row = ctk.CTkFrame(form, fg_color="transparent")
        opt_row.pack(fill="x", pady=(0, Spacing.PAD_MD))
        
        self._hourly_opt_slope_check = ctk.CTkCheckBox(opt_row, text="Optimize slope", font=Fonts.BODY,
                        text_color=Colors.TEXT_PRIMARY, fg_color=Colors.ACCENT,
                        hover_color=Colors.ACCENT_DIM, command=self._on_optimize_change)
        self._hourly_opt_slope_check.pack(side="left", padx=(0, Spacing.PAD_LG))
        
        self._hourly_opt_both_check = ctk.CTkCheckBox(opt_row, text="Optimize slope and azimuth", font=Fonts.BODY,
                        text_color=Colors.TEXT_PRIMARY, fg_color=Colors.ACCENT,
                        hover_color=Colors.ACCENT_DIM, command=self._on_optimize_change)
        self._hourly_opt_both_check.pack(side="left")
        
        # PV power checkbox and fields
        self._hourly_pv_check = ctk.CTkCheckBox(form, text="PV power", font=Fonts.BODY_BOLD,
                                    text_color=Colors.TEXT_PRIMARY, fg_color=Colors.ACCENT,
                                    hover_color=Colors.ACCENT_DIM, command=self._on_pv_check_change)
        self._hourly_pv_check.pack(anchor="w", pady=(0, Spacing.PAD_SM))
        self._hourly_pv_check.select()  # Selected by default
        
        # PV Technology
        tech_frame = ctk.CTkFrame(form, fg_color="transparent")
        tech_frame.pack(fill="x", pady=(0, Spacing.PAD_SM))
        ctk.CTkLabel(tech_frame, text="PV technology", font=Fonts.CAPTION,
                     text_color=Colors.TEXT_SECONDARY).pack(anchor="w")
        self._hourly_pvtech_menu = ctk.CTkOptionMenu(tech_frame, values=["Crystalline silicon (crystSi)", "CIS", "CdTe", "Unknown"],
                          font=Fonts.BODY, height=38, fg_color=Colors.BG_ELEVATED,
                          button_color=Colors.BORDER_DEFAULT, dropdown_fg_color=Colors.BG_CARD,
                          text_color=Colors.TEXT_PRIMARY)
        self._hourly_pvtech_menu.pack(fill="x", pady=(4, 0))
        self._hourly_pvtech_menu.set("Crystalline silicon (crystSi)")
        
        # Installed peak PV power and System loss (side by side)
        pv_row = ctk.CTkFrame(form, fg_color="transparent")
        pv_row.pack(fill="x", pady=(0, Spacing.PAD_MD))
        pv_row.grid_columnconfigure((0, 1), weight=1)
        
        peak_f = ctk.CTkFrame(pv_row, fg_color="transparent")
        peak_f.grid(row=0, column=0, sticky="ew", padx=(0, Spacing.PAD_SM))
        ctk.CTkLabel(peak_f, text="Installed peak PV power [kWp]", font=Fonts.CAPTION,
                     text_color=Colors.TEXT_SECONDARY).pack(anchor="w")
        self._hourly_peakpower_entry = ctk.CTkEntry(peak_f, placeholder_text="1", font=Fonts.BODY, height=38,
                     fg_color=Colors.BG_ELEVATED, border_color=Colors.BORDER_SUBTLE,
                     text_color=Colors.TEXT_PRIMARY)
        self._hourly_peakpower_entry.pack(fill="x", pady=(4, 0))
        self._hourly_peakpower_entry.insert(0, "1")
        
        loss_f = ctk.CTkFrame(pv_row, fg_color="transparent")
        loss_f.grid(row=0, column=1, sticky="ew", padx=(Spacing.PAD_SM, 0))
        ctk.CTkLabel(loss_f, text="System loss [%]", font=Fonts.CAPTION,
                     text_color=Colors.TEXT_SECONDARY).pack(anchor="w")
        self._hourly_loss_entry = ctk.CTkEntry(loss_f, placeholder_text="14", font=Fonts.BODY, height=38,
                     fg_color=Colors.BG_ELEVATED, border_color=Colors.BORDER_SUBTLE,
                     text_color=Colors.TEXT_PRIMARY)
        self._hourly_loss_entry.pack(fill="x", pady=(4, 0))
        self._hourly_loss_entry.insert(0, "14")
        
        # Radiation components checkbox
        self._hourly_components_check = ctk.CTkCheckBox(form, text="Radiation components", font=Fonts.BODY_BOLD,
                        text_color=Colors.TEXT_PRIMARY, fg_color=Colors.ACCENT,
                        hover_color=Colors.ACCENT_DIM)
        self._hourly_components_check.pack(anchor="w", pady=(0, Spacing.PAD_LG))
        
        # Fetch button
        # Status label
        self._hourly_status_label = ctk.CTkLabel(
            form, text="", font=Fonts.CAPTION,
            text_color=Colors.TEXT_SECONDARY
        )
        self._hourly_status_label.pack(pady=(0, Spacing.PAD_SM))
        
        # Button row
        btn_row = ctk.CTkFrame(form, fg_color="transparent")
        btn_row.pack(fill="x", pady=Spacing.PAD_SM)
        
        ctk.CTkButton(btn_row, text="Fetch Hourly Data",
                      command=self._fetch_hourly_data,
                      font=Fonts.BODY_BOLD, height=42,
                      fg_color=Colors.ACCENT, hover_color=Colors.ACCENT_DIM,
                      text_color=Colors.TEXT_ON_ACCENT, corner_radius=Spacing.RADIUS_MD,
                      width=180).pack(side="left", padx=(0, Spacing.PAD_SM))
        
        self._hourly_viz_btn = ctk.CTkButton(btn_row, text="Visualization & Export",
                      command=lambda: self._show_data_popup("hourly"),
                      font=Fonts.BODY_BOLD, height=42,
                      fg_color="#6366f1", hover_color="#4f46e5",
                      text_color="#ffffff", corner_radius=Spacing.RADIUS_MD,
                      width=200, state="disabled")
        self._hourly_viz_btn.pack(side="left")
        
        # Initialize input states based on default mounting type (Fixed)
        self._on_mounting_change()
        
        # Restore saved Hourly form data
        self._restore_hourly_form_data()
        
        # Enable viz button if data exists
        if self._hourly_data is not None:
            self._hourly_viz_btn.configure(state="normal")
        
    def _render_ninja_form(self):
        # Distinct Header for User Verification
        header = ctk.CTkFrame(self._import_content, fg_color=Colors.BG_ELEVATED, height=50, corner_radius=0)
        header.pack(fill="x")
        header.pack_propagate(False)

        # Indicator line (Different color logic if we had distinct colors, but using ACCENT for consistency)
        ctk.CTkFrame(header, width=4, height=24, fg_color=Colors.SUCCESS).pack(side="left", padx=(Spacing.PAD_LG, Spacing.PAD_SM))

        ctk.CTkLabel(header, text="Renewable Ninja Configuration", font=Fonts.H3,
                     text_color=Colors.TEXT_DARK).pack(side="left")

        ctk.CTkLabel(header, text="(Global Solar & Wind Data)", font=Fonts.MICRO,
                     text_color=Colors.TEXT_MUTED).pack(side="left", padx=Spacing.PAD_SM)

        # Scrollable form container
        scroll_frame = ctk.CTkScrollableFrame(
            self._import_content,
            fg_color="transparent",
            scrollbar_button_color=Colors.BORDER_DEFAULT,
            scrollbar_button_hover_color=Colors.BORDER_SUBTLE
        )
        scroll_frame.pack(fill="both", expand=True, padx=Spacing.PAD_XL, pady=Spacing.PAD_LG)
        
        # Main form container
        form = ctk.CTkFrame(scroll_frame, fg_color="transparent")
        form.pack(fill="both", expand=True)
        
        # Latitude and Longitude (side by side)
        latlon_row = ctk.CTkFrame(form, fg_color="transparent")
        latlon_row.pack(fill="x", pady=(0, Spacing.PAD_MD))
        latlon_row.grid_columnconfigure((0, 1), weight=1)
        
        lat_f = ctk.CTkFrame(latlon_row, fg_color="transparent")
        lat_f.grid(row=0, column=0, sticky="ew", padx=(0, Spacing.PAD_SM))
        ctk.CTkLabel(lat_f, text="Latitude", font=Fonts.CAPTION,
                     text_color=Colors.TEXT_SECONDARY).pack(anchor="w")
        self._ninja_lat_entry = ctk.CTkEntry(lat_f, placeholder_text="e.g., 56.0", font=Fonts.BODY, height=38,
                     fg_color=Colors.BG_ELEVATED, border_color=Colors.BORDER_SUBTLE,
                     text_color=Colors.TEXT_PRIMARY)
        self._ninja_lat_entry.pack(fill="x", pady=(4, 0))
        # Bind to auto-fetch altitude when value changes
        self._ninja_lat_entry.bind("<FocusOut>", lambda e: self._on_coordinates_changed())
        self._ninja_lat_entry.bind("<Return>", lambda e: self._on_coordinates_changed())
        
        lon_f = ctk.CTkFrame(latlon_row, fg_color="transparent")
        lon_f.grid(row=0, column=1, sticky="ew", padx=(Spacing.PAD_SM, 0))
        ctk.CTkLabel(lon_f, text="Longitude", font=Fonts.CAPTION,
                     text_color=Colors.TEXT_SECONDARY).pack(anchor="w")
        self._ninja_lon_entry = ctk.CTkEntry(lon_f, placeholder_text="e.g., -3.0", font=Fonts.BODY, height=38,
                     fg_color=Colors.BG_ELEVATED, border_color=Colors.BORDER_SUBTLE,
                     text_color=Colors.TEXT_PRIMARY)
        self._ninja_lon_entry.pack(fill="x", pady=(4, 0))
        # Bind to auto-fetch altitude when value changes
        self._ninja_lon_entry.bind("<FocusOut>", lambda e: self._on_coordinates_changed())
        self._ninja_lon_entry.bind("<Return>", lambda e: self._on_coordinates_changed())
        
        # Dataset (MERRA-2 only)
        dataset_frame = ctk.CTkFrame(form, fg_color="transparent")
        dataset_frame.pack(fill="x", pady=(0, Spacing.PAD_MD))
        ctk.CTkLabel(dataset_frame, text="Dataset", font=Fonts.CAPTION,
                     text_color=Colors.TEXT_SECONDARY).pack(anchor="w")
        self._ninja_dataset_menu = ctk.CTkOptionMenu(
            dataset_frame,
            values=["MERRA-2 (global)"],
            font=Fonts.BODY, height=38,
            fg_color=Colors.BG_ELEVATED,
            button_color=Colors.BORDER_DEFAULT,
            dropdown_fg_color=Colors.BG_CARD,
            text_color=Colors.TEXT_PRIMARY
        )
        self._ninja_dataset_menu.pack(fill="x", pady=(4, 0))
        self._ninja_dataset_menu.set("MERRA-2 (global)")
        
        # Select a year of data
        year_frame = ctk.CTkFrame(form, fg_color="transparent")
        year_frame.pack(fill="x", pady=(0, Spacing.PAD_MD))
        ctk.CTkLabel(year_frame, text="Select a year of data", font=Fonts.CAPTION,
                     text_color=Colors.TEXT_SECONDARY).pack(anchor="w")
        years = [str(y) for y in range(2024, 1979, -1)]  # 2024 down to 1980 (MERRA-2 dataset range)
        self._ninja_year_menu = ctk.CTkOptionMenu(
            year_frame,
            values=years,
            font=Fonts.BODY, height=38,
            fg_color=Colors.BG_ELEVATED,
            button_color=Colors.BORDER_DEFAULT,
            dropdown_fg_color=Colors.BG_CARD,
            text_color=Colors.TEXT_PRIMARY
        )
        self._ninja_year_menu.pack(fill="x", pady=(4, 0))
        self._ninja_year_menu.set("2024")
        
        # Capacity (kW)
        capacity_frame = ctk.CTkFrame(form, fg_color="transparent")
        capacity_frame.pack(fill="x", pady=(0, Spacing.PAD_MD))
        ctk.CTkLabel(capacity_frame, text="Capacity (kW)", font=Fonts.CAPTION,
                     text_color=Colors.TEXT_SECONDARY).pack(anchor="w")
        self._ninja_capacity_entry = ctk.CTkEntry(capacity_frame, placeholder_text="1", font=Fonts.BODY, height=38,
                     fg_color=Colors.BG_ELEVATED, border_color=Colors.BORDER_SUBTLE,
                     text_color=Colors.TEXT_PRIMARY)
        self._ninja_capacity_entry.pack(fill="x", pady=(4, 0))
        self._ninja_capacity_entry.insert(0, "1")
        
        # Hub height (m)
        hub_frame = ctk.CTkFrame(form, fg_color="transparent")
        hub_frame.pack(fill="x", pady=(0, Spacing.PAD_MD))
        ctk.CTkLabel(hub_frame, text="Hub height (m)", font=Fonts.CAPTION,
                     text_color=Colors.TEXT_SECONDARY).pack(anchor="w")
        self._ninja_height_entry = ctk.CTkEntry(hub_frame, placeholder_text="80", font=Fonts.BODY, height=38,
                     fg_color=Colors.BG_ELEVATED, border_color=Colors.BORDER_SUBTLE,
                     text_color=Colors.TEXT_PRIMARY)
        self._ninja_height_entry.pack(fill="x", pady=(4, 0))
        self._ninja_height_entry.insert(0, "80")
        
        # Turbine model
        turbine_frame = ctk.CTkFrame(form, fg_color="transparent")
        turbine_frame.pack(fill="x", pady=(0, Spacing.PAD_MD))
        ctk.CTkLabel(turbine_frame, text="Turbine model", font=Fonts.CAPTION,
                     text_color=Colors.TEXT_SECONDARY).pack(anchor="w")
        self._ninja_turbine_menu = ctk.CTkOptionMenu(
            turbine_frame,
            values=TURBINE_MODELS,
            font=Fonts.BODY, height=38,
            fg_color=Colors.BG_ELEVATED,
            button_color=Colors.BORDER_DEFAULT,
            dropdown_fg_color=Colors.BG_CARD,
            text_color=Colors.TEXT_PRIMARY
        )
        self._ninja_turbine_menu.pack(fill="x", pady=(4, 0))
        self._ninja_turbine_menu.set("Vestas V80 2000")
        
        # Add callback to sync with Energy page
        self._ninja_turbine_menu.configure(command=self._on_ninja_turbine_changed)
        
        # Raw data checkbox
        self._ninja_raw_check = ctk.CTkCheckBox(
            form,
            text="Include raw data",
            font=Fonts.BODY,
            text_color=Colors.TEXT_PRIMARY,
            fg_color=Colors.SUCCESS,
            hover_color="#059669"
        )
        self._ninja_raw_check.pack(anchor="w", pady=(0, Spacing.PAD_LG))
        
        # Restore saved form data for Ninja
        self._restore_ninja_form_data()
        
        # Status label for API feedback
        self._ninja_status_label = ctk.CTkLabel(
            form, text="", font=Fonts.CAPTION,
            text_color=Colors.TEXT_SECONDARY
        )
        self._ninja_status_label.pack(pady=(0, Spacing.PAD_SM))
        
        # Button row
        btn_row = ctk.CTkFrame(form, fg_color="transparent")
        btn_row.pack(fill="x")
        
        ctk.CTkButton(
            btn_row,
            text="Fetch Ninja Data",
            command=self._fetch_ninja_data,
            font=Fonts.BODY_BOLD, height=42,
            fg_color=Colors.SUCCESS,
            hover_color="#059669",
            text_color=Colors.TEXT_ON_ACCENT,
            corner_radius=Spacing.RADIUS_MD,
            width=180
        ).pack(side="left", padx=(0, Spacing.PAD_SM))
        
        self._ninja_viz_btn = ctk.CTkButton(
            btn_row,
            text="Visualization & Export",
            command=lambda: self._show_data_popup("ninja"),
            font=Fonts.BODY_BOLD, height=42,
            fg_color="#6366f1", hover_color="#4f46e5",
            text_color="#ffffff",
            corner_radius=Spacing.RADIUS_MD,
            width=200, state="disabled"
        )
        self._ninja_viz_btn.pack(side="left")
        
        # Enable viz button if data exists
        if self._ninja_data is not None:
            self._ninja_viz_btn.configure(state="normal")

    def _update_energy_provider_status(self):
        """Update Energy tab to show only the active provider's status."""
        try:
            if not hasattr(self, '_energy_provider_label'):
                return
                
            if self._active_provider == "ninja":
                self._energy_provider_label.configure(
                    text="● Renewable Ninja (Active Provider)",
                    text_color=Colors.SUCCESS
                )
                # Hide PVGIS status, show Ninja status
                self._energy_pvgis_status.configure(
                    text="⏸ PVGIS: Not active (using Renewable Ninja)",
                    text_color=Colors.TEXT_MUTED
                )
                self._energy_ninja_status.configure(
                    text="⟳ Renewable Ninja: Ready to fetch data" if not self._ninja_data 
                         else "✓ Renewable Ninja: Data loaded",
                    text_color=Colors.TEXT_SECONDARY if not self._ninja_data else "#10b981"
                )
                # Update source labels
                if hasattr(self, '_energy_temp_source_label'):
                    self._energy_temp_source_label.configure(text="from Ninja")
                if hasattr(self, '_energy_wind_source_label'):
                    self._energy_wind_source_label.configure(text="from Ninja")
                    
            elif self._active_provider == "pvgis":
                self._energy_provider_label.configure(
                    text="● PVGIS (Active Provider)",
                    text_color=Colors.ACCENT
                )
                # Show PVGIS status, hide Ninja status  
                self._energy_pvgis_status.configure(
                    text="⟳ PVGIS: Ready to fetch data" if not self._hourly_data
                         else "✓ PVGIS: Data loaded",
                    text_color=Colors.TEXT_SECONDARY if not self._hourly_data else "#10b981"
                )
                self._energy_ninja_status.configure(
                    text="⏸ Renewable Ninja: Not active (using PVGIS)",
                    text_color=Colors.TEXT_MUTED
                )
                # Update source labels
                if hasattr(self, '_energy_temp_source_label'):
                    self._energy_temp_source_label.configure(text="from PVGIS")
                if hasattr(self, '_energy_wind_source_label'):
                    self._energy_wind_source_label.configure(text="from PVGIS")
            else:
                self._energy_provider_label.configure(
                    text="No data provider selected - Choose Renewable Ninja or PVGIS in Import Data tab",
                    text_color=Colors.TEXT_MUTED
                )
                self._energy_pvgis_status.configure(
                    text="⚠ PVGIS: Not active",
                    text_color=Colors.TEXT_MUTED
                )
                self._energy_ninja_status.configure(
                    text="⚠ Renewable Ninja: Not active", 
                    text_color=Colors.TEXT_MUTED
                )
        except Exception as e:
            print(f"Error updating energy provider status: {e}")

    def _on_ninja_turbine_changed(self, selected_turbine):
        """Callback when turbine selection changes in Ninja tab - sync with Energy page."""
        try:
            if hasattr(self, '_energy_turbine_menu') and self._energy_turbine_menu.winfo_exists():
                self._energy_turbine_menu.set(selected_turbine)
        except Exception:
            pass  # Silently fail if widget doesn't exist

    def _sync_energy_turbine_with_ninja(self):
        """Sync Energy page turbine selection with Ninja tab's current selection."""
        try:
            if hasattr(self, '_ninja_turbine_menu') and self._ninja_turbine_menu.winfo_exists():
                if hasattr(self, '_energy_turbine_menu') and self._energy_turbine_menu.winfo_exists():
                    ninja_turbine = self._ninja_turbine_menu.get()
                    current_values = self._energy_turbine_menu.cget("values")
                    if ninja_turbine in current_values:
                        self._energy_turbine_menu.set(ninja_turbine)
                        # Update specs display
                        self._update_turbine_specs_display()
        except Exception:
            pass  # Silently fail if widgets don't exist yet
    
    def _on_turbine_changed(self, selected_turbine):
        """Handle turbine selection change - update specs display and trigger recalculation."""
        self._update_turbine_specs_display()
        # Also sync with Ninja tab if this is the Energy page turbine
        self._on_ninja_turbine_changed(selected_turbine)
    
    def _on_coordinates_changed(self):
        """Handle coordinate changes - auto-fetch altitude with debouncing."""
        # Cancel any pending altitude fetch
        if hasattr(self, '_altitude_fetch_after_id'):
            try:
                self.after_cancel(self._altitude_fetch_after_id)
            except Exception:
                pass
        
        # Schedule new fetch after delay (debounce)
        self._altitude_fetch_after_id = self.after(800, self._fetch_altitude_if_valid)
    
    def _fetch_altitude_if_valid(self):
        """Fetch altitude only if coordinates are valid."""
        lat, lon, source = self._get_coordinates_for_altitude()
        
        if lat is None or lon is None:
            return
        
        # Validate coordinate ranges
        if not (-90 <= lat <= 90) or not (-180 <= lon <= 180):
            return
        
        print(f"Auto-fetching altitude for coordinates: ({lat}, {lon})")
        
        # Update UI to show fetching
        if hasattr(self, '_energy_alt_status'):
            self._energy_alt_status.configure(text="⟳ auto-fetching...", text_color=Colors.ACCENT)
        
        # Fetch in background thread
        def fetch():
            elevation = None
            source_name = None
            
            # Try PVGIS first
            try:
                url = "https://re.jrc.ec.europa.eu/api/v5_3/datarequest"
                params = {"lat": lat, "lon": lon, "outputformat": "json"}
                response = requests.get(url, params=params, timeout=15)
                response.raise_for_status()
                data = response.json()
                
                if "meta" in data and "elevation" in data["meta"]:
                    elevation = data["meta"]["elevation"]
                elif "inputs" in data and "location" in data["inputs"]:
                    elevation = data["inputs"]["location"].get("elevation")
                
                if elevation is not None:
                    source_name = "PVGIS"
            except Exception:
                pass
            
            # Try Open-Elevation as fallback
            if elevation is None:
                try:
                    elevation = self._fetch_from_open_elevation(lat, lon)
                    if elevation is not None:
                        source_name = "Open-Elevation"
                except Exception:
                    pass
            
            # Update UI
            if elevation is not None and hasattr(self, '_energy_alt_label'):
                self.after(0, lambda: self._update_altitude_auto(elevation, source_name))
            elif hasattr(self, '_energy_alt_status'):
                self.after(0, lambda: self._energy_alt_status.configure(
                    text="⚠ auto-fetch failed", text_color="#f59e0b"))
        
        thread = threading.Thread(target=fetch)
        thread.daemon = True
        thread.start()
    
    def _update_altitude_auto(self, elevation, source_name):
        """Update altitude field after auto-fetch."""
        if hasattr(self, '_energy_alt_label'):
            self._energy_alt_label.configure(text=str(int(round(elevation))))
        self._current_altitude = elevation  # Store for calculations
        if hasattr(self, '_energy_alt_status'):
            self._energy_alt_status.configure(
                text=f"✓ auto-fetched from {source_name}", 
                text_color="#10b981"
            )
        if hasattr(self, '_energy_alt_fetch_btn') and self._energy_alt_fetch_btn.winfo_exists():
            self._energy_alt_fetch_btn.configure(state="normal")
        if hasattr(self, '_update_air_density'):
            self._update_air_density()
        print(f"Auto-updated altitude: {elevation}m from {source_name}")
    
    def _update_turbine_specs_display(self):
        """Update Vd, Vn, Pn display based on selected turbine (static display only)."""
        try:
            if hasattr(self, '_energy_turbine_menu'):
                turbine_name = self._energy_turbine_menu.get()
                specs = self._get_turbine_specs(turbine_name)
                
                # Store current specs for calculations
                self._current_turbine_specs = specs
                
                # Update Vd label (static display)
                if hasattr(self, '_energy_vd_label'):
                    self._energy_vd_label.configure(text=str(specs["cut_in_speed"]))
                
                # Update Vn label (static display)
                if hasattr(self, '_energy_vn_label'):
                    self._energy_vn_label.configure(text=str(specs["rated_speed"]))
                
                # Update Pn label (static display)
                if hasattr(self, '_energy_pn_label'):
                    self._energy_pn_label.configure(text=str(specs["rated_power_kw"]))
                
                # Update Voff label (static display)
                if hasattr(self, '_energy_voff_label'):
                    voff = specs.get("cut_out_speed", 25.0)
                    self._energy_voff_label.configure(text=str(voff))
                
                print(f"Updated specs for {turbine_name}: Vd={specs['cut_in_speed']}, Vn={specs['rated_speed']}, Voff={specs.get('cut_out_speed', 25.0)}, Pn={specs['rated_power_kw']}")
        except Exception as e:
            print(f"Error updating turbine specs display: {e}")

    def _build_energy(self, parent):
        """Build Energy tab with wind power calculation."""
        import customtkinter as ctk
        
        # Main scrollable container
        scroll_frame = ctk.CTkScrollableFrame(parent, fg_color="transparent")
        scroll_frame.pack(fill="both", expand=True, padx=Spacing.PAD_SM, pady=Spacing.PAD_SM)
        
        # Data Source Status Section
        source_frame = ctk.CTkFrame(scroll_frame, fg_color=Colors.BG_CARD,
                                    corner_radius=Spacing.RADIUS_LG,
                                    border_width=1, border_color=Colors.BORDER_SUBTLE)
        source_frame.pack(fill="x", pady=(0, Spacing.PAD_MD))
        
        source_inner = ctk.CTkFrame(source_frame, fg_color="transparent")
        source_inner.pack(fill="x", padx=Spacing.PAD_LG, pady=Spacing.PAD_MD)
        
        ctk.CTkLabel(source_inner, text="Data Sources", font=Fonts.H3,
                     text_color=Colors.TEXT_DARK).pack(anchor="w")
        
        # Active provider indicator
        self._energy_provider_label = ctk.CTkLabel(source_inner, text="No data provider selected",
                                                    font=Fonts.BODY_BOLD, text_color=Colors.TEXT_MUTED)
        self._energy_provider_label.pack(anchor="w", pady=(Spacing.PAD_XS, 0))
        
        # Status indicators (conditionally shown based on active provider)
        self._energy_pvgis_status = ctk.CTkLabel(source_inner, text="⚠ PVGIS: Not active",
                                                  font=Fonts.BODY, text_color=Colors.TEXT_MUTED)
        self._energy_pvgis_status.pack(anchor="w", pady=(Spacing.PAD_XS, 0))
        
        self._energy_ninja_status = ctk.CTkLabel(source_inner, text="⚠ Renewable Ninja: Not active",
                                                  font=Fonts.BODY, text_color=Colors.TEXT_MUTED)
        self._energy_ninja_status.pack(anchor="w", pady=(Spacing.PAD_XS, 0))
        
        # Active fetch status
        self._energy_fetch_status = ctk.CTkLabel(source_inner, text="",
                                                  font=Fonts.CAPTION, text_color=Colors.ACCENT)
        self._energy_fetch_status.pack(anchor="w", pady=(Spacing.PAD_XS, 0))
        
        # Input Parameters Section
        input_frame = ctk.CTkFrame(scroll_frame, fg_color=Colors.BG_CARD,
                                   corner_radius=Spacing.RADIUS_LG,
                                   border_width=1, border_color=Colors.BORDER_SUBTLE)
        input_frame.pack(fill="x", pady=(0, Spacing.PAD_MD))
        
        input_inner = ctk.CTkFrame(input_frame, fg_color="transparent")
        input_inner.pack(fill="x", padx=Spacing.PAD_LG, pady=Spacing.PAD_MD)
        
        ctk.CTkLabel(input_inner, text="Input Parameters", font=Fonts.H3,
                     text_color=Colors.TEXT_DARK).pack(anchor="w", pady=(0, Spacing.PAD_SM))
        
        # Input grid - organized in rows for consistent alignment
        input_grid = ctk.CTkFrame(input_inner, fg_color="transparent")
        input_grid.pack(fill="x")
        input_grid.grid_columnconfigure((0, 1, 2, 3), weight=1)
        
        # ROW 1: Parameter Labels
        ctk.CTkLabel(input_grid, text="Temperature T₂ₘ (°C)", font=Fonts.CAPTION,
                     text_color=Colors.TEXT_SECONDARY).grid(row=0, column=0, sticky="w", padx=Spacing.PAD_XS)
        ctk.CTkLabel(input_grid, text="Altitude (m)", font=Fonts.CAPTION,
                     text_color=Colors.TEXT_SECONDARY).grid(row=0, column=1, sticky="w", padx=Spacing.PAD_XS)
        ctk.CTkLabel(input_grid, text="Wind Speed V (m/s)", font=Fonts.CAPTION,
                     text_color=Colors.TEXT_SECONDARY).grid(row=0, column=2, sticky="w", padx=Spacing.PAD_XS)
        ctk.CTkLabel(input_grid, text="Turbine Model", font=Fonts.CAPTION,
                     text_color=Colors.TEXT_SECONDARY).grid(row=0, column=3, sticky="w", padx=Spacing.PAD_XS)
        
        # ROW 2: Source indicators
        self._energy_temp_source_label = ctk.CTkLabel(input_grid, text="", font=Fonts.CAPTION,
                                                       text_color=Colors.TEXT_MUTED)
        self._energy_temp_source_label.grid(row=1, column=0, sticky="w", padx=Spacing.PAD_XS)
        
        self._energy_alt_status = ctk.CTkLabel(input_grid, text="⚠ manual", font=Fonts.CAPTION,
                                                text_color="#f59e0b")
        self._energy_alt_status.grid(row=1, column=1, sticky="w", padx=Spacing.PAD_XS)
        
        self._energy_wind_source_label = ctk.CTkLabel(input_grid, text="", font=Fonts.CAPTION,
                                                       text_color=Colors.TEXT_MUTED)
        self._energy_wind_source_label.grid(row=1, column=2, sticky="w", padx=Spacing.PAD_XS)
        
        # Empty cell for turbine model (no source indicator needed)
        ctk.CTkLabel(input_grid, text="", font=Fonts.CAPTION).grid(row=1, column=3, sticky="w")
        
        # ROW 3: Values (display only)
        self._energy_temp_label = ctk.CTkLabel(input_grid, text="--", font=Fonts.BODY,
                                               text_color=Colors.TEXT_PRIMARY)
        self._energy_temp_label.grid(row=2, column=0, sticky="w", padx=Spacing.PAD_XS, pady=(4, 0))
        
        self._energy_alt_label = ctk.CTkLabel(input_grid, text="0", font=Fonts.BODY,
                                              text_color=Colors.TEXT_PRIMARY)
        self._energy_alt_label.grid(row=2, column=1, sticky="w", padx=Spacing.PAD_XS, pady=(4, 0))
        
        self._energy_wind_label = ctk.CTkLabel(input_grid, text="--", font=Fonts.BODY,
                                               text_color=Colors.TEXT_PRIMARY)
        self._energy_wind_label.grid(row=2, column=2, sticky="w", padx=Spacing.PAD_XS, pady=(4, 0))
        
        self._energy_turbine_menu = ctk.CTkOptionMenu(
            input_grid,
            values=TURBINE_MODELS,
            font=Fonts.BODY,
            fg_color=Colors.BG_ELEVATED,
            button_color=Colors.BORDER_DEFAULT,
            dropdown_fg_color=Colors.BG_CARD,
            text_color=Colors.TEXT_PRIMARY,
            corner_radius=Spacing.RADIUS_MD,
            height=38,
        )
        self._energy_turbine_menu.grid(row=2, column=3, sticky="ew", padx=Spacing.PAD_XS, pady=(4, 0))
        
        # ROW 4: Action button (only for altitude)
        ctk.CTkLabel(input_grid, text="", font=Fonts.CAPTION).grid(row=3, column=0, sticky="w")
        
        self._energy_alt_fetch_btn = ctk.CTkButton(
            input_grid, text="Auto-fetch altitude", font=Fonts.CAPTION, height=24,
            fg_color=Colors.BG_ELEVATED, hover_color=Colors.BG_HOVER,
            text_color=Colors.TEXT_SECONDARY, corner_radius=Spacing.RADIUS_SM,
            command=self._fetch_altitude_auto
        )
        self._energy_alt_fetch_btn.grid(row=3, column=1, sticky="ew", padx=Spacing.PAD_XS, pady=(8, 0))
        
        # Sync with Ninja tab turbine selection
        self._sync_energy_turbine_with_ninja()
        
        # Hourly data availability info
        self._energy_hourly_info = ctk.CTkLabel(input_inner, text="",
                                                 font=Fonts.CAPTION, text_color=Colors.TEXT_MUTED)
        self._energy_hourly_info.pack(anchor="w", pady=(Spacing.PAD_SM, 0))
        
        # Air Density Display
        rho_frame = ctk.CTkFrame(scroll_frame, fg_color=Colors.ACCENT_BG,
                                 corner_radius=Spacing.RADIUS_MD)
        rho_frame.pack(fill="x", pady=(0, Spacing.PAD_MD))
        
        rho_inner = ctk.CTkFrame(rho_frame, fg_color="transparent")
        rho_inner.pack(fill="x", padx=Spacing.PAD_MD, pady=Spacing.PAD_MD)
        
        ctk.CTkLabel(rho_inner, text="Calculated Air Density (ρ)", font=Fonts.CAPTION,
                     text_color=Colors.ACCENT).pack(anchor="w")
        self._energy_rho_label = ctk.CTkLabel(rho_inner, text="1.225 kg/m³ (sea level, 15°C)",
                                              font=Fonts.H3, text_color=Colors.TEXT_DARK)
        self._energy_rho_label.pack(anchor="w", pady=(Spacing.PAD_XS, 0))
        
        # Update air density when inputs change
        # Air density updates automatically when data is fetched
        # (No manual binding needed since fields are now display-only)
        
        # Model Selection Section
        model_frame = ctk.CTkFrame(scroll_frame, fg_color=Colors.BG_CARD,
                                   corner_radius=Spacing.RADIUS_LG,
                                   border_width=1, border_color=Colors.BORDER_SUBTLE)
        model_frame.pack(fill="x", pady=(0, Spacing.PAD_MD))
        
        model_inner = ctk.CTkFrame(model_frame, fg_color="transparent")
        model_inner.pack(fill="x", padx=Spacing.PAD_LG, pady=Spacing.PAD_MD)
        
        ctk.CTkLabel(model_inner, text="Calculation Model", font=Fonts.H3,
                     text_color=Colors.TEXT_DARK).pack(anchor="w")
        
        # Model dropdown
        self._energy_model_var = ctk.StringVar(value="Model 1")
        self._energy_model_menu = ctk.CTkOptionMenu(
            model_inner,
            values=["Model 1", "Model 2"],
            variable=self._energy_model_var,
            command=self._update_energy_display,
            font=Fonts.BODY,
            fg_color=Colors.BG_ELEVATED,
            button_color=Colors.BORDER_DEFAULT,
            dropdown_fg_color=Colors.BG_CARD,
            text_color=Colors.TEXT_PRIMARY,
            corner_radius=Spacing.RADIUS_MD,
            height=38,
        )
        self._energy_model_menu.pack(fill="x", pady=(Spacing.PAD_SM, 0))
        
        # Dynamic display area for model details
        self._energy_display_frame = ctk.CTkFrame(scroll_frame, fg_color=Colors.BG_CARD,
                                                  corner_radius=Spacing.RADIUS_LG,
                                                  border_width=1, border_color=Colors.BORDER_SUBTLE)
        self._energy_display_frame.pack(fill="x", expand=False)
        
        # Calculate Button
        calc_frame = ctk.CTkFrame(scroll_frame, fg_color="transparent")
        calc_frame.pack(fill="x", pady=Spacing.PAD_MD)
        
        ctk.CTkButton(
            calc_frame,
            text="  Calculate  ",
            font=Fonts.BODY_BOLD,
            fg_color=Colors.ACCENT,
            hover_color=Colors.ACCENT_DIM,
            text_color=Colors.TEXT_ON_ACCENT,
            corner_radius=Spacing.RADIUS_MD,
            height=48,
            command=self._show_energy_results
        ).pack(anchor="center")
        
        # Auto-fill from fetched data after all widgets are created
        self._auto_fill_energy_data()
        
        # Start periodic status check
        self._schedule_fetch_status_check()
    
    def _update_energy_display(self, selected_model):
        """Update the display area based on selected model."""
        import customtkinter as ctk
        
        # Clear previous content
        for widget in self._energy_display_frame.winfo_children():
            widget.destroy()
        
        # Header
        header = ctk.CTkFrame(self._energy_display_frame, fg_color=Colors.BG_ELEVATED,
                              height=42, corner_radius=0)
        header.pack(fill="x")
        header.pack_propagate(False)
        ctk.CTkLabel(header, text=f"  {selected_model} Configuration", font=Fonts.H3,
                     text_color=Colors.TEXT_DARK).pack(side="left", padx=Spacing.PAD_LG)
        
        ctk.CTkFrame(self._energy_display_frame, height=1, fg_color=Colors.BORDER_SUBTLE).pack(fill="x")
        
        # Content area
        content = ctk.CTkFrame(self._energy_display_frame, fg_color="transparent")
        content.pack(fill="both", expand=True, padx=Spacing.PAD_XL, pady=Spacing.PAD_MD)
        
        if selected_model == "Model 1":
            # Model 1: P = ½ * ρ * A * V³
            self._build_model1_display(content)
        else:
            # Model 2: P = Pm * (V³ - Vc³) / (Vm³ - Vc³)
            self._build_model2_display(content)
    
    def _build_model1_display(self, parent):
        """Build display for Model 1: Wind Power Equation."""
        import customtkinter as ctk
        
        # Equation section - clean format like the photo
        eq_card = ctk.CTkFrame(parent, fg_color=Colors.ACCENT_BG,
                               corner_radius=Spacing.RADIUS_MD)
        eq_card.pack(fill="x", pady=(0, Spacing.PAD_MD))
        
        ctk.CTkLabel(eq_card, text="Equation", font=Fonts.CAPTION,
                     text_color=Colors.ACCENT).pack(anchor="w", padx=Spacing.PAD_MD, pady=(Spacing.PAD_MD, 0))
        
        # LaTeX-style equation
        ctk.CTkLabel(eq_card, text="P = ½ · ρ · A · Cₚ · V³", font=Fonts.H3,
                     text_color=Colors.TEXT_DARK).pack(anchor="w", padx=Spacing.PAD_MD, pady=(Spacing.PAD_XS, Spacing.PAD_MD))
        
        # Description
        ctk.CTkLabel(parent, text="Standard wind power equation calculating theoretical power from wind kinetic energy.",
                     font=Fonts.BODY, text_color=Colors.TEXT_SECONDARY,
                     wraplength=600).pack(anchor="w", pady=(0, Spacing.PAD_MD))
    
    def _build_model2_display(self, parent):
        """Build display for Model 2: Cubic Power Curve Model."""
        import customtkinter as ctk
        
        # Equation section
        eq_card = ctk.CTkFrame(parent, fg_color=Colors.ACCENT_BG,
                               corner_radius=Spacing.RADIUS_MD)
        eq_card.pack(fill="x", pady=(0, Spacing.PAD_MD))
        
        ctk.CTkLabel(eq_card, text="Equation", font=Fonts.CAPTION,
                     text_color=Colors.ACCENT).pack(anchor="w", padx=Spacing.PAD_MD, pady=(Spacing.PAD_MD, 0))
        
        # LaTeX-style fraction with proper alignment
        eq_container = ctk.CTkFrame(eq_card, fg_color="transparent")
        eq_container.pack(anchor="w", padx=Spacing.PAD_MD, pady=(Spacing.PAD_XS, Spacing.PAD_SM))
        
        # Use grid for precise alignment
        eq_container.grid_columnconfigure(0, weight=0)  # Left part (P = Pn)
        eq_container.grid_columnconfigure(1, weight=0)  # Center part (dot)
        eq_container.grid_columnconfigure(2, weight=0)  # Fraction
        
        # Row 0: Left side aligned with fraction center
        left_label = ctk.CTkLabel(eq_container, text="P = Pₙ", font=Fonts.H3,
                                   text_color=Colors.TEXT_DARK)
        left_label.grid(row=0, column=0, sticky="e", padx=(0, 4))
        
        # Multiplication dot
        dot_label = ctk.CTkLabel(eq_container, text="·", font=Fonts.H3,
                                  text_color=Colors.TEXT_DARK)
        dot_label.grid(row=0, column=1, sticky="w", padx=(0, 8))
        
        # Fraction in its own frame
        frac_container = ctk.CTkFrame(eq_container, fg_color="transparent")
        frac_container.grid(row=0, column=2, sticky="w")
        
        # Numerator (centered)
        num_label = ctk.CTkLabel(frac_container, text="V³ - Vd³", font=Fonts.H3,
                                  text_color=Colors.TEXT_DARK)
        num_label.pack(anchor="center")
        
        # Fraction bar (full width)
        bar = ctk.CTkFrame(frac_container, height=2, fg_color=Colors.TEXT_DARK)
        bar.pack(fill="x", expand=True, padx=2)
        
        # Denominator (centered)
        den_label = ctk.CTkLabel(frac_container, text="Vn³ - Vd³", font=Fonts.H3,
                                  text_color=Colors.TEXT_DARK)
        den_label.pack(anchor="center")
        
        # Description
        ctk.CTkLabel(parent, text="Cubic power curve model for wind speeds between cut-in and rated speed.",
                     font=Fonts.BODY, text_color=Colors.TEXT_SECONDARY,
                     wraplength=600).pack(anchor="w", pady=(0, Spacing.PAD_MD))
        
        # Turbine Specs section with editable Vd and Vn
        specs_frame = ctk.CTkFrame(parent, fg_color=Colors.BG_CARD,
                                   corner_radius=Spacing.RADIUS_MD,
                                   border_width=1, border_color=Colors.BORDER_SUBTLE)
        specs_frame.pack(fill="x", pady=(0, Spacing.PAD_MD))
        
        specs_inner = ctk.CTkFrame(specs_frame, fg_color="transparent")
        specs_inner.pack(fill="x", padx=Spacing.PAD_MD, pady=Spacing.PAD_MD)
        
        ctk.CTkLabel(specs_inner, text="Turbine Specifications (from database)", font=Fonts.H3,
                     text_color=Colors.TEXT_DARK).pack(anchor="w", pady=(0, Spacing.PAD_SM))
        
        # Input fields for Vd, Vn, Voff, Pn
        input_grid = ctk.CTkFrame(specs_inner, fg_color="transparent")
        input_grid.pack(fill="x")
        input_grid.grid_columnconfigure((0, 1, 2, 3), weight=1)
        
        # Cut-in speed (Vd) - display only (static)
        vd_frame = ctk.CTkFrame(input_grid, fg_color="transparent")
        vd_frame.grid(row=0, column=0, sticky="ew", padx=(0, Spacing.PAD_XS))
        ctk.CTkLabel(vd_frame, text="Cut-in Speed Vd (m/s)", font=Fonts.CAPTION,
                     text_color=Colors.TEXT_SECONDARY).pack(anchor="w")
        self._energy_vd_label = ctk.CTkLabel(vd_frame, text="3.0", font=Fonts.BODY,
                                              text_color=Colors.TEXT_PRIMARY)
        self._energy_vd_label.pack(anchor="w", pady=(Spacing.PAD_XS + 8, 0))
        
        # Rated speed (Vn) - display only (static)
        vn_frame = ctk.CTkFrame(input_grid, fg_color="transparent")
        vn_frame.grid(row=0, column=1, sticky="ew", padx=Spacing.PAD_XS)
        ctk.CTkLabel(vn_frame, text="Rated Speed Vn (m/s)", font=Fonts.CAPTION,
                     text_color=Colors.TEXT_SECONDARY).pack(anchor="w")
        self._energy_vn_label = ctk.CTkLabel(vn_frame, text="12.0", font=Fonts.BODY,
                                              text_color=Colors.TEXT_PRIMARY)
        self._energy_vn_label.pack(anchor="w", pady=(Spacing.PAD_XS + 8, 0))
        
        # Cut-out speed (Voff) - display only (static)
        voff_frame = ctk.CTkFrame(input_grid, fg_color="transparent")
        voff_frame.grid(row=0, column=2, sticky="ew", padx=Spacing.PAD_XS)
        ctk.CTkLabel(voff_frame, text="Cut-out Speed Voff (m/s)", font=Fonts.CAPTION,
                     text_color=Colors.TEXT_SECONDARY).pack(anchor="w")
        self._energy_voff_label = ctk.CTkLabel(voff_frame, text="25.0", font=Fonts.BODY,
                                               text_color=Colors.TEXT_PRIMARY)
        self._energy_voff_label.pack(anchor="w", pady=(Spacing.PAD_XS + 8, 0))
        
        # Rated power (Pn) - display only (static)
        pn_frame = ctk.CTkFrame(input_grid, fg_color="transparent")
        pn_frame.grid(row=0, column=3, sticky="ew", padx=(Spacing.PAD_XS, 0))
        ctk.CTkLabel(pn_frame, text="Rated Power Pn (kW)", font=Fonts.CAPTION,
                     text_color=Colors.TEXT_SECONDARY).pack(anchor="w")
        self._energy_pn_label = ctk.CTkLabel(pn_frame, text="2000", font=Fonts.BODY,
                                              text_color=Colors.TEXT_PRIMARY)
        self._energy_pn_label.pack(anchor="w", pady=(Spacing.PAD_XS + 8, 0))
        
        # Update specs when turbine changes
        self._energy_turbine_menu.configure(command=self._on_turbine_changed)
        
        # Initialize with current turbine
        self._update_turbine_specs_display()

    def _auto_fill_energy_data_ninja(self):
        """Auto-fill energy inputs from Renewable Ninja data (wind + temp from solar API)."""
        if not self._ninja_data or not self._ninja_temp_data:
            return
        
        try:
            # Process wind data
            wind_data_entries = self._ninja_data.get("data", {})
            wind_count = len(wind_data_entries) if isinstance(wind_data_entries, dict) else 0
            
            # Process temperature data from solar API
            solar_data_entries = self._ninja_temp_data.get("data", {})
            temp_count = len(solar_data_entries) if isinstance(solar_data_entries, dict) else 0
            
            # Extract and populate wind speed
            if isinstance(wind_data_entries, dict) and wind_data_entries:
                wind_speeds = [v.get("wind_speed", 0) for v in wind_data_entries.values() 
                              if isinstance(v, dict) and "wind_speed" in v]
                if wind_speeds:
                    avg_wind = sum(wind_speeds) / len(wind_speeds)
                    self._energy_wind_label.configure(text=f"{avg_wind:.2f}")
                    self._current_wind = avg_wind  # Store for calculations
                    print(f"Energy: Average wind speed {avg_wind:.2f} m/s from Ninja")
            
            # Extract and populate temperature from solar data
            temps = []
            if isinstance(solar_data_entries, dict):
                temps = [v.get("temperature") for v in solar_data_entries.values() 
                        if isinstance(v, dict) and v.get("temperature") is not None]
            elif isinstance(solar_data_entries, list):
                temps = [r.get("temperature") for r in solar_data_entries 
                        if isinstance(r, dict) and r.get("temperature") is not None]
            
            if temps:
                avg_temp = sum(temps) / len(temps)
                self._energy_temp_label.configure(text=f"{avg_temp:.1f}")
                self._current_temp = avg_temp  # Store for calculations
                print(f"Energy: Average temperature {avg_temp:.1f}°C from Ninja Solar API")
            
            # Update status indicators
            self._energy_ninja_status.configure(
                text=f"✓ Renewable Ninja: Wind ({wind_count}h) + Temp ({temp_count}h) synchronized",
                text_color="#10b981"
            )
            
            # Hide PVGIS status when Ninja is active
            self._energy_pvgis_status.configure(
                text="⏸ PVGIS: Not active (using Renewable Ninja)",
                text_color=Colors.TEXT_MUTED
            )
            
            # Update hourly info label
            if hasattr(self, '_energy_hourly_info'):
                self._energy_hourly_info.configure(
                    text=f"✓ Renewable Ninja data ready: {wind_count} hourly records (Wind + Temperature synchronized)",
                    text_color="#10b981"
                )
            
            # Update fetch status
            self._energy_fetch_status.configure(
                text="✓ Wind and Temperature data synchronized from Renewable Ninja. Ready for calculation.",
                text_color="#10b981"
            )
            
            # Update air density display
            self._update_air_density()
            
        except Exception as e:
            print(f"Error in _auto_fill_energy_data_ninja: {e}")
    
    def _auto_fill_energy_data(self):
        """Auto-fill energy inputs from fetched data based on active provider."""
        # STRICT ISOLATION: Only process data from the active provider
        if self._active_provider == "ninja":
            # Only use Ninja data
            if self._ninja_data and self._ninja_temp_data:
                self._auto_fill_energy_data_ninja()
            else:
                # Waiting for Ninja data
                self._energy_fetch_status.configure(
                    text="⟳ Fetching data from Renewable Ninja...",
                    text_color=Colors.ACCENT
                )
        elif self._active_provider == "pvgis":
            # Only use PVGIS data
            self._auto_fill_energy_data_pvgis()
        else:
            # No active provider selected yet - show neutral status
            self._energy_fetch_status.configure(
                text="Select a data provider (Renewable Ninja or PVGIS) in the Import Data tab",
                text_color=Colors.TEXT_MUTED
            )
    
    def _auto_fill_energy_data_pvgis(self):
        """Auto-fill energy inputs from PVGIS data only."""
        pvgis_records = 0
        pvgis_hourly_data = None
        pvgis_data_source = None
        
        # Use Hourly data (priority)
        if hasattr(self, '_hourly_data') and self._hourly_data:
            try:
                outputs = self._hourly_data.get("outputs", {})
                hourly = outputs.get("hourly", [])
                if hourly and len(hourly) > 0:
                    pvgis_hourly_data = hourly
                    pvgis_records = len(hourly)
                    pvgis_data_source = self._hourly_data
                    print(f"Energy: Found {pvgis_records} PVGIS Hourly records")
            except Exception as e:
                print(f"Energy: Error parsing Hourly data: {e}")
        # Fallback to TMY data
        elif hasattr(self, '_tmy_data') and self._tmy_data:
            try:
                outputs = self._tmy_data.get("outputs", {})
                hourly = outputs.get("hourly", [])
                if hourly and len(hourly) > 0:
                    pvgis_hourly_data = hourly
                    pvgis_records = len(hourly)
                    pvgis_data_source = self._tmy_data
                    print(f"Energy: Found {pvgis_records} PVGIS TMY records")
            except Exception as e:
                print(f"Energy: Error parsing TMY data: {e}")
        
        # Process PVGIS data if available
        if pvgis_hourly_data:
            try:
                temps = [h.get("T2m", 15.0) for h in pvgis_hourly_data if "T2m" in h]
                if temps:
                    avg_temp = sum(temps) / len(temps)
                    self._energy_temp_label.configure(text=f"{avg_temp:.1f}")
                    self._current_temp = avg_temp  # Store for calculations
                    print(f"Energy: Average temperature {avg_temp:.1f}°C from PVGIS")
                
                # Update status
                self._energy_pvgis_status.configure(
                    text=f"✓ PVGIS: Temperature data loaded ({pvgis_records} records)",
                    text_color="#10b981"
                )
                
                # Get altitude from PVGIS metadata
                altitude = None
                if pvgis_data_source:
                    meta = pvgis_data_source.get("meta", {})
                    altitude = meta.get("elevation")
                    if altitude is None:
                        inputs = pvgis_data_source.get("inputs", {})
                        location = inputs.get("location", {})
                        altitude = location.get("elevation")
                    
                    if altitude is not None:
                        self._energy_alt_label.configure(text=str(int(round(altitude))))
                        self._current_altitude = altitude  # Store for calculations
                        if hasattr(self, '_energy_alt_status'):
                            self._energy_alt_status.configure(text="✓ from PVGIS", text_color="#10b981")
                
                # Hide Ninja status when PVGIS is active
                self._energy_ninja_status.configure(
                    text="⏸ Renewable Ninja: Not active (using PVGIS)",
                    text_color=Colors.TEXT_MUTED
                )
                
                # Update hourly info
                if hasattr(self, '_energy_hourly_info'):
                    self._energy_hourly_info.configure(
                        text=f"✓ PVGIS data ready: {pvgis_records} hourly records",
                        text_color="#10b981"
                    )
                
                self._energy_fetch_status.configure(
                    text="✓ PVGIS data loaded. Ready for calculation.",
                    text_color="#10b981"
                )
                
            except Exception as e:
                print(f"Energy: Error processing PVGIS data: {e}")
        else:
            # Waiting for PVGIS data
            self._energy_pvgis_status.configure(
                text="⟳ Waiting for PVGIS data...",
                text_color=Colors.TEXT_SECONDARY
            )
        
        # Update air density
        self._update_air_density()
    
    def _schedule_fetch_status_check(self):
        """Schedule periodic checks for fetch status updates."""
        self._check_fetch_status()
        # Check again in 500ms
        self.after(500, self._schedule_fetch_status_check)
    
    def _check_fetch_status(self):
        """Check and display fetch status based on active provider."""
        try:
            # STRICT ISOLATION: Only check status for the active provider
            if self._active_provider == "ninja":
                try:
                    ninja_text = self._ninja_status_label.cget("text")
                    ninja_fetching = "Fetching" in ninja_text or "Loading" in ninja_text
                    
                    if ninja_fetching:
                        self._energy_fetch_status.configure(
                            text="⟳ Renewable Ninja is fetching wind and temperature data...",
                            text_color=Colors.ACCENT
                        )
                    elif self._ninja_data and self._ninja_temp_data:
                        self._energy_fetch_status.configure(
                            text="✓ Renewable Ninja data ready! Click Calculate to compute hourly wind power.",
                            text_color="#10b981"
                        )
                    else:
                        self._energy_fetch_status.configure(
                            text="⏳ Click 'Fetch Ninja Data' to load wind and temperature data",
                            text_color=Colors.TEXT_MUTED
                        )
                except Exception:
                    pass
                    
            elif self._active_provider == "pvgis":
                try:
                    pvgis_text = self._hourly_status_label.cget("text")
                    pvgis_fetching = "Fetching" in pvgis_text or "Please wait" in pvgis_text
                    
                    if pvgis_fetching:
                        self._energy_fetch_status.configure(
                            text="⟳ PVGIS is fetching data...",
                            text_color=Colors.ACCENT
                        )
                    elif self._hourly_data or self._tmy_data:
                        self._energy_fetch_status.configure(
                            text="✓ PVGIS data ready! Click Calculate to compute hourly wind power.",
                            text_color="#10b981"
                        )
                    else:
                        self._energy_fetch_status.configure(
                            text="⏳ Fetch PVGIS data in the Import Data tab",
                            text_color=Colors.TEXT_MUTED
                        )
                except Exception:
                    pass
            else:
                # No provider selected
                self._energy_fetch_status.configure(
                    text="Select a data provider (Renewable Ninja or PVGIS) in the Import Data tab",
                    text_color=Colors.TEXT_MUTED
                )
        except Exception:
            pass
    
    def _update_air_density(self):
        """Calculate and update air density based on temperature and altitude."""
        try:
            # Get values from stored data (fields are display-only)
            temp_c = float(getattr(self, '_current_temp', 15.0))
            altitude = float(getattr(self, '_current_altitude', 0))
        except ValueError:
            return
        
        # Calculate air density using barometric formula
        # ρ = ρ₀ * (P/P₀) * (T₀/T)
        # Simplified: ρ = ρ₀ * exp(-M*g*h/(R*T)) * (T₀/T)
        # Where ρ₀ = 1.225 kg/m³ at sea level, 15°C
        
        rho_0 = 1.225  # kg/m³ at sea level
        T_0 = 288.15   # K (15°C)
        g = 9.80665    # m/s²
        M = 0.0289644  # kg/mol (molar mass of air)
        R = 8.31447    # J/(mol·K)
        
        T_kelvin = temp_c + 273.15
        
        # Pressure at altitude
        P_alt = 101325 * ((T_kelvin) / T_0) ** (-g * M / (R * 0.0065))
        
        # Density at altitude and temperature
        rho = (P_alt * M) / (R * T_kelvin)
        
        self._energy_rho_label.configure(
            text=f"{rho:.3f} kg/m³ (at {altitude}m, {temp_c:.1f}°C)"
        )
    
    def _get_turbine_specs(self, turbine_name):
        """Get turbine specifications from database or parse from name."""
        # First try to get from database
        if turbine_name in TURBINE_DATABASE:
            specs = TURBINE_DATABASE[turbine_name].copy()
            specs["name"] = turbine_name
            return specs
        
        # Fallback: parse from name and use defaults
        specs = TURBINE_DATABASE["_default"].copy()
        specs["name"] = turbine_name
        
        try:
            import re
            power_match = re.search(r'(\d+)\s*$', turbine_name)
            if power_match:
                specs["rated_power_kw"] = int(power_match.group(1))
            
            diameter_match = re.search(r'V(\d+)|SWT[\-\.](\d+\.?\d*)|G(\d+)|E(\d+)', turbine_name)
            if diameter_match:
                for group in diameter_match.groups():
                    if group:
                        specs["rotor_diameter_m"] = int(float(group))
                        break
        except Exception:
            pass
        
        return specs
    
    def _calculate_model1(self, rho, V, turbine_specs):
        """Calculate power using Model 1: P = ½ × ρ × A × Cp × V³."""
        # Swept area
        r = turbine_specs["rotor_diameter_m"] / 2
        A = 3.14159 * r ** 2
        
        # Power coefficient (typical value 0.35-0.45, max 0.59 Betz limit)
        Cp = 0.40
        
        # Power in Watts
        P = 0.5 * rho * A * Cp * (V ** 3)
        
        # Cap at rated power
        rated_power_w = turbine_specs["rated_power_kw"] * 1000
        P = min(P, rated_power_w)
        
        return P, Cp, A
    
    def _calculate_model2(self, V, turbine_specs):
        """Calculate power using Model 2 piecewise power curve.
        
        P(v) = {
            0                   if v < Vd (cut-in)
            Pn × (v³-Vd³)/(Vn³-Vd³)  if Vd ≤ v ≤ Vn (cubic region)
            Pn                  if Vn < v ≤ Voff (rated region)
            0                   if v > Voff (cut-out/shutdown)
        }
        """
        Vd = turbine_specs["cut_in_speed"]      # Cut-in speed (Vn in LaTeX)
        Vn = turbine_specs["rated_speed"]       # Rated speed (Vr in LaTeX)
        Voff = turbine_specs.get("cut_out_speed", 25.0)  # Cut-out/shutdown speed
        Pn = turbine_specs["rated_power_kw"] * 1000  # Nominal power in Watts
        
        if V < Vd or V > Voff:
            # Region 1: Below cut-in or above cut-out - no power
            P = 0
        elif Vd <= V <= Vn:
            # Region 2: Cubic power curve between cut-in and rated speed
            P = Pn * ((V ** 3) - (Vd ** 3)) / ((Vn ** 3) - (Vd ** 3))
        else:  # Vn < V <= Voff
            # Region 3: Rated power between rated speed and cut-out
            P = Pn
        
        return P
    
    def _show_energy_results(self):
        """Show calculation results in a popup window with hourly data."""
        import customtkinter as ctk
        import tkinter as tk
        from tkinter import ttk
        import tkinter.messagebox as mb
        from datetime import datetime
        
        # Get selected model
        selected_model = self._energy_model_var.get()
        
        # Get turbine specs (from database - static values)
        turbine_name = self._energy_turbine_menu.get()
        turbine_specs = self._get_turbine_specs(turbine_name)
        
        # Get altitude
        try:
            altitude = float(getattr(self, '_current_altitude', 0))
        except ValueError:
            altitude = 0
        
        # Determine data source based on active provider
        wind_data = {}
        temp_data = {}
        data_source_name = ""
        
        if self._active_provider == "ninja" and self._ninja_data and self._ninja_temp_data:
            # DUAL-NINJA MODE: Both wind and temperature from Renewable Ninja
            data_source_name = "Renewable Ninja"
            
            # Process wind data from Ninja
            raw_wind = self._ninja_data.get("data", {})
            if isinstance(raw_wind, dict):
                for ts, values in raw_wind.items():
                    if isinstance(values, dict):
                        try:
                            ts_ms = int(ts)
                            from datetime import datetime, timezone
                            dt = datetime.fromtimestamp(ts_ms / 1000.0, tz=timezone.utc)
                            time_key = dt.strftime("%Y%m%d:%H%M")
                            wind_data[time_key] = values
                        except Exception:
                            pass
            
            # Process temperature data from Ninja Solar API
            raw_solar = self._ninja_temp_data.get("data", {})
            if isinstance(raw_solar, dict):
                for ts, values in raw_solar.items():
                    if isinstance(values, dict):
                        try:
                            ts_ms = int(ts)
                            from datetime import datetime, timezone
                            dt = datetime.fromtimestamp(ts_ms / 1000.0, tz=timezone.utc)
                            time_key = dt.strftime("%Y%m%d:%H%M")
                            temp_data[time_key] = values
                        except Exception:
                            pass
                            
        elif self._active_provider == "pvgis" and self._hourly_data:
            # PVGIS MODE: Temperature from PVGIS, wind would need to be provided separately
            data_source_name = "PVGIS"
            mb.showwarning(
                "Missing Wind Data",
                "PVGIS provides temperature data but not wind speed data.\n\n"
                "To calculate wind power, please:\n"
                "1. Switch to 'Renewable Ninja' provider in the Import Data tab\n"
                "2. Fetch data from Renewable Ninja (which provides both wind and temperature)\n\n"
                "Alternatively, manually enter wind speed values in the Input Parameters."
            )
            return
        else:
            # No data available
            mb.showwarning(
                "No Data Available",
                "No data has been fetched from the active provider.\n\n"
                "Please go to the Import Data tab and:\n"
                "1. Select either 'Renewable Ninja' or 'PVGIS'\n"
                "2. Fetch data for your location\n\n"
                "Note: Renewable Ninja is recommended as it provides both wind speed and temperature data."
            )
            return
        
        # Find common timestamps between wind and temperature data
        common_timestamps = sorted(set(wind_data.keys()) & set(temp_data.keys()))
        has_hourly_data = len(common_timestamps) > 0
        
        if not has_hourly_data:
            mb.showwarning(
                "Data Synchronization Error",
                f"Could not synchronize wind and temperature data from {data_source_name}.\n\n"
                f"Wind records: {len(wind_data)}\n"
                f"Temperature records: {len(temp_data)}\n"
                f"Matching timestamps: {len(common_timestamps)}\n\n"
                "Please try fetching the data again."
            )
            return
        
        n_hours = len(common_timestamps)
        
        # === CALCULATE ALL DATA FIRST (before creating popup) ===
        hourly_results = []
        total_power = 0
        
        # Constants for air density calculation
        rho_0 = 1.225
        T_0 = 288.15
        g = 9.80665
        M = 0.0289644
        R = 8.31447
        
        # Get hub height from Ninja form
        try:
            hub_height = float(self._ninja_height_entry.get() or 80)
        except ValueError:
            hub_height = 80.0
        
        # Total height = altitude + hub height
        total_height = altitude + hub_height
        
        # Swept area
        r = turbine_specs["rotor_diameter_m"] / 2
        A = 3.14159 * r ** 2
        Cp = 0.40
        
        for i, time_key in enumerate(common_timestamps):
            # Get synchronized wind and temperature data
            wind_record = wind_data[time_key]
            temp_record = temp_data[time_key]
            
            # Get temperature from Ninja solar data (temperature field)
            temp_c = temp_record.get("temperature", 15.0)
            
            # Get wind speed from Ninja
            V = wind_record.get("wind_speed", 0)
            
            # Get pressure if available (from wind data)
            pressure = wind_record.get("pressure", 101325)
            
            # Calculate air density using total height (altitude + hub height)
            T_kelvin = temp_c + 273.15
            # Barometric formula with total height
            P_alt = 101325 * ((T_kelvin - 0.0065 * total_height) / T_kelvin) ** (g * M / (R * 0.0065))
            rho = (P_alt * M) / (R * T_kelvin)
            
            # Calculate power for SELECTED MODEL ONLY
            if selected_model == "Model 1":
                P = 0.5 * rho * A * Cp * (V ** 3)
                # Cap at rated power
                rated_power_w = turbine_specs["rated_power_kw"] * 1000
                P = min(P, rated_power_w)
            else:  # Model 2
                P = self._calculate_model2(V, turbine_specs)
            
            # Format timestamp for display (YYYYMMDDHHMM -> YYYY-MM-DD HH:MM)
            try:
                year = time_key[0:4]
                month = time_key[4:6]
                day = time_key[6:8]
                hour = time_key[9:11]
                minute = time_key[11:13]
                timestamp = f"{year}-{month}-{day} {hour}:{minute}"
            except Exception:
                timestamp = f"Hour {i+1}"
            
            # Calculate air density ratio (ρ/ρ₀) where ρ₀ = 1.225 kg/m³
            rho_0 = 1.225
            rho_ratio = rho / rho_0
            
            hourly_results.append({
                "hour": i + 1,
                "timestamp": timestamp,
                "temp_c": temp_c,
                "wind_speed": V,
                "pressure": pressure / 100,
                "air_density": rho,
                "rho_ratio": rho_ratio,
                "power_kw": P / 1000,
                "Cp": Cp,
                "A": A,
            })
            
            total_power += P / 1000  # kW
        
        # Calculate averages
        avg_power = total_power / n_hours
        capacity_factor = (avg_power / turbine_specs["rated_power_kw"]) * 100
        
        # === STORE RESULTS FOR EXPORT ===
        self._last_energy_results = {
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "selected_model": selected_model,
            "turbine": turbine_name,
            "altitude_m": altitude,
            "hub_height_m": hub_height,
            "total_height_m": total_height,
            "swept_area_m2": A,
            "power_coefficient": Cp,
            "cut_in_speed": turbine_specs["cut_in_speed"],
            "rated_speed": turbine_specs["rated_speed"],
            "rated_power_kw": turbine_specs["rated_power_kw"],
            "total_hours": n_hours,
            "total_energy_kwh": total_power,
            "avg_power_kw": avg_power,
            "capacity_factor": capacity_factor,
            "hourly_data": hourly_results,
            "turbine_specs": turbine_specs,  # Store full specs for export
        }
        
        # === CREATE POPUP WINDOW ===
        popup = ctk.CTkToplevel(self)
        popup.title(f"{selected_model} Wind Power Calculation - {n_hours} Hours ({data_source_name})")
        popup.geometry("1200x800")
        popup.grab_set()
        
        # Header with title
        header = ctk.CTkFrame(popup, fg_color=Colors.BG_ELEVATED, height=60, corner_radius=0)
        header.pack(fill="x")
        header.pack_propagate(False)
        
        ctk.CTkFrame(header, width=4, height=24, fg_color=Colors.ACCENT).pack(side="left", padx=(16, 8), pady=18)
        title_text = f"{selected_model} - {n_hours} Hourly Records | Turbine: {turbine_name} | Altitude: {altitude:.0f}m | Source: {data_source_name}"
        ctk.CTkLabel(header, text=title_text, font=Fonts.H3, 
                     text_color=Colors.TEXT_DARK).pack(side="left", pady=18)
        
        # Tab bar
        tab_bar = ctk.CTkFrame(popup, fg_color=Colors.BG_ELEVATED, height=40, corner_radius=0)
        tab_bar.pack(fill="x")
        tab_bar.pack_propagate(False)
        
        # Content container
        content_frame = ctk.CTkFrame(popup, fg_color="transparent")
        content_frame.pack(fill="both", expand=True, padx=16, pady=12)
        
        # Tab buttons
        tab_names = ["Metadata", "Data Table", "Graphs", "Export"]
        tab_buttons = []
        
        def switch_tab(tab_name):
            # Clear content
            for child in content_frame.winfo_children():
                child.destroy()
            
            # Update button styles
            for btn in tab_buttons:
                if btn.cget("text") == tab_name:
                    btn.configure(fg_color="#6366f1", text_color="#ffffff")
                else:
                    btn.configure(fg_color="transparent", text_color=Colors.TEXT_SECONDARY)
            
            # Show appropriate content
            if tab_name == "Metadata":
                self._build_energy_metadata_tab(content_frame, selected_model, turbine_specs, 
                                                n_hours, total_power, avg_power, capacity_factor, 
                                                A, Cp, data_source_name, hub_height, total_height)
            elif tab_name == "Data Table":
                self._build_energy_data_table_tab(content_frame, hourly_results, selected_model, 
                                                  turbine_specs, Cp, n_hours, total_power, 
                                                  avg_power, capacity_factor, A)
            elif tab_name == "Graphs":
                self._build_energy_graphs_tab(content_frame, hourly_results, selected_model, 
                                              turbine_specs, Cp, A, n_hours)
            elif tab_name == "Export":
                self._build_energy_export_tab(content_frame, popup)
        
        for name in tab_names:
            btn = ctk.CTkButton(
                tab_bar, text=name, font=Fonts.BODY_BOLD, height=32,
                fg_color="transparent", text_color=Colors.TEXT_SECONDARY,
                hover_color=Colors.BG_CARD, corner_radius=6, width=120,
                command=lambda n=name: switch_tab(n)
            )
            btn.pack(side="left", padx=4, pady=4)
            tab_buttons.append(btn)
        
        # Summary statistics frame (shown in all tabs)
        summary_frame = ctk.CTkFrame(popup, fg_color=Colors.BG_CARD,
                                     corner_radius=Spacing.RADIUS_MD)
        summary_frame.pack(fill="x", padx=16, pady=(0, 12))
        
        # Display summary for SELECTED MODEL
        summary_inner = ctk.CTkFrame(summary_frame, fg_color="transparent")
        summary_inner.pack(fill="x", padx=Spacing.PAD_MD, pady=Spacing.PAD_MD)
        
        summary_grid = ctk.CTkFrame(summary_inner, fg_color="transparent")
        summary_grid.pack(fill="x")
        summary_grid.grid_columnconfigure((0, 1, 2, 3), weight=1)
        
        # Show stats based on selected model
        if selected_model == "Model 1":
            stats = [
                (f"Total Energy ({selected_model})", f"{total_power:.2f}", "kWh"),
                (f"Avg Power ({selected_model})", f"{avg_power:.3f}", "kW"),
                (f"Capacity Factor ({selected_model})", f"{capacity_factor:.1f}", "%"),
                ("Swept Area", f"{A:.2f}", "m²"),
            ]
        else:
            # Model 2: Show turbine constants in summary
            Vd = turbine_specs["cut_in_speed"]
            Vn = turbine_specs["rated_speed"]
            Voff = turbine_specs.get("cut_out_speed", 25.0)
            Pn = turbine_specs["rated_power_kw"]
            stats = [
                (f"Total Energy ({selected_model})", f"{total_power:.2f}", "kWh"),
                (f"Capacity Factor", f"{capacity_factor:.1f}", "%"),
                (f"Vd / Vn / Voff", f"{Vd:.1f} / {Vn:.1f} / {Voff:.1f}", "m/s"),
                (f"Pn (Nominal)", f"{Pn:.1f}", "kW"),
            ]
        
        for i, (label, value, unit) in enumerate(stats):
            f = ctk.CTkFrame(summary_grid, fg_color=Colors.BG_ELEVATED,
                            corner_radius=Spacing.RADIUS_SM)
            f.grid(row=0, column=i, sticky="ew", padx=Spacing.PAD_XS)
            ctk.CTkLabel(f, text=label, font=Fonts.CAPTION,
                         text_color=Colors.TEXT_SECONDARY).pack(anchor="w", padx=Spacing.PAD_SM, pady=(Spacing.PAD_SM, 0))
            val_frame = ctk.CTkFrame(f, fg_color="transparent")
            val_frame.pack(anchor="w", padx=Spacing.PAD_SM, pady=(0, Spacing.PAD_SM))
            ctk.CTkLabel(val_frame, text=value, font=Fonts.BODY_BOLD,
                         text_color=Colors.TEXT_PRIMARY).pack(side="left")
            ctk.CTkLabel(val_frame, text=f" {unit}", font=Fonts.CAPTION,
                         text_color=Colors.TEXT_MUTED).pack(side="left")
        
        # Initialize with Data Table tab
        switch_tab("Data Table")
        
        # Close button at bottom
        btn_frame = ctk.CTkFrame(popup, fg_color="transparent")
        btn_frame.pack(fill="x", padx=16, pady=(0, 12))
        
        ctk.CTkButton(
            btn_frame,
            text="Close",
            font=Fonts.BODY_BOLD,
            fg_color=Colors.BG_ELEVATED,
            hover_color=Colors.BG_HOVER,
            text_color=Colors.TEXT_PRIMARY,
            height=36,
            command=popup.destroy
        ).pack(side="right")
    
    def _export_hourly_energy_results(self, parent_window):
        """Export hourly energy calculation results to CSV."""
        import tkinter.filedialog as fd
        import csv
        
        if not hasattr(self, '_last_energy_results'):
            return
        
        filepath = fd.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
            initialfile="wind_power_hourly.csv",
            parent=parent_window
        )
        
        if not filepath:
            return
        
        try:
            with open(filepath, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                
                # Header info
                data = self._last_energy_results
                writer.writerow([f"{data['selected_model']} Wind Power Hourly Calculation Results"])
                writer.writerow(["Timestamp", data["timestamp"]])
                writer.writerow(["Model", data["selected_model"]])
                writer.writerow(["Turbine", data["turbine"]])
                writer.writerow(["Altitude", f"{data['altitude_m']:.0f}", "m"])
                writer.writerow(["Swept Area", f"{data['swept_area_m2']:.2f}", "m²"])
                writer.writerow([])
                
                # Summary
                writer.writerow(["Summary Statistics"])
                writer.writerow(["Total Hours", data["total_hours"]])
                writer.writerow(["Total Energy", f"{data['total_energy_kwh']:.2f}", "kWh"])
                writer.writerow(["Average Power", f"{data['avg_power_kw']:.3f}", "kW"])
                writer.writerow(["Capacity Factor", f"{data['capacity_factor']:.1f}", "%"])
                writer.writerow([])
                
                # Model-specific headers and data
                if data["selected_model"] == "Model 1":
                    # Model 1: Include Cp
                    writer.writerow(["Power Coefficient (Cp)", f"{data['power_coefficient']:.2f}"])
                    writer.writerow([])
                    writer.writerow(["Hour", "Timestamp", "Temperature (°C)", "Wind Speed (m/s)", 
                                   "Air Density (kg/m³)", "Cp", "Power (kW)"])
                    
                    for row in data["hourly_data"]:
                        writer.writerow([
                            row["hour"],
                            row["timestamp"],
                            f"{row['temp_c']:.2f}",
                            f"{row['wind_speed']:.2f}",
                            f"{row['air_density']:.4f}",
                            f"{data['power_coefficient']:.2f}",
                            f"{row['power_kw']:.3f}",
                        ])
                else:
                    # Model 2: Include Vd, Vn, Pn
                    writer.writerow(["Cut-in Speed (Vd)", f"{data['cut_in_speed']:.1f}", "m/s"])
                    writer.writerow(["Rated Speed (Vn)", f"{data['rated_speed']:.1f}", "m/s"])
                    writer.writerow(["Rated Power (Pn)", f"{data['rated_power_kw']:.1f}", "kW"])
                    writer.writerow([])
                    writer.writerow(["Hour", "Timestamp", "Vd (m/s)", "Vn (m/s)", "Pn (kW)", "Power (kW)"])
                    
                    for row in data["hourly_data"]:
                        writer.writerow([
                            row["hour"],
                            row["timestamp"],
                            f"{data['cut_in_speed']:.1f}",
                            f"{data['rated_speed']:.1f}",
                            f"{data['rated_power_kw']:.1f}",
                            f"{row['power_kw']:.3f}",
                        ])
            
            import tkinter.messagebox as mb
            mb.showinfo("Export Complete", f"Hourly results exported to:\n{filepath}", parent=parent_window)
        except Exception as e:
            import tkinter.messagebox as mb
            mb.showerror("Export Error", f"Failed to export:\n{str(e)}", parent=parent_window)

    def _export_energy_csv_multi(self, parent_window, hourly=False, daily=False, monthly=False, annual=False):
        """Export energy calculation results to CSV with selected aggregation mode."""
        import tkinter.filedialog as fd
        import csv
        
        if not hasattr(self, '_last_energy_results'):
            import tkinter.messagebox as mb
            mb.showwarning("No Data", "No calculation results available to export.", parent=parent_window)
            return
        
        # Determine mode and filename suffix
        if daily:
            mode = "daily"
            suffix = "daily"
        elif monthly:
            mode = "monthly"
            suffix = "monthly"
        elif annual:
            mode = "annual"
            suffix = "annual"
        else:  # Default to hourly
            mode = "hourly"
            suffix = "hourly"
        
        filepath = fd.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
            initialfile=f"wind_power_{suffix}.csv",
            parent=parent_window
        )
        
        if not filepath:
            return
        
        try:
            data = self._last_energy_results
            hourly_data = data["hourly_data"]
            selected_model = data["selected_model"]
            turbine_specs = data.get("turbine_specs", {
                "cut_in_speed": data.get("cut_in_speed", 3.0),
                "rated_speed": data.get("rated_speed", 12.0),
                "rated_power_kw": data.get("rated_power_kw", 2000),
                "cut_out_speed": data.get("cut_out_speed", 25.0)
            })
            Cp = data.get("power_coefficient", 0.35)
            A = data["swept_area_m2"]
            
            # Aggregate data based on mode
            if mode == "hourly":
                display_data = hourly_data
            elif mode == "daily":
                display_data = self._aggregate_energy_daily(hourly_data)
            elif mode == "monthly":
                display_data = self._aggregate_energy_monthly(hourly_data)
            else:  # annual
                display_data = self._aggregate_energy_annual(hourly_data)
            
            with open(filepath, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                
                # Header info
                writer.writerow([f"{selected_model} Wind Power {mode.capitalize()} Calculation Results"])
                writer.writerow(["Timestamp", data["timestamp"]])
                writer.writerow(["Model", selected_model])
                writer.writerow(["Turbine", data["turbine"]])
                writer.writerow(["Altitude", f"{data['altitude_m']:.0f}", "m"])
                writer.writerow(["Swept Area", f"{A:.2f}", "m²"])
                writer.writerow([])
                
                # Summary
                writer.writerow(["Summary Statistics"])
                if mode == "hourly":
                    writer.writerow(["Total Hours", len(hourly_data)])
                elif mode == "daily":
                    writer.writerow(["Total Days", len(display_data)])
                elif mode == "monthly":
                    writer.writerow(["Total Months", len(display_data)])
                else:
                    writer.writerow(["Total Yearly Records", len(display_data)])
                writer.writerow(["Total Energy", f"{data['total_energy_kwh']:.2f}", "kWh"])
                writer.writerow(["Average Power", f"{data['avg_power_kw']:.3f}", "kW"])
                writer.writerow(["Capacity Factor", f"{data['capacity_factor']:.1f}", "%"])
                writer.writerow([])
                
                # Data headers and rows based on model and mode
                if selected_model == "Model 1":
                    writer.writerow(["Power Coefficient (Cp)", f"{Cp:.2f}"])
                    writer.writerow([])
                    
                    if mode == "hourly":
                        writer.writerow(["Hour", "Timestamp", "Wind Speed (m/s)", "Temperature (°C)", 
                                       "Air Density (kg/m³)", "Cp", "Swept Area (m²)", "Power (kW)"])
                        for row in display_data:
                            writer.writerow([
                                row["hour"], row["timestamp"],
                                f"{row['wind_speed']:.2f}", f"{row['temp_c']:.2f}",
                                f"{row['air_density']:.4f}", f"{Cp:.2f}", f"{A:.2f}",
                                f"{row['power_kw']:.3f}"
                            ])
                    else:  # Daily, Monthly, Annual aggregates
                        time_col = "Day" if mode == "daily" else ("Month" if mode == "monthly" else "Year")
                        writer.writerow([time_col, "Avg Wind (m/s)", "Avg Temp (°C)", "Avg ρ (kg/m³)", 
                                       "Cp", "Swept Area (m²)", "Energy (kWh)"])
                        for row in display_data:
                            writer.writerow([
                                row.get("period", "-"),
                                f"{row['avg_wind_speed']:.2f}", f"{row['avg_temp_c']:.2f}",
                                f"{row['avg_air_density']:.4f}", f"{Cp:.2f}", f"{A:.2f}",
                                f"{row['total_energy_kwh']:.2f}"
                            ])
                else:
                    # Model 2
                    Vd = turbine_specs["cut_in_speed"]
                    Vn = turbine_specs["rated_speed"]
                    Pn = turbine_specs["rated_power_kw"]
                    Voff = turbine_specs.get("cut_out_speed", 25.0)
                    
                    writer.writerow(["Cut-in Speed (Vd)", f"{Vd:.1f}", "m/s"])
                    writer.writerow(["Rated Speed (Vn)", f"{Vn:.1f}", "m/s"])
                    writer.writerow(["Cut-out Speed (Voff)", f"{Voff:.1f}", "m/s"])
                    writer.writerow(["Rated Power (Pn)", f"{Pn:.1f}", "kW"])
                    writer.writerow([])
                    
                    if mode == "hourly":
                        writer.writerow(["Hour", "Timestamp", "Wind Speed (m/s)", "Air Density (kg/m³)",
                                       "Power (kW)", "Vd (m/s)", "Vn (m/s)", "Pn (kW)", "Voff (m/s)"])
                        for row in display_data:
                            writer.writerow([
                                row["hour"], row["timestamp"],
                                f"{row['wind_speed']:.2f}", f"{row['air_density']:.4f}",
                                f"{row['power_kw']:.3f}", f"{Vd:.1f}", f"{Vn:.1f}", f"{Pn:.1f}", f"{Voff:.1f}"
                            ])
                    else:  # Aggregates
                        time_col = "Day" if mode == "daily" else ("Month" if mode == "monthly" else "Year")
                        writer.writerow([time_col, "Avg Wind (m/s)", "Avg Temp (°C)", "Avg ρ (kg/m³)",
                                       "Vd (m/s)", "Vn (m/s)", "Pn (kW)", "Voff (m/s)", 
                                       "Energy (kWh)"])
                        for row in display_data:
                            writer.writerow([
                                row.get("period", "-"),
                                f"{row['avg_wind_speed']:.2f}", f"{row['avg_temp_c']:.2f}",
                                f"{row['avg_air_density']:.4f}",
                                f"{Vd:.1f}", f"{Vn:.1f}", f"{Pn:.1f}", f"{Voff:.1f}",
                                f"{row['total_energy_kwh']:.2f}"
                            ])
            
            import tkinter.messagebox as mb
            mb.showinfo("Export Complete", f"{mode.capitalize()} results exported to:\n{filepath}", parent=parent_window)
        except Exception as e:
            import tkinter.messagebox as mb
            mb.showerror("Export Error", f"Failed to export:\n{str(e)}", parent=parent_window)

    def _export_energy_excel_multi(self, parent_window):
        """Export energy calculation results to Excel with all 4 sheets."""
        import tkinter.filedialog as fd
        
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            import tkinter.messagebox as mb
            mb.showerror("Missing Dependency", 
                        "openpyxl is required for Excel export.\nInstall with: pip install openpyxl",
                        parent=parent_window)
            return
        
        if not hasattr(self, '_last_energy_results'):
            import tkinter.messagebox as mb
            mb.showwarning("No Data", "No calculation results available to export.", parent=parent_window)
            return
        
        filepath = fd.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile="wind_power_all_sheets.xlsx",
            parent=parent_window
        )
        
        if not filepath:
            return
        
        try:
            data = self._last_energy_results
            hourly_data = data["hourly_data"]
            selected_model = data["selected_model"]
            turbine_specs = data.get("turbine_specs", {
                "cut_in_speed": data.get("cut_in_speed", 3.0),
                "rated_speed": data.get("rated_speed", 12.0),
                "rated_power_kw": data.get("rated_power_kw", 2000),
                "cut_out_speed": data.get("cut_out_speed", 25.0)
            })
            Cp = data.get("power_coefficient", 0.35)
            A = data["swept_area_m2"]
            
            # Create workbook
            wb = openpyxl.Workbook()
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4B5563", end_color="4B5563", fill_type="solid")
            info_font = Font(bold=True)
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            def create_info_sheet(ws, title, agg_data, mode):
                """Create a sheet with info header and data table."""
                # Title
                ws['A1'] = f"{selected_model} Wind Power {title}"
                ws['A1'].font = Font(bold=True, size=14)
                
                # Metadata
                ws['A3'] = "Timestamp:"
                ws['B3'] = data["timestamp"]
                ws['A4'] = "Model:"
                ws['B4'] = selected_model
                ws['A5'] = "Turbine:"
                ws['B5'] = data["turbine"]
                ws['A6'] = "Altitude:"
                ws['B6'] = f"{data['altitude_m']:.0f} m"
                ws['A7'] = "Swept Area:"
                ws['B7'] = f"{A:.2f} m²"
                
                for row in range(3, 8):
                    ws[f'A{row}'].font = info_font
                
                # Summary
                ws['A9'] = "Summary Statistics"
                ws['A9'].font = info_font
                
                if mode == "hourly":
                    ws['A10'] = "Total Hours:"
                    ws['B10'] = len(hourly_data)
                elif mode == "daily":
                    ws['A10'] = "Total Days:"
                    ws['B10'] = len(agg_data)
                elif mode == "monthly":
                    ws['A10'] = "Total Months:"
                    ws['B10'] = len(agg_data)
                else:  # annual
                    ws['A10'] = "Total Yearly Records:"
                    ws['B10'] = len(agg_data)
                
                ws['A11'] = "Total Energy:"
                ws['B11'] = f"{data['total_energy_kwh']:.2f} kWh"
                ws['A12'] = "Average Power:"
                ws['B12'] = f"{data['avg_power_kw']:.3f} kW"
                ws['A13'] = "Capacity Factor:"
                ws['B13'] = f"{data['capacity_factor']:.1f}%"
                
                # Data table
                row_start = 15
                
                if selected_model == "Model 1":
                    ws[f'A{row_start}'] = f"Power Coefficient (Cp): {Cp:.2f}"
                    row_start += 2
                    
                    if mode == "hourly":
                        headers = ["Hour", "Timestamp", "Wind Speed (m/s)", "Temp (°C)", 
                                  "Air Density (kg/m³)", "ρ/ρ₀", "Cp", "Swept Area (m²)", "Power (kW)"]
                        for col, header in enumerate(headers, 1):
                            cell = ws.cell(row=row_start, column=col, value=header)
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.border = thin_border
                        
                        for i, row_data in enumerate(agg_data, row_start + 1):  # Export all hours
                            ws.cell(row=i, column=1, value=row_data["hour"])
                            ws.cell(row=i, column=2, value=row_data["timestamp"])
                            ws.cell(row=i, column=3, value=f"{row_data['wind_speed']:.2f}")
                            ws.cell(row=i, column=4, value=f"{row_data['temp_c']:.2f}")
                            ws.cell(row=i, column=5, value=f"{row_data['air_density']:.4f}")
                            ws.cell(row=i, column=6, value=f"{row_data.get('rho_ratio', 1.0):.4f}")
                            ws.cell(row=i, column=7, value=f"{Cp:.2f}")
                            ws.cell(row=i, column=8, value=f"{A:.2f}")
                            ws.cell(row=i, column=9, value=f"{row_data['power_kw']:.3f}")
                    else:
                        time_col = "Day" if mode == "daily" else ("Month" if mode == "monthly" else "Year")
                        headers = [time_col, "Avg Wind (m/s)", "Avg Temp (°C)", "Avg ρ (kg/m³)", 
                                  "Avg ρ/ρ₀", "Cp", "Swept Area (m²)", "Energy (kWh)"]
                        for col, header in enumerate(headers, 1):
                            cell = ws.cell(row=row_start, column=col, value=header)
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.border = thin_border
                        
                        for i, row_data in enumerate(agg_data, row_start + 1):
                            time_val = row_data.get("period", "-")
                            ws.cell(row=i, column=1, value=time_val)
                            ws.cell(row=i, column=2, value=f"{row_data['avg_wind_speed']:.2f}")
                            ws.cell(row=i, column=3, value=f"{row_data['avg_temp_c']:.2f}")
                            ws.cell(row=i, column=4, value=f"{row_data['avg_air_density']:.4f}")
                            ws.cell(row=i, column=5, value=f"{row_data.get('avg_rho_ratio', 1.0):.4f}")
                            ws.cell(row=i, column=6, value=f"{Cp:.2f}")
                            ws.cell(row=i, column=7, value=f"{A:.2f}")
                            ws.cell(row=i, column=8, value=f"{row_data['total_energy_kwh']:.2f}")
                else:
                    # Model 2
                    Vd = turbine_specs["cut_in_speed"]
                    Vn = turbine_specs["rated_speed"]
                    Pn = turbine_specs["rated_power_kw"]
                    Voff = turbine_specs.get("cut_out_speed", 25.0)
                    
                    ws[f'A{row_start}'] = f"Cut-in Speed (Vd): {Vd:.1f} m/s"
                    ws[f'A{row_start+1}'] = f"Rated Speed (Vn): {Vn:.1f} m/s"
                    ws[f'A{row_start+2}'] = f"Cut-out Speed (Voff): {Voff:.1f} m/s"
                    ws[f'A{row_start+3}'] = f"Rated Power (Pn): {Pn:.1f} kW"
                    row_start += 5
                    
                    if mode == "hourly":
                        headers = ["Hour", "Timestamp", "Wind Speed (m/s)", "Air Density (kg/m³)",
                                  "ρ/ρ₀", "Power (kW)", "Vd (m/s)", "Vn (m/s)", "Pn (kW)", "Voff (m/s)"]
                        for col, header in enumerate(headers, 1):
                            cell = ws.cell(row=row_start, column=col, value=header)
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.border = thin_border
                        
                        for i, row_data in enumerate(agg_data, row_start + 1):  # Export all hours
                            ws.cell(row=i, column=1, value=row_data["hour"])
                            ws.cell(row=i, column=2, value=row_data["timestamp"])
                            ws.cell(row=i, column=3, value=f"{row_data['wind_speed']:.2f}")
                            ws.cell(row=i, column=4, value=f"{row_data['air_density']:.4f}")
                            ws.cell(row=i, column=5, value=f"{row_data.get('rho_ratio', 1.0):.4f}")
                            ws.cell(row=i, column=6, value=f"{row_data['power_kw']:.3f}")
                            ws.cell(row=i, column=7, value=f"{Vd:.1f}")
                            ws.cell(row=i, column=8, value=f"{Vn:.1f}")
                            ws.cell(row=i, column=9, value=f"{Pn:.1f}")
                            ws.cell(row=i, column=10, value=f"{Voff:.1f}")
                    else:
                        time_col = "Day" if mode == "daily" else ("Month" if mode == "monthly" else "Year")
                        headers = [time_col, "Avg Wind (m/s)", "Avg Temp (°C)", "Avg ρ (kg/m³)",
                                  "Avg ρ/ρ₀", "Vd (m/s)", "Vn (m/s)", "Pn (kW)", "Voff (m/s)", 
                                  "Energy (kWh)"]
                        for col, header in enumerate(headers, 1):
                            cell = ws.cell(row=row_start, column=col, value=header)
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.border = thin_border
                        
                        for i, row_data in enumerate(agg_data, row_start + 1):
                            time_val = row_data.get("period", "-")
                            ws.cell(row=i, column=1, value=time_val)
                            ws.cell(row=i, column=2, value=f"{row_data['avg_wind_speed']:.2f}")
                            ws.cell(row=i, column=3, value=f"{row_data['avg_temp_c']:.2f}")
                            ws.cell(row=i, column=4, value=f"{row_data['avg_air_density']:.4f}")
                            ws.cell(row=i, column=5, value=f"{row_data.get('avg_rho_ratio', 1.0):.4f}")
                            ws.cell(row=i, column=6, value=f"{Vd:.1f}")
                            ws.cell(row=i, column=7, value=f"{Vn:.1f}")
                            ws.cell(row=i, column=8, value=f"{Pn:.1f}")
                            ws.cell(row=i, column=9, value=f"{Voff:.1f}")
                            ws.cell(row=i, column=10, value=f"{row_data['total_energy_kwh']:.2f}")
            
            # Create sheets
            modes = [("Hourly", hourly_data, "hourly"),
                     ("Daily", self._aggregate_energy_daily(hourly_data), "daily"),
                     ("Monthly", self._aggregate_energy_monthly(hourly_data), "monthly"),
                     ("Yearly", self._aggregate_energy_annual(hourly_data), "annual")]
            
            for idx, (title, agg_data, mode) in enumerate(modes):
                if idx == 0:
                    ws = wb.active
                    ws.title = title
                else:
                    ws = wb.create_sheet(title=title)
                create_info_sheet(ws, title, agg_data, mode)
            
            # Adjust column widths for all sheets
            for ws in wb.worksheets:
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(filepath)
            
            import tkinter.messagebox as mb
            mb.showinfo("Export Complete", f"Excel file exported with 4 sheets:\n{filepath}", parent=parent_window)
            
        except Exception as e:
            import tkinter.messagebox as mb
            mb.showerror("Export Error", f"Failed to export:\n{str(e)}", parent=parent_window)

    def _refresh_energy_table(self, parent, hourly_results, selected_model, mode, turbine_specs, Cp, n_hours, total_power, avg_power, capacity_factor, A):
        """Refresh the energy results table with aggregated data."""
        import customtkinter as ctk
        
        # Clear existing content
        for child in parent.winfo_children():
            child.destroy()
        
        # Aggregate data based on mode
        if mode == "hourly":
            display_data = hourly_results
            agg_label = f"Hourly Results (showing first 100 of {len(hourly_results)})"
        elif mode == "daily":
            display_data = self._aggregate_energy_daily(hourly_results)
            agg_label = f"Daily Aggregates ({len(display_data)} days)"
        elif mode == "monthly":
            display_data = self._aggregate_energy_monthly(hourly_results)
            agg_label = f"Monthly Aggregates ({len(display_data)} months)"
        else:  # annual
            display_data = self._aggregate_energy_annual(hourly_results)
            agg_label = f"Yearly Aggregates ({len(display_data)} years)"
        
        # Create table frame
        table_frame = ctk.CTkFrame(parent, fg_color=Colors.BG_CARD,
                                   corner_radius=Spacing.RADIUS_MD)
        table_frame.pack(fill="both", expand=True)
        
        ctk.CTkLabel(table_frame, text=f"{selected_model} {agg_label}", font=Fonts.H3,
                     text_color=Colors.TEXT_DARK).pack(anchor="w", padx=Spacing.PAD_MD, pady=(Spacing.PAD_SM, 0))
        
        # Use CTkScrollableFrame for better performance
        text_container = ctk.CTkScrollableFrame(table_frame, fg_color="transparent", height=400)
        text_container.pack(fill="both", expand=True, padx=Spacing.PAD_MD, pady=Spacing.PAD_SM)
        
        # Create header - different columns for each model and mode
        header_frame = ctk.CTkFrame(text_container, fg_color=Colors.BG_ELEVATED, height=30)
        header_frame.pack(fill="x", pady=(0, 2))
        
        # Get turbine constants for display
        Vd = turbine_specs["cut_in_speed"]          # Cut-in speed (Model 2: Vn)
        Vn = turbine_specs["rated_speed"]           # Rated speed (Model 2: Vr)
        Voff = turbine_specs.get("cut_out_speed", 25.0)  # Cut-out/shutdown speed
        Pn = turbine_specs["rated_power_kw"]        # Nominal power
        rated_power_kw = turbine_specs["rated_power_kw"]
        
        # Column definitions based on mode and model
        if selected_model == "Model 1":
            # Model 1: Standard wind power equation
            if mode == "hourly":
                headers = ["Hour", "Timestamp", "Wind(m/s)", "Temp(°C)", "P(hPa)", "ρ(kg/m³)", "ρ/ρ₀", "Cp", "A(m²)", "Power(kW)"]
                widths = [5, 14, 9, 8, 8, 9, 7, 6, 8, 10]
            elif mode == "daily":
                headers = ["Day", "Avg Wind(m/s)", "Avg Temp(°C)", "Avg P(hPa)", "Avg ρ(kg/m³)", "Avg ρ/ρ₀", "Cp", "A(m²)", "Energy(kWh)", "CF(%)"]
                widths = [10, 11, 10, 10, 11, 8, 6, 8, 12, 8]
            elif mode == "monthly":
                headers = ["Month", "Avg Wind(m/s)", "Avg Temp(°C)", "Avg P(hPa)", "Avg ρ(kg/m³)", "Avg ρ/ρ₀", "Cp", "A(m²)", "Energy(kWh)", "CF(%)"]
                widths = [8, 11, 10, 10, 11, 8, 6, 8, 12, 8]
            else:  # annual
                headers = ["Year", "Avg Wind(m/s)", "Avg Temp(°C)", "Avg P(hPa)", "Avg ρ(kg/m³)", "Avg ρ/ρ₀", "Cp", "A(m²)", "Energy(kWh)", "CF(%)"]
                widths = [6, 11, 10, 10, 11, 8, 6, 8, 12, 8]
        else:
            # Model 2: Power curve model with turbine constants
            if mode == "hourly":
                # Hourly: Dynamic (v, ρ, ρ/ρ₀, P) + Constants (Vd, Vn, Pn, Voff)
                headers = ["Hour", "Timestamp", "v(m/s)", "ρ(kg/m³)", "ρ/ρ₀", "P(kW)", "Vd", "Vn", "Pn(kW)", "Voff"]
                widths = [5, 14, 8, 9, 7, 9, 6, 6, 9, 6]
            elif mode == "daily":
                # Daily: Averages (v, ρ, ρ/ρ₀) + Energy + Constants
                headers = ["Day", "Avg v(m/s)", "Avg ρ(kg/m³)", "Avg ρ/ρ₀", "Vd", "Vn", "Pn(kW)", "Voff", "E(kWh)", "CF(%)"]
                widths = [10, 10, 11, 8, 6, 6, 9, 6, 12, 8]
            elif mode == "monthly":
                # Monthly: Averages (v, ρ, ρ/ρ₀) + Energy + Constants
                headers = ["Month", "Avg v(m/s)", "Avg ρ(kg/m³)", "Avg ρ/ρ₀", "Vd", "Vn", "Pn(kW)", "Voff", "E(kWh)", "CF(%)"]
                widths = [8, 10, 11, 8, 6, 6, 9, 6, 12, 8]
            else:  # annual
                # Yearly: Averages (v, ρ, ρ/ρ₀) + Energy + Constants
                headers = ["Year", "Avg v(m/s)", "Avg ρ(kg/m³)", "Avg ρ/ρ₀", "Vd", "Vn", "Pn(kW)", "Voff", "E(kWh)", "CF(%)"]
                widths = [6, 10, 11, 8, 6, 6, 9, 6, 12, 8]
        
        for i, (h, w) in enumerate(zip(headers, widths)):
            ctk.CTkLabel(header_frame, text=h, font=Fonts.BODY_BOLD, 
                         text_color=Colors.TEXT_PRIMARY, width=w*8).pack(side="left", padx=2)
        
        # Add data rows (first 100 for performance)
        display_count = min(100, len(display_data))
        for result in display_data[:display_count]:
            row_frame = ctk.CTkFrame(text_container, fg_color="transparent", height=25)
            row_frame.pack(fill="x", pady=1)
            
            if selected_model == "Model 1":
                # Model 1 values
                if mode == "hourly":
                    values = [
                        f"{result['hour']}", result['timestamp'],
                        f"{result['wind_speed']:.2f}", f"{result['temp_c']:.1f}",
                        f"{result.get('pressure', 1013):.0f}", f"{result['air_density']:.4f}",
                        f"{result.get('rho_ratio', 1.0):.4f}",
                        f"{Cp:.2f}", f"{A:.2f}", f"{result['power_kw']:.3f}",
                    ]
                else:  # Aggregated modes
                    count = result.get('count', 1)
                    avg_power = result['total_energy_kwh'] / count if count > 0 else 0
                    cf = (avg_power / rated_power_kw) * 100 if rated_power_kw > 0 else 0
                    
                    values = [
                        result['period'], f"{result['avg_wind_speed']:.2f}", f"{result['avg_temp_c']:.1f}",
                        f"{result.get('avg_pressure', 1013):.0f}", f"{result['avg_air_density']:.4f}",
                        f"{result.get('avg_rho_ratio', 1.0):.4f}",
                        f"{Cp:.2f}", f"{A:.2f}", f"{result['total_energy_kwh']:.2f}", f"{cf:.1f}",
                    ]
            else:
                # Model 2 values with turbine constants
                if mode == "hourly":
                    # Hourly: Dynamic (v, ρ, ρ/ρ₀, P) + Constants (Vd, Vn, Pn, Voff)
                    values = [
                        f"{result['hour']}", result['timestamp'],
                        f"{result['wind_speed']:.2f}", f"{result['air_density']:.4f}",
                        f"{result.get('rho_ratio', 1.0):.4f}",
                        f"{result['power_kw']:.3f}",
                        f"{Vd:.1f}", f"{Vn:.1f}", f"{Pn:.1f}", f"{Voff:.1f}",
                    ]
                else:  # Aggregated modes: Averages (v, ρ, ρ/ρ₀) + Energy + Constants
                    count = result.get('count', 1)
                    avg_power = result['total_energy_kwh'] / count if count > 0 else 0
                    cf = (avg_power / rated_power_kw) * 100 if rated_power_kw > 0 else 0
                    
                    values = [
                        result['period'], f"{result['avg_wind_speed']:.2f}", f"{result['avg_air_density']:.4f}",
                        f"{result.get('avg_rho_ratio', 1.0):.4f}",
                        f"{Vd:.1f}", f"{Vn:.1f}", f"{Pn:.1f}", f"{Voff:.1f}",
                        f"{result['total_energy_kwh']:.2f}", f"{cf:.1f}",
                    ]
            
            for i, (val, w) in enumerate(zip(values, widths)):
                ctk.CTkLabel(row_frame, text=val, font=Fonts.BODY, 
                             text_color=Colors.TEXT_PRIMARY, width=w*8).pack(side="left", padx=2)
        
        if len(display_data) > display_count:
            ctk.CTkLabel(text_container, text=f"... and {len(display_data) - display_count} more rows", 
                         font=Fonts.CAPTION, text_color=Colors.TEXT_MUTED).pack(pady=Spacing.PAD_SM)
    
    def _build_energy_metadata_tab(self, parent, selected_model, turbine_specs, n_hours, 
                                    total_power, avg_power, capacity_factor, A, Cp, data_source_name,
                                    hub_height=80, total_height=80):
        """Build the Metadata tab for energy results popup."""
        import customtkinter as ctk
        
        scroll = ctk.CTkScrollableFrame(parent, fg_color="transparent",
                    scrollbar_button_color=Colors.BORDER_DEFAULT)
        scroll.pack(fill="both", expand=True)
        
        # Model Information
        ctk.CTkLabel(scroll, text="Calculation Model", font=Fonts.H3,
                     text_color=Colors.TEXT_DARK).pack(anchor="w", pady=(12, 4))
        ctk.CTkFrame(scroll, height=1, fg_color=Colors.BORDER_DEFAULT).pack(fill="x", pady=(0, 8))
        
        model_info = [
            ("Selected Model", selected_model),
            ("Data Source", data_source_name),
            ("Total Hours", str(n_hours)),
        ]
        
        for key, value in model_info:
            row = ctk.CTkFrame(scroll, fg_color="transparent")
            row.pack(fill="x", pady=1)
            ctk.CTkLabel(row, text=key, font=Fonts.BODY_BOLD,
                         text_color=Colors.TEXT_SECONDARY, width=200,
                         anchor="w").pack(side="left", padx=Spacing.PAD_SM)
            ctk.CTkLabel(row, text=str(value), font=Fonts.BODY,
                         text_color=Colors.TEXT_PRIMARY,
                         anchor="w").pack(side="left", fill="x", expand=True)
        
        # Turbine Specifications
        ctk.CTkLabel(scroll, text="Turbine Specifications", font=Fonts.H3,
                     text_color=Colors.TEXT_DARK).pack(anchor="w", pady=(16, 4))
        ctk.CTkFrame(scroll, height=1, fg_color=Colors.BORDER_DEFAULT).pack(fill="x", pady=(0, 8))
        
        Vd = turbine_specs["cut_in_speed"]
        Vn = turbine_specs["rated_speed"]
        Pn = turbine_specs["rated_power_kw"]
        
        turbine_info = [
            ("Turbine Model", turbine_specs.get("name", "Unknown")),
            ("Rated Power (Pn)", f"{Pn} kW"),
            ("Rotor Diameter", f"{turbine_specs['rotor_diameter_m']} m"),
            ("Cut-in Speed (Vd)", f"{Vd} m/s"),
            ("Rated Speed (Vn)", f"{Vn} m/s"),
        ]
        
        if selected_model == "Model 2":
            Voff = turbine_specs.get("cut_out_speed", 25.0)
            turbine_info.append(("Cut-out Speed (Voff)", f"{Voff} m/s"))
        
        turbine_info.extend([
            ("Swept Area (A)", f"{A:.2f} m²"),
        ])
        
        if selected_model == "Model 1":
            turbine_info.append(("Power Coefficient (Cp)", f"{Cp:.2f}"))
        
        for key, value in turbine_info:
            row = ctk.CTkFrame(scroll, fg_color="transparent")
            row.pack(fill="x", pady=1)
            ctk.CTkLabel(row, text=key, font=Fonts.BODY_BOLD,
                         text_color=Colors.TEXT_SECONDARY, width=200,
                         anchor="w").pack(side="left", padx=Spacing.PAD_SM)
            ctk.CTkLabel(row, text=str(value), font=Fonts.BODY,
                         text_color=Colors.TEXT_PRIMARY,
                         anchor="w").pack(side="left", fill="x", expand=True)
        
        # Results Summary
        ctk.CTkLabel(scroll, text="Results Summary", font=Fonts.H3,
                     text_color=Colors.TEXT_DARK).pack(anchor="w", pady=(16, 4))
        ctk.CTkFrame(scroll, height=1, fg_color=Colors.BORDER_DEFAULT).pack(fill="x", pady=(0, 8))
        
        results_info = [
            ("Total Energy", f"{total_power:.2f} kWh"),
            ("Average Power", f"{avg_power:.3f} kW"),
            ("Capacity Factor", f"{capacity_factor:.1f} %"),
        ]
        
        for key, value in results_info:
            row = ctk.CTkFrame(scroll, fg_color="transparent")
            row.pack(fill="x", pady=1)
            ctk.CTkLabel(row, text=key, font=Fonts.BODY_BOLD,
                         text_color=Colors.TEXT_SECONDARY, width=200,
                         anchor="w").pack(side="left", padx=Spacing.PAD_SM)
            ctk.CTkLabel(row, text=str(value), font=Fonts.BODY,
                         text_color=Colors.ACCENT,
                         anchor="w").pack(side="left", fill="x", expand=True)
        
        # Equation
        ctk.CTkLabel(scroll, text="Equation Used", font=Fonts.H3,
                     text_color=Colors.TEXT_DARK).pack(anchor="w", pady=(16, 4))
        ctk.CTkFrame(scroll, height=1, fg_color=Colors.BORDER_DEFAULT).pack(fill="x", pady=(0, 8))
        
        eq_card = ctk.CTkFrame(scroll, fg_color=Colors.ACCENT_BG,
                               corner_radius=Spacing.RADIUS_MD)
        eq_card.pack(fill="x", pady=(0, Spacing.PAD_MD))
        
        if selected_model == "Model 1":
            ctk.CTkLabel(eq_card, text="P = ½ · ρ · A · Cₚ · V³", font=Fonts.H3,
                         text_color=Colors.TEXT_DARK).pack(anchor="w", padx=Spacing.PAD_MD, 
                                                            pady=(Spacing.PAD_MD, Spacing.PAD_SM))
            ctk.CTkLabel(eq_card, text="Standard wind power equation", font=Fonts.CAPTION,
                         text_color=Colors.TEXT_MUTED).pack(anchor="w", padx=Spacing.PAD_MD, 
                                                             pady=(0, Spacing.PAD_MD))
            
            # Pressure and Air Density Calculations for Model 1
            ctk.CTkLabel(scroll, text="Pressure & Air Density Calculations", font=Fonts.H3,
                         text_color=Colors.TEXT_DARK).pack(anchor="w", pady=(16, 4))
            ctk.CTkFrame(scroll, height=1, fg_color=Colors.BORDER_DEFAULT).pack(fill="x", pady=(0, 8))
            
            calc_card = ctk.CTkFrame(scroll, fg_color=Colors.BG_CARD,
                                   corner_radius=Spacing.RADIUS_MD)
            calc_card.pack(fill="x", pady=(0, Spacing.PAD_MD))
            
            calc_inner = ctk.CTkFrame(calc_card, fg_color="transparent")
            calc_inner.pack(fill="x", padx=Spacing.PAD_MD, pady=Spacing.PAD_MD)
            
            # Two-column layout: Variables | Equations
            calc_columns = ctk.CTkFrame(calc_inner, fg_color="transparent")
            calc_columns.pack(fill="x")
            calc_columns.grid_columnconfigure((0, 1), weight=1)
            
            # Left column - Variables (clean table format)
            vars_frame = ctk.CTkFrame(calc_columns, fg_color="transparent")
            vars_frame.grid(row=0, column=0, sticky="nw", padx=(0, Spacing.PAD_MD))
            
            ctk.CTkLabel(vars_frame, text="Variables", font=Fonts.BODY_BOLD,
                        text_color=Colors.TEXT_DARK).pack(anchor="w", pady=(0, 8))
            
            # Table-style variable list: Symbol | Description | Unit
            variables_table = [
                ("P", "Power output", "W"),
                ("ρ", "Air density", "kg/m³"),
                ("A", "Swept area", "m²"),
                ("Cp", "Power coefficient", "-"),
                ("V", "Wind speed", "m/s"),
                ("", "", ""),  # Separator
                ("ρ₀", "Sea level air density", "1.225 kg/m³"),
                ("T₀", "Sea level temperature", "288.15 K"),
                ("g", "Gravity", "9.80665 m/s²"),
                ("M", "Molar mass of air", "0.0289644 kg/mol"),
                ("R", "Gas constant", "8.31447 J/(mol·K)"),
                ("", "", ""),  # Separator
                ("h_hub", "Hub height", f"{hub_height:.0f} m"),
                ("h_alt", "Altitude", f"{total_height - hub_height:.0f} m"),
                ("h_total", "Total height", f"{total_height:.0f} m"),
            ]
            
            for symbol, desc, unit in variables_table:
                if not symbol and not desc:  # Separator
                    sep = ctk.CTkFrame(vars_frame, height=1, fg_color=Colors.BORDER_DEFAULT)
                    sep.pack(fill="x", pady=6)
                    continue
                
                # Compact row: height 24px, dense layout
                var_row = ctk.CTkFrame(vars_frame, fg_color=Colors.BG_ELEVATED, corner_radius=4, height=24)
                var_row.pack(fill="x", pady=2)
                var_row.pack_propagate(False)
                
                # Symbol (left): 13px, Bold, Primary Blue
                ctk.CTkLabel(var_row, text=symbol, font=("Inter", 13, "bold"),
                            text_color=Colors.ACCENT, width=40, anchor="w").pack(side="left", padx=(12, 0), pady=2)
                # Description (center): 12px, Regular
                ctk.CTkLabel(var_row, text=desc, font=("Inter", 12),
                            text_color=Colors.TEXT_PRIMARY).pack(side="left", fill="x", expand=True, pady=2)
                # Unit (right): 11px, Muted Gray
                ctk.CTkLabel(var_row, text=unit, font=("Inter", 11),
                            text_color=Colors.TEXT_MUTED, width=90, anchor="e").pack(side="right", padx=(0, 12), pady=2)
            
            # Right column - Equations
            eqs_frame = ctk.CTkFrame(calc_columns, fg_color="transparent")
            eqs_frame.grid(row=0, column=1, sticky="nw", padx=(Spacing.PAD_MD, 0))
            
            ctk.CTkLabel(eqs_frame, text="Equations", font=Fonts.BODY_BOLD,
                        text_color=Colors.TEXT_DARK).pack(anchor="w", pady=(0, 8))
            
            # Pressure equation
            ctk.CTkLabel(eqs_frame, text="Pressure at total height (altitude + hub):", font=Fonts.BODY,
                        text_color=Colors.TEXT_SECONDARY).pack(anchor="w", pady=(4, 2))
            ctk.CTkLabel(eqs_frame, text="P = P₀ · ((T - 0.0065·h_total)/T₀)^(g·M/(R·0.0065))", font=Fonts.BODY_BOLD,
                        text_color=Colors.TEXT_PRIMARY).pack(anchor="w")
            ctk.CTkLabel(eqs_frame, text=f"where P₀ = 101325 Pa, h_total = {total_height:.0f} m", font=Fonts.CAPTION,
                        text_color=Colors.TEXT_MUTED).pack(anchor="w", pady=(2, 8))
            
            # Air density equation
            ctk.CTkLabel(eqs_frame, text="Air density:", font=Fonts.BODY,
                        text_color=Colors.TEXT_SECONDARY).pack(anchor="w", pady=(4, 2))
            ctk.CTkLabel(eqs_frame, text="ρ = (P · M) / (R · T)", font=Fonts.BODY_BOLD,
                        text_color=Colors.TEXT_PRIMARY).pack(anchor="w")
            ctk.CTkLabel(eqs_frame, text="Ideal gas law", font=Fonts.CAPTION,
                        text_color=Colors.TEXT_MUTED).pack(anchor="w", pady=(2, 8))
            
            # Swept Area equation
            rotor_diameter = turbine_specs.get("rotor_diameter_m", 80)
            ctk.CTkLabel(eqs_frame, text="Swept area:", font=Fonts.BODY,
                        text_color=Colors.TEXT_SECONDARY).pack(anchor="w", pady=(4, 2))
            ctk.CTkLabel(eqs_frame, text=f"A = π · D²/4 = π · ({rotor_diameter})²/4 = {A:.2f} m²", font=Fonts.BODY_BOLD,
                        text_color=Colors.TEXT_PRIMARY).pack(anchor="w")
            ctk.CTkLabel(eqs_frame, text=f"where D = {rotor_diameter} m (rotor diameter)", font=Fonts.CAPTION,
                        text_color=Colors.TEXT_MUTED).pack(anchor="w", pady=(2, 0))
            
        else:
            # Model 2 equation with fraction
            frac_container = ctk.CTkFrame(eq_card, fg_color="transparent")
            frac_container.pack(anchor="w", padx=Spacing.PAD_MD, 
                                pady=(Spacing.PAD_MD, Spacing.PAD_SM))
            
            ctk.CTkLabel(frac_container, text="P = Pₙ · ", font=Fonts.H3,
                         text_color=Colors.TEXT_DARK).pack(side="left", anchor="center")
            
            fraction = ctk.CTkFrame(frac_container, fg_color="transparent")
            fraction.pack(side="left", anchor="center")
            
            ctk.CTkLabel(fraction, text="V³ - Vd³", font=Fonts.H3,
                         text_color=Colors.TEXT_DARK).pack(anchor="center")
            ctk.CTkFrame(fraction, height=2, fg_color=Colors.TEXT_DARK).pack(fill="x", padx=2)
            ctk.CTkLabel(fraction, text="Vn³ - Vd³", font=Fonts.H3,
                         text_color=Colors.TEXT_DARK).pack(anchor="center")
            
            ctk.CTkLabel(eq_card, text="Cubic power curve model", font=Fonts.CAPTION,
                         text_color=Colors.TEXT_MUTED).pack(anchor="w", padx=Spacing.PAD_MD, 
                                                             pady=(0, Spacing.PAD_MD))
    
    def _build_energy_data_table_tab(self, parent, hourly_results, selected_model, turbine_specs, 
                                      Cp, n_hours, total_power, avg_power, capacity_factor, A):
        """Build the Data Table tab for energy results popup."""
        import customtkinter as ctk
        
        # Variable to track aggregation mode
        agg_mode = ctk.StringVar(value="hourly")
        
        # Aggregation mode selection frame
        agg_frame = ctk.CTkFrame(parent, fg_color=Colors.BG_CARD, corner_radius=Spacing.RADIUS_MD)
        agg_frame.pack(fill="x", pady=(0, Spacing.PAD_MD))
        
        agg_inner = ctk.CTkFrame(agg_frame, fg_color="transparent")
        agg_inner.pack(fill="x", padx=Spacing.PAD_MD, pady=Spacing.PAD_MD)
        
        ctk.CTkLabel(agg_inner, text="View Mode:", font=Fonts.BODY_BOLD,
                     text_color=Colors.TEXT_DARK).pack(side="left", padx=(0, Spacing.PAD_SM))
        
        # Container for table content that will be refreshed
        table_container = ctk.CTkFrame(parent, fg_color="transparent")
        table_container.pack(fill="both", expand=True)
        
        # Radio buttons
        ctk.CTkRadioButton(agg_inner, text="Hourly", variable=agg_mode, value="hourly",
                           font=Fonts.BODY, text_color=Colors.TEXT_PRIMARY,
                           command=lambda: self._refresh_energy_table(table_container, hourly_results, 
                                                                      selected_model, agg_mode.get(), 
                                                                      turbine_specs, Cp, n_hours, 
                                                                      total_power, avg_power, 
                                                                      capacity_factor, A)).pack(side="left", 
                                                                                                      padx=Spacing.PAD_SM)
        ctk.CTkRadioButton(agg_inner, text="Daily", variable=agg_mode, value="daily",
                           font=Fonts.BODY, text_color=Colors.TEXT_PRIMARY,
                           command=lambda: self._refresh_energy_table(table_container, hourly_results, 
                                                                      selected_model, agg_mode.get(), 
                                                                      turbine_specs, Cp, n_hours, 
                                                                      total_power, avg_power, 
                                                                      capacity_factor, A)).pack(side="left", 
                                                                                                      padx=Spacing.PAD_SM)
        ctk.CTkRadioButton(agg_inner, text="Monthly", variable=agg_mode, value="monthly",
                           font=Fonts.BODY, text_color=Colors.TEXT_PRIMARY,
                           command=lambda: self._refresh_energy_table(table_container, hourly_results, 
                                                                      selected_model, agg_mode.get(), 
                                                                      turbine_specs, Cp, n_hours, 
                                                                      total_power, avg_power, 
                                                                      capacity_factor, A)).pack(side="left", 
                                                                                                      padx=Spacing.PAD_SM)
        ctk.CTkRadioButton(agg_inner, text="Yearly", variable=agg_mode, value="annual",
                           font=Fonts.BODY, text_color=Colors.TEXT_PRIMARY,
                           command=lambda: self._refresh_energy_table(table_container, hourly_results, 
                                                                      selected_model, agg_mode.get(), 
                                                                      turbine_specs, Cp, n_hours, 
                                                                      total_power, avg_power, 
                                                                      capacity_factor, A)).pack(side="left", 
                                                                                                      padx=Spacing.PAD_SM)
        
        # Initial render with hourly data
        self._refresh_energy_table(table_container, hourly_results, selected_model, "hourly", 
                                   turbine_specs, Cp, n_hours, total_power, avg_power, 
                                   capacity_factor, A)
    
    def _build_energy_export_tab(self, parent, popup_window):
        """Build the Export tab for energy results popup with CSV and Excel options."""
        import customtkinter as ctk
        
        scroll = ctk.CTkScrollableFrame(parent, fg_color="transparent",
                    scrollbar_button_color=Colors.BORDER_DEFAULT)
        scroll.pack(fill="both", expand=True)
        
        # Title
        ctk.CTkLabel(scroll, text="Export Data", font=Fonts.H2,
                     text_color=Colors.TEXT_DARK).pack(pady=(20, 4))
        
        # === CSV Export Section ===
        csv_frame = ctk.CTkFrame(scroll, fg_color=Colors.BG_CARD, corner_radius=Spacing.RADIUS_MD)
        csv_frame.pack(fill="x", pady=(20, 12), padx=40)
        
        csv_inner = ctk.CTkFrame(csv_frame, fg_color="transparent")
        csv_inner.pack(fill="x", padx=Spacing.PAD_LG, pady=Spacing.PAD_LG)
        
        ctk.CTkLabel(csv_inner, text="Export as CSV (.csv)", font=Fonts.BODY_BOLD,
                     text_color=Colors.TEXT_DARK).pack(anchor="w", pady=(0, Spacing.PAD_SM))
        
        # Checkboxes for CSV export modes - mutually exclusive (radio button behavior)
        checkbox_frame = ctk.CTkFrame(csv_inner, fg_color="transparent")
        checkbox_frame.pack(fill="x", pady=(0, Spacing.PAD_MD))
        
        # Variables for checkboxes - only one can be selected at a time
        export_hourly = ctk.BooleanVar(value=True)
        export_daily = ctk.BooleanVar(value=False)
        export_monthly = ctk.BooleanVar(value=False)
        export_years = ctk.BooleanVar(value=False)
        
        # Store all export variables for mutual exclusivity
        export_vars = [export_hourly, export_daily, export_monthly, export_years]
        export_modes = ["hourly", "daily", "monthly", "annual"]
        
        def make_exclusive(selected_index):
            """Ensure only the selected checkbox is checked, others are unchecked."""
            for i, var in enumerate(export_vars):
                if i != selected_index:
                    var.set(False)
            # Ensure at least one is always selected
            if not any(var.get() for var in export_vars):
                export_vars[selected_index].set(True)
        
        ctk.CTkCheckBox(checkbox_frame, text="Hourly", variable=export_hourly,
                        command=lambda: make_exclusive(0),
                        font=Fonts.BODY, text_color=Colors.TEXT_PRIMARY).pack(side="left", padx=(0, 24))
        ctk.CTkCheckBox(checkbox_frame, text="Daily", variable=export_daily,
                        command=lambda: make_exclusive(1),
                        font=Fonts.BODY, text_color=Colors.TEXT_PRIMARY).pack(side="left", padx=(0, 24))
        ctk.CTkCheckBox(checkbox_frame, text="Monthly", variable=export_monthly,
                        command=lambda: make_exclusive(2),
                        font=Fonts.BODY, text_color=Colors.TEXT_PRIMARY).pack(side="left", padx=(0, 24))
        ctk.CTkCheckBox(checkbox_frame, text="Yearly", variable=export_years,
                        command=lambda: make_exclusive(3),
                        font=Fonts.BODY, text_color=Colors.TEXT_PRIMARY).pack(side="left", padx=(0, 24))
        
        # Export CSV button
        ctk.CTkButton(
            csv_inner, text="Export Selected as CSV",
            font=Fonts.BODY_BOLD, height=42, width=250,
            fg_color=Colors.SUCCESS, hover_color="#059669",
            text_color="#ffffff", corner_radius=Spacing.RADIUS_MD,
            command=lambda: self._export_energy_csv_multi(
                popup_window,
                export_hourly.get(), export_daily.get(), 
                export_monthly.get(), export_years.get()
            )
        ).pack(pady=(0, Spacing.PAD_SM))
        
        # === Excel Export Section ===
        excel_frame = ctk.CTkFrame(scroll, fg_color=Colors.BG_CARD, corner_radius=Spacing.RADIUS_MD)
        excel_frame.pack(fill="x", pady=(12, 12), padx=40)
        
        excel_inner = ctk.CTkFrame(excel_frame, fg_color="transparent")
        excel_inner.pack(fill="x", padx=Spacing.PAD_LG, pady=Spacing.PAD_LG)
        
        ctk.CTkLabel(excel_inner, text="Export as Excel (.xlsx)", font=Fonts.BODY_BOLD,
                     text_color=Colors.TEXT_DARK).pack(anchor="w", pady=(0, Spacing.PAD_SM))
        
        ctk.CTkLabel(excel_inner, text="Exports all 4 sheets: Hourly, Daily, Monthly, Yearly",
                     font=Fonts.CAPTION, text_color=Colors.TEXT_MUTED).pack(anchor="w", pady=(0, Spacing.PAD_MD))
        
        # Export Excel button
        ctk.CTkButton(
            excel_inner, text="Export All as Excel",
            font=Fonts.BODY_BOLD, height=42, width=250,
            fg_color="#6366f1", hover_color="#4f46e5",
            text_color="#ffffff", corner_radius=Spacing.RADIUS_MD,
            command=lambda: self._export_energy_excel_multi(popup_window)
        ).pack(pady=(0, Spacing.PAD_SM))
    
    def _aggregate_energy_monthly(self, hourly_results):
        """Aggregate hourly energy results to monthly totals."""
        from collections import defaultdict
        
        monthly_data = defaultdict(lambda: {
            'temp_c_sum': 0, 'wind_speed_sum': 0, 'pressure_sum': 0, 'air_density_sum': 0,
            'rho_ratio_sum': 0, 'power_sum': 0, 'count': 0
        })
        
        for result in hourly_results:
            # Extract month from timestamp (format: "YYYY-MM-DD HH:MM")
            timestamp = result['timestamp']
            try:
                month_key = timestamp[:7]  # YYYY-MM
            except:
                continue
            
            monthly_data[month_key]['temp_c_sum'] += result.get('temp_c', 0)
            monthly_data[month_key]['wind_speed_sum'] += result.get('wind_speed', 0)
            monthly_data[month_key]['pressure_sum'] += result.get('pressure', 1013)
            monthly_data[month_key]['air_density_sum'] += result.get('air_density', 0)
            monthly_data[month_key]['rho_ratio_sum'] += result.get('rho_ratio', 1.0)
            monthly_data[month_key]['power_sum'] += result.get('power_kw', 0)
            monthly_data[month_key]['count'] += 1
        
        # Build aggregated results
        agg_results = []
        for month_key in sorted(monthly_data.keys()):
            data = monthly_data[month_key]
            count = data['count']
            if count > 0:
                agg_results.append({
                    'period': month_key,
                    'avg_temp_c': data['temp_c_sum'] / count,
                    'avg_wind_speed': data['wind_speed_sum'] / count,
                    'avg_pressure': data['pressure_sum'] / count,
                    'avg_air_density': data['air_density_sum'] / count,
                    'avg_rho_ratio': data['rho_ratio_sum'] / count,
                    'total_energy_kwh': data['power_sum'],  # Sum of hourly power = total energy (kWh)
                    'count': count,
                })
        
        return agg_results
    
    def _aggregate_energy_daily(self, hourly_results):
        """Aggregate hourly energy results to daily totals."""
        from collections import defaultdict
        
        daily_data = defaultdict(lambda: {
            'temp_c_sum': 0, 'wind_speed_sum': 0, 'pressure_sum': 0, 'air_density_sum': 0,
            'rho_ratio_sum': 0, 'power_sum': 0, 'count': 0
        })
        
        for result in hourly_results:
            # Extract day from timestamp (format: "YYYY-MM-DD HH:MM")
            timestamp = result['timestamp']
            try:
                day_key = timestamp[:10]  # YYYY-MM-DD
            except:
                continue
            
            daily_data[day_key]['temp_c_sum'] += result.get('temp_c', 0)
            daily_data[day_key]['wind_speed_sum'] += result.get('wind_speed', 0)
            daily_data[day_key]['pressure_sum'] += result.get('pressure', 1013)
            daily_data[day_key]['air_density_sum'] += result.get('air_density', 0)
            daily_data[day_key]['rho_ratio_sum'] += result.get('rho_ratio', 1.0)
            daily_data[day_key]['power_sum'] += result.get('power_kw', 0)
            daily_data[day_key]['count'] += 1
        
        # Build aggregated results
        agg_results = []
        for day_key in sorted(daily_data.keys()):
            data = daily_data[day_key]
            count = data['count']
            if count > 0:
                agg_results.append({
                    'period': day_key,
                    'avg_temp_c': data['temp_c_sum'] / count,
                    'avg_wind_speed': data['wind_speed_sum'] / count,
                    'avg_pressure': data['pressure_sum'] / count,
                    'avg_air_density': data['air_density_sum'] / count,
                    'avg_rho_ratio': data['rho_ratio_sum'] / count,
                    'total_energy_kwh': data['power_sum'],  # Sum of hourly power = total energy (kWh)
                    'count': count,
                })
        
        return agg_results
    
    def _aggregate_energy_annual(self, hourly_results):
        """Aggregate hourly energy results to annual totals."""
        from collections import defaultdict
        
        annual_data = defaultdict(lambda: {
            'temp_c_sum': 0, 'wind_speed_sum': 0, 'pressure_sum': 0, 'air_density_sum': 0,
            'rho_ratio_sum': 0, 'power_sum': 0, 'count': 0
        })
        
        for result in hourly_results:
            # Extract year from timestamp (format: "YYYY-MM-DD HH:MM")
            timestamp = result['timestamp']
            try:
                year_key = timestamp[:4]  # YYYY
            except:
                continue
            
            annual_data[year_key]['temp_c_sum'] += result.get('temp_c', 0)
            annual_data[year_key]['wind_speed_sum'] += result.get('wind_speed', 0)
            annual_data[year_key]['pressure_sum'] += result.get('pressure', 1013)
            annual_data[year_key]['air_density_sum'] += result.get('air_density', 0)
            annual_data[year_key]['rho_ratio_sum'] += result.get('rho_ratio', 1.0)
            annual_data[year_key]['power_sum'] += result.get('power_kw', 0)
            annual_data[year_key]['count'] += 1
        
        # Build aggregated results
        agg_results = []
        for year_key in sorted(annual_data.keys()):
            data = annual_data[year_key]
            count = data['count']
            if count > 0:
                agg_results.append({
                    'period': year_key,
                    'avg_temp_c': data['temp_c_sum'] / count,
                    'avg_wind_speed': data['wind_speed_sum'] / count,
                    'avg_pressure': data['pressure_sum'] / count,
                    'avg_air_density': data['air_density_sum'] / count,
                    'avg_rho_ratio': data['rho_ratio_sum'] / count,
                    'total_energy_kwh': data['power_sum'],  # Sum of hourly power = total energy (kWh)
                    'count': count,
                })
        
        return agg_results

    def _fetch_tmy_data(self):
        """Fetch TMY data from PVGIS API."""
        try:
            lat = float(self._tmy_lat_entry.get())
            lon = float(self._tmy_lon_entry.get())
        except ValueError:
            self._tmy_status_label.configure(
                text="Error: Please enter valid latitude and longitude values",
                text_color="#ef4444"
            )
            return
        
        if not (-90 <= lat <= 90):
            self._tmy_status_label.configure(
                text="Error: Latitude must be between -90 and 90",
                text_color="#ef4444"
            )
            return
        
        if not (-180 <= lon <= 180):
            self._tmy_status_label.configure(
                text="Error: Longitude must be between -180 and 180",
                text_color="#ef4444"
            )
            return
        
        database_selection = self._tmy_database_menu.get()
        raddatabase = database_selection.split(":")[0].strip()
        
        self._tmy_status_label.configure(
            text="Fetching TMY data from PVGIS... Please wait.",
            text_color=Colors.TEXT_SECONDARY
        )
        
        thread = threading.Thread(target=self._call_tmy_api, args=(lat, lon, raddatabase))
        thread.daemon = True
        thread.start()

    def _call_tmy_api(self, lat, lon, raddatabase):
        """Call PVGIS TMY API in a separate thread."""
        try:
            url = "https://re.jrc.ec.europa.eu/api/v5_3/tmy"
            params = {"lat": lat, "lon": lon, "outputformat": "json"}
            response = requests.get(url, params=params, timeout=30)
            response.raise_for_status()
            data = response.json()
            self.after(0, lambda: self._handle_tmy_success(data))
        except requests.exceptions.Timeout:
            self.after(0, lambda: self._tmy_status_label.configure(
                text="Error: Request timed out. Please try again.", text_color="#ef4444"))
        except requests.exceptions.RequestException as err:
            error_msg = str(err)
            # Clean up error message - remove URL if present
            if "REQUEST for url:" in error_msg or "http" in error_msg.lower():
                error_msg = error_msg.split("REQUEST for url:")[0].strip()
                error_msg = error_msg.split("http")[0].strip()
                if not error_msg:
                    error_msg = "API request failed. Check your connection and coordinates."
            self.after(0, lambda msg=error_msg: self._tmy_status_label.configure(
                text=f"Error: {msg}", text_color="#ef4444"))
        except Exception as err:
            error_msg = str(err)
            self.after(0, lambda msg=error_msg: self._tmy_status_label.configure(
                text=f"Error: {msg}", text_color="#ef4444"))

    def _handle_tmy_success(self, data):
        """Handle successful TMY API response."""
        self._tmy_data = data
        inputs = data.get("inputs", {})
        location = inputs.get("location", {})
        meteo_data = inputs.get("meteo_data", {})
        
        success_msg = f"✓ TMY data received! Lat: {location.get('latitude', 'N/A')}, Lon: {location.get('longitude', 'N/A')}, DB: {meteo_data.get('radiation_db', 'N/A')}"
        
        # Check if widgets still exist (page may have been destroyed)
        try:
            self._tmy_status_label.configure(text=success_msg, text_color="#10b981")
            self._tmy_viz_btn.configure(state="normal")
        except Exception:
            pass
        
        print(f"TMY Data received: {len(str(data))} bytes")
        
        # Auto-fill energy page with PVGIS data (if PVGIS is active provider)
        if self._active_provider == "pvgis":
            try:
                self._auto_fill_energy_data_pvgis()
            except Exception:
                pass

    def _fetch_altitude_auto(self):
        """Fetch altitude automatically - tries PVGIS first, then Open-Elevation as fallback."""
        import tkinter.messagebox as mb
        
        # Get coordinates from active provider
        lat, lon, source = self._get_coordinates_for_altitude()
        
        if lat is None or lon is None:
            mb.showwarning("Missing Coordinates", 
                          "Please enter valid latitude and longitude in the Import Data tab first.")
            return
        
        # Update UI to show fetching
        self._energy_alt_status.configure(text="⟳ fetching...", text_color=Colors.ACCENT)
        self._energy_alt_fetch_btn.configure(state="disabled")
        
        # Fetch in background thread - try PVGIS first, then Open-Elevation
        def fetch():
            elevation = None
            source_name = None
            
            # Try 1: PVGIS datarequest endpoint
            try:
                print(f"Fetching altitude from PVGIS for ({lat}, {lon})...")
                url = "https://re.jrc.ec.europa.eu/api/v5_3/datarequest"
                params = {
                    "lat": lat,
                    "lon": lon,
                    "outputformat": "json"
                }
                response = requests.get(url, params=params, timeout=15)
                response.raise_for_status()
                data = response.json()
                
                # Extract elevation from response
                if "meta" in data and "elevation" in data["meta"]:
                    elevation = data["meta"]["elevation"]
                elif "inputs" in data and "location" in data["inputs"]:
                    elevation = data["inputs"]["location"].get("elevation")
                elif "outputs" in data and "location" in data.get("outputs", {}):
                    elevation = data["outputs"]["location"].get("elevation")
                
                if elevation is not None:
                    source_name = "PVGIS"
                    print(f"✓ Altitude from PVGIS: {elevation}m")
                    
            except Exception as e:
                print(f"PVGIS altitude fetch failed: {e}")
            
            # Try 2: Open-Elevation API (fallback)
            if elevation is None:
                try:
                    print(f"Fetching altitude from Open-Elevation for ({lat}, {lon})...")
                    elevation = self._fetch_from_open_elevation(lat, lon)
                    if elevation is not None:
                        source_name = "Open-Elevation"
                        print(f"✓ Altitude from Open-Elevation: {elevation}m")
                except Exception as e:
                    print(f"Open-Elevation fetch failed: {e}")
            
            # Update UI with result
            if elevation is not None:
                self.after(0, lambda: self._update_altitude_success(elevation, lat, lon, source_name))
            else:
                self.after(0, lambda: self._update_altitude_error("All sources failed (PVGIS and Open-Elevation)"))
        
        thread = threading.Thread(target=fetch)
        thread.daemon = True
        thread.start()
    
    def _fetch_from_open_elevation(self, lat, lon):
        """Fetch elevation from Open-Elevation API.
        
        Open-Elevation is a free and open-source elevation API.
        Documentation: https://open-elevation.com/
        """
        # Open-Elevation API endpoint
        url = "https://api.open-elevation.com/api/v1/lookup"
        
        params = {
            "locations": f"{lat},{lon}"
        }
        
        response = requests.get(url, params=params, timeout=15)
        response.raise_for_status()
        data = response.json()
        
        # Parse response
        # Response format: {"results": [{"latitude": X, "longitude": Y, "elevation": Z}]}
        results = data.get("results", [])
        if results and len(results) > 0:
            elevation = results[0].get("elevation")
            return elevation
        
        return None
    
    def _get_coordinates_for_altitude(self):
        """Get coordinates from the active provider for altitude lookup.
        
        Returns:
            tuple: (lat, lon, source_name) or (None, None, None) if not available
        """
        lat = None
        lon = None
        source = None
        
        # Try active provider first
        if self._active_provider == "ninja":
            try:
                lat = float(self._ninja_lat_entry.get())
                lon = float(self._ninja_lon_entry.get())
                source = "ninja"
            except (ValueError, AttributeError):
                pass
        elif self._active_provider == "pvgis":
            try:
                lat = float(self._hourly_lat_entry.get())
                lon = float(self._hourly_lon_entry.get())
                source = "pvgis"
            except (ValueError, AttributeError):
                pass
        
        # If active provider doesn't have coordinates, try any available
        if lat is None or lon is None:
            # Try Ninja
            try:
                lat = float(self._ninja_lat_entry.get())
                lon = float(self._ninja_lon_entry.get())
                source = "ninja"
            except (ValueError, AttributeError):
                # Try PVGIS Hourly
                try:
                    lat = float(self._hourly_lat_entry.get())
                    lon = float(self._hourly_lon_entry.get())
                    source = "pvgis"
                except (ValueError, AttributeError):
                    pass
        
        return lat, lon, source
    
    def _update_altitude_success(self, elevation, lat, lon, source_name):
        """Update altitude field with successfully fetched value."""
        self._energy_alt_label.configure(text=str(int(round(elevation))))
        self._current_altitude = elevation  # Store for calculations
        self._energy_alt_status.configure(
            text=f"✓ from {source_name}", 
            text_color="#10b981"
        )
        self._energy_alt_fetch_btn.configure(state="normal")
        self._update_air_density()
        print(f"Altitude updated: {elevation}m at ({lat:.4f}, {lon:.4f}) from {source_name}")
    
    def _update_altitude_error(self, error_msg):
        """Show error when altitude fetch fails."""
        self._energy_alt_status.configure(text="⚠ fetch failed", text_color="#ef4444")
        self._energy_alt_fetch_btn.configure(state="normal")
        print(f"Error fetching altitude: {error_msg}")

    def _fetch_hourly_data(self):
        """Fetch hourly data from PVGIS API."""
        try:
            lat = float(self._hourly_lat_entry.get())
            lon = float(self._hourly_lon_entry.get())
        except ValueError:
            self._hourly_status_label.configure(
                text="Error: Please enter valid latitude and longitude values",
                text_color="#ef4444"
            )
            return
        
        if not (-90 <= lat <= 90):
            self._hourly_status_label.configure(
                text="Error: Latitude must be between -90 and 90",
                text_color="#ef4444"
            )
            return
        
        if not (-180 <= lon <= 180):
            self._hourly_status_label.configure(
                text="Error: Longitude must be between -180 and 180",
                text_color="#ef4444"
            )
            return
        
        # Get database and years
        raddatabase = self._hourly_database_menu.get()
        startyear = int(self._hourly_startyear_menu.get())
        endyear = int(self._hourly_endyear_menu.get())
        
        if startyear > endyear:
            self._hourly_status_label.configure(
                text="Error: Start year must be before or equal to end year",
                text_color="#ef4444"
            )
            return
        
        # Get all parameters
        params = {
            'mounting': self._mount_var.get(),
            'opt_slope': self._hourly_opt_slope_check.get(),
            'opt_both': self._hourly_opt_both_check.get(),
            'slope': self._hourly_slope_entry.get(),
            'azimuth': self._hourly_azimuth_entry.get(),
            'pv_enabled': self._hourly_pv_check.get(),
            'pvtech': self._hourly_pvtech_menu.get(),
            'peakpower': self._hourly_peakpower_entry.get(),
            'loss': self._hourly_loss_entry.get(),
            'components': self._hourly_components_check.get()
        }
        
        self._hourly_status_label.configure(
            text="Fetching hourly data from PVGIS...",
            text_color=Colors.TEXT_SECONDARY
        )
        
        # Set active provider to PVGIS when fetch is initiated
        self._active_provider = "pvgis"
        self._update_energy_provider_status()
        
        thread = threading.Thread(
            target=self._call_hourly_api, 
            args=(lat, lon, raddatabase, startyear, endyear, params)
        )
        thread.daemon = True
        thread.start()
    
    def _call_hourly_api(self, lat, lon, raddatabase, startyear, endyear, params):
        """Call PVGIS hourly API with all parameters."""
        try:
            url = "https://re.jrc.ec.europa.eu/api/v5_3/seriescalc"
            api_params = {
                "lat": lat,
                "lon": lon,
                "raddatabase": raddatabase,
                "startyear": startyear,
                "endyear": endyear,
                "outputformat": "json"
            }
            
            # PV calculation
            if params['pv_enabled']:
                api_params["pvcalculation"] = 1
                pvtech_map = {
                    "Crystalline silicon (crystSi)": "crystSi",
                    "CIS": "CIS",
                    "CdTe": "CdTe",
                    "Unknown": "Unknown"
                }
                api_params["pvtechchoice"] = pvtech_map.get(params['pvtech'], "crystSi")
                api_params["peakpower"] = float(params['peakpower'] or 1)
                api_params["loss"] = float(params['loss'] or 14)
            else:
                api_params["pvcalculation"] = 0
            
            # Mounting
            mount_map = {"Fixed": "free", "Vertical axis": "vertical", 
                        "Inclined axis": "inclined", "Two axis": "2axis"}
            api_params["mountingplace"] = mount_map.get(params['mounting'], "free")
            
            # Slope/azimuth optimization
            if params['opt_both']:
                api_params["optimalangles"] = 1
            elif params['opt_slope']:
                api_params["optimalinclination"] = 1
                if params['azimuth']:
                    api_params["aspect"] = float(params['azimuth'])
            else:
                if params['slope']:
                    api_params["angle"] = float(params['slope'])
                if params['azimuth']:
                    api_params["aspect"] = float(params['azimuth'])
            
            # Components
            if params['components']:
                api_params["components"] = 1
            
            response = requests.get(url, params=api_params, timeout=60)
            response.raise_for_status()
            data = response.json()
            self.after(0, lambda: self._handle_hourly_success(data, startyear, endyear))
        except Exception as err:
            error_msg = str(err)
            # Clean up error message - remove URL if present (can be very long)
            if "REQUEST for url:" in error_msg or "http" in error_msg.lower():
                # Extract just the main error without the URL
                error_msg = error_msg.split("REQUEST for url:")[0].strip()
                error_msg = error_msg.split("http")[0].strip()
                if not error_msg:
                    error_msg = "API request failed. Check your connection and coordinates."
            self.after(0, lambda msg=error_msg: self._hourly_status_label.configure(
                text=f"Error: {msg}", text_color="#ef4444"))

    def _handle_hourly_success(self, data, startyear, endyear):
        """Handle successful hourly API response."""
        self._hourly_data = data
        inputs = data.get("inputs", {})
        location = inputs.get("location", {})
        outputs = data.get("outputs", {})
        hourly_data = outputs.get("hourly", [])
        
        success_msg = f"✓ Hourly data received! Lat: {location.get('latitude', 'N/A')}, Lon: {location.get('longitude', 'N/A')}, Years: {startyear}-{endyear}, Records: {len(hourly_data)}"
        
        # Check if widgets still exist (page may have been destroyed)
        try:
            self._hourly_status_label.configure(text=success_msg, text_color="#10b981")
            self._hourly_viz_btn.configure(state="normal")
        except Exception:
            pass
        print(f"Hourly Data received: {len(str(data))} bytes, {len(hourly_data)} hourly records")
        
        # Auto-fill energy page with PVGIS data (if PVGIS is active provider)
        if self._active_provider == "pvgis":
            try:
                self._auto_fill_energy_data_pvgis()
            except Exception:
                pass
    
    def _on_mounting_change(self):
        """Handle mounting type changes - enable/disable slope/azimuth based on tracker type."""
        mount = self._mount_var.get()
        opt_slope = self._hourly_opt_slope_check.get()
        opt_both = self._hourly_opt_both_check.get()
        
        # Base state from mounting type (before optimization overrides)
        if mount == "Fixed":
            # Fixed mounting: user can set both slope and azimuth
            slope_enabled = True
            azimuth_enabled = True
            
        elif mount == "Vertical axis":
            # Vertical axis tracking: slope can be set, azimuth follows sun
            slope_enabled = True
            azimuth_enabled = False
            
        elif mount == "Inclined axis":
            # Inclined/tilted axis tracking: axis tilt is fixed, but orientation can be set
            slope_enabled = False
            azimuth_enabled = True
            
        elif mount == "Two axis":
            # Two axis tracking: both dimensions optimized automatically
            slope_enabled = False
            azimuth_enabled = False
        
        # Apply optimization overrides (optimization can further restrict)
        if opt_both:
            # If optimizing both, disable both regardless of mounting
            slope_enabled = False
            azimuth_enabled = False
        elif opt_slope:
            # If optimizing slope only, disable slope regardless of mounting
            slope_enabled = False
        
        # Apply the final state
        self._hourly_slope_entry.configure(state="normal" if slope_enabled else "disabled")
        self._hourly_azimuth_entry.configure(state="normal" if azimuth_enabled else "disabled")

    def _on_optimize_change(self):
        """Handle optimize checkbox changes - disable manual inputs when optimize is checked."""
        opt_slope = self._hourly_opt_slope_check.get()
        opt_both = self._hourly_opt_both_check.get()
        
        # If "optimize both" is checked, uncheck "optimize slope only"
        if opt_both and opt_slope:
            self._hourly_opt_slope_check.deselect()
        
        # If "optimize slope only" is checked, uncheck "optimize both"
        elif opt_slope and opt_both:
            self._hourly_opt_both_check.deselect()
        
        # Let mounting change handle the input states (respects both mounting and optimization)
        self._on_mounting_change()
    
    def _on_pv_check_change(self):
        """Handle PV power checkbox changes - enable/disable PV parameter inputs."""
        pv_enabled = self._hourly_pv_check.get()
        state = "normal" if pv_enabled else "disabled"
        
        # Enable/disable all PV-related widgets
        self._hourly_pvtech_menu.configure(state=state)
        self._hourly_peakpower_entry.configure(state=state)
        self._hourly_loss_entry.configure(state=state)
    
    def _fetch_ninja_data(self):
        """Fetch wind and temperature data from Renewable Ninja API (dual-API call)."""
        # Validate latitude and longitude
        lat_text = self._ninja_lat_entry.get().strip()
        lon_text = self._ninja_lon_entry.get().strip()
        
        # Highlight empty fields
        if not lat_text:
            self._ninja_lat_entry.configure(border_color="#ef4444")
        else:
            self._ninja_lat_entry.configure(border_color=Colors.BORDER_SUBTLE)
            
        if not lon_text:
            self._ninja_lon_entry.configure(border_color="#ef4444")
        else:
            self._ninja_lon_entry.configure(border_color=Colors.BORDER_SUBTLE)
        
        if not lat_text or not lon_text:
            self._ninja_status_label.configure(
                text="Error: Please enter both latitude and longitude values (shown above)",
                text_color="#ef4444"
            )
            # Scroll to top to show lat/lon fields
            self._ninja_lat_entry.focus_set()
            return
        
        try:
            lat = float(lat_text)
            lon = float(lon_text)
        except ValueError:
            self._ninja_status_label.configure(
                text="Error: Please enter valid numeric latitude and longitude values",
                text_color="#ef4444"
            )
            return
        
        if not (-90 <= lat <= 90):
            self._ninja_status_label.configure(
                text="Error: Latitude must be between -90 and 90",
                text_color="#ef4444"
            )
            return
        
        if not (-180 <= lon <= 180):
            self._ninja_status_label.configure(
                text="Error: Longitude must be between -180 and 180",
                text_color="#ef4444"
            )
            return
        
        # Get year
        year = self._ninja_year_menu.get()
        
        # Validate numeric inputs
        try:
            capacity = float(self._ninja_capacity_entry.get() or 1)
            height = int(self._ninja_height_entry.get() or 100)
        except ValueError:
            self._ninja_status_label.configure(
                text="Error: Please enter valid numeric values for capacity and height",
                text_color="#ef4444"
            )
            return
        
        if capacity <= 0 or height <= 0:
            self._ninja_status_label.configure(
                text="Error: Capacity and height must be positive numbers",
                text_color="#ef4444"
            )
            return
        
        # Get all parameters
        dataset_map = {"MERRA-2 (global)": "merra2"}
        dataset = dataset_map.get(self._ninja_dataset_menu.get(), "merra2")
        
        # Get turbine model - need to format for API
        turbine_display = self._ninja_turbine_menu.get()
        # Convert display name to API format (replace spaces with appropriate format if needed)
        turbine = self._format_turbine_for_api(turbine_display)
        
        raw = self._ninja_raw_check.get()
        
        print(f"DEBUG - Fetching with lat={lat}, lon={lon}, turbine={turbine}")
        
        self._ninja_status_label.configure(
            text="Fetching wind and temperature data from Renewable Ninja...",
            text_color=Colors.TEXT_SECONDARY
        )
        
        # Set active provider to Ninja when fetch is initiated
        self._active_provider = "ninja"
        self._update_energy_provider_status()
        
        # Start dual API calls in background thread
        thread = threading.Thread(
            target=self._call_ninja_api_dual,
            args=(lat, lon, year, capacity, height, turbine, dataset, raw)
        )
        thread.daemon = True
        thread.start()
    
    def _format_turbine_for_api(self, turbine_display):
        """Convert turbine display name to API format."""
        # The API expects specific turbine names - let's try the display name first
        # but provide a fallback to a known working turbine if needed
        
        # Map common turbines to their API names if different
        turbine_map = {
            # Add mappings here if API uses different names
            # Example: "Display Name": "api_name"
        }
        
        if turbine_display in turbine_map:
            return turbine_map[turbine_display]
        
        # Try the display name as-is (the API might accept it)
        return turbine_display

    def _call_ninja_api_dual(self, lat, lon, year, capacity, height, turbine, dataset, raw):
        """Call both Renewable Ninja wind and solar APIs to get wind speed and temperature."""
        import json
        
        # Build date range from year
        date_from = f"{year}-01-01"
        date_to = f"{year}-12-31"
        
        # Add token authentication header
        headers = {
            "Authorization": "Token 92fffc450a4b4f379d5499c0205d161a65a091a6"
        }
        
        # API Request 1: Wind data (for wind speed)
        wind_url = "https://www.renewables.ninja/api/data/wind"
        wind_params = {
            "lat": lat,
            "lon": lon,
            "date_from": date_from,
            "date_to": date_to,
            "capacity": capacity,
            "height": height,
            "turbine": turbine,
            "dataset": dataset,
            "format": "json",
            "raw": "true" if raw else "false"
        }
        
        # API Request 2: Solar/PV data (for temperature T2m)
        # Solar API requires additional parameters for PV calculation
        solar_url = "https://www.renewables.ninja/api/data/pv"
        solar_params = {
            "lat": lat,
            "lon": lon,
            "date_from": date_from,
            "date_to": date_to,
            "capacity": 1,  # Minimal capacity, we don't use power output
            "dataset": "merra2",  # Use merra2 for solar too
            "format": "json",
            "raw": "true",  # Include raw weather data including temperature
            # Required PV-specific parameters (we use defaults since we only need temperature)
            "tracking": 0,       # Fixed mounting (0 = fixed, 1 = single axis, 2 = dual axis)
            "azim": 180,         # Azimuth (180 = south-facing)
            "tilt": 30,          # Tilt angle in degrees
            "system_loss": 0.14  # System losses as decimal (0.14 = 14%)
        }
        
        print(f"DEBUG - Wind API params: {wind_params}")
        print(f"DEBUG - Solar API params: {solar_params}")
        
        try:
            # Fetch wind data
            self.after(0, lambda: self._ninja_status_label.configure(
                text="Fetching wind data from Renewable Ninja...",
                text_color=Colors.TEXT_SECONDARY))
            
            wind_response = requests.get(wind_url, params=wind_params, headers=headers, timeout=60)
            print(f"DEBUG - Wind API status: {wind_response.status_code}")
            if wind_response.status_code != 200:
                print(f"DEBUG - Wind API response: {wind_response.text[:500]}")
            wind_response.raise_for_status()
            wind_data = wind_response.json()
            print(f"DEBUG - Wind data received, keys: {wind_data.keys() if isinstance(wind_data, dict) else 'not dict'}")
            
            # Fetch solar data (for temperature)
            self.after(0, lambda: self._ninja_status_label.configure(
                text="Fetching temperature data from Renewable Ninja (via solar API)...",
                text_color=Colors.TEXT_SECONDARY))
            
            solar_response = requests.get(solar_url, params=solar_params, headers=headers, timeout=60)
            print(f"DEBUG - Solar API status: {solar_response.status_code}")
            if solar_response.status_code != 200:
                print(f"DEBUG - Solar API response: {solar_response.text[:500]}")
            solar_response.raise_for_status()
            solar_data = solar_response.json()
            print(f"DEBUG - Solar data received, keys: {solar_data.keys() if isinstance(solar_data, dict) else 'not dict'}")
            
            # Both successful - call combined success handler
            self.after(0, lambda: self._handle_ninja_success_dual(wind_data, solar_data, year))
            
        except requests.exceptions.Timeout:
            self.after(0, lambda: self._ninja_status_label.configure(
                text="Error: Request timed out. Please try again.",
                text_color="#ef4444"))
        except requests.exceptions.HTTPError as err:
            # Handle HTTP errors (400, 401, 403, etc.)
            error_msg = str(err)
            print(f"DEBUG - HTTP Error: {error_msg}")
            try:
                # Try to get more details from response
                if 'wind_response' in locals() and wind_response.status_code != 200:
                    status_code = wind_response.status_code
                    error_detail = wind_response.text[:500]
                    print(f"DEBUG - Wind error detail ({status_code}): {error_detail}")
                    
                    # Provide user-friendly error messages based on status code
                    if status_code == 400:
                        if "turbine" in error_detail.lower():
                            error_msg = f"Invalid turbine model '{turbine}'. Try a different turbine from the list."
                        elif "capacity" in error_detail.lower():
                            error_msg = f"Invalid capacity value. Must match turbine rated power (kW)."
                        elif "height" in error_detail.lower():
                            error_msg = f"Invalid hub height. Check the turbine specifications."
                        elif "coordinates" in error_detail.lower() or "lat" in error_detail.lower():
                            error_msg = f"Invalid coordinates ({lat}, {lon}). Try different values."
                        elif "date" in error_detail.lower():
                            error_msg = f"Invalid date range: {date_from} to {date_to}."
                        else:
                            error_msg = f"Invalid request (400). The turbine model or parameters may not be supported."
                    elif status_code == 401:
                        error_msg = "Authentication failed. The API token may be invalid."
                    elif status_code == 403:
                        error_msg = "Access denied. Check your API token or rate limits."
                    elif status_code == 404:
                        error_msg = "API endpoint not found. The service may be unavailable."
                    elif status_code >= 500:
                        error_msg = "Server error. Please try again later."
                    else:
                        error_msg = f"API Error ({status_code}): {error_detail[:100]}"
                        
                elif 'solar_response' in locals() and solar_response.status_code != 200:
                    status_code = solar_response.status_code
                    error_detail = solar_response.text[:500]
                    print(f"DEBUG - Solar error detail ({status_code}): {error_detail}")
                    error_msg = f"Solar API Error ({status_code}): Unable to fetch temperature data."
            except Exception as e:
                print(f"DEBUG - Error parsing error response: {e}")
            
            self.after(0, lambda msg=error_msg: self._ninja_status_label.configure(
                text=f"Error: {msg}",
                text_color="#ef4444"))
        except requests.exceptions.RequestException as err:
            error_msg = str(err)
            print(f"DEBUG - Request Exception: {error_msg}")
            # Clean up error message - remove URL if present
            if "REQUEST for url:" in error_msg or "http" in error_msg.lower():
                error_msg = error_msg.split("REQUEST for url:")[0].strip()
                error_msg = error_msg.split("http")[0].strip()
                if not error_msg:
                    error_msg = "API request failed. Check your connection and coordinates."
            self.after(0, lambda msg=error_msg: self._ninja_status_label.configure(
                text=f"Error: {msg}",
                text_color="#ef4444"))
        except Exception as err:
            error_msg = str(err)
            print(f"DEBUG - General Exception: {error_msg}")
            import traceback
            traceback.print_exc()
            self.after(0, lambda msg=error_msg: self._ninja_status_label.configure(
                text=f"Error: {msg}",
                text_color="#ef4444"))

    def _handle_ninja_success_dual(self, wind_data, solar_data, year):
        """Handle successful dual Renewable Ninja API responses (wind + solar for temp)."""
        self._ninja_data = wind_data
        self._ninja_temp_data = solar_data
        
        # Process wind data
        wind_entries = wind_data.get("data", {})
        wind_metadata = wind_data.get("metadata", {})
        wind_count = len(wind_entries) if isinstance(wind_entries, dict) else len(wind_entries)
        
        # Process solar data (for temperature)
        solar_entries = solar_data.get("data", {})
        solar_count = len(solar_entries) if isinstance(solar_entries, dict) else len(solar_entries)
        
        # Extract average temperature if available
        avg_temp = None
        try:
            if isinstance(solar_entries, dict):
                temps = [v.get("temperature", v.get("T2m")) for v in solar_entries.values() 
                        if isinstance(v, dict) and ("temperature" in v or "T2m" in v)]
                if temps:
                    avg_temp = sum(temps) / len(temps)
            elif isinstance(solar_entries, list) and solar_entries:
                temps = [r.get("temperature", r.get("T2m")) for r in solar_entries 
                        if isinstance(r, dict) and ("temperature" in r or "T2m" in r)]
                if temps:
                    avg_temp = sum(temps) / len(temps)
        except Exception as e:
            print(f"Could not extract temperature from solar data: {e}")
        
        # Build success message
        temp_info = f" | Avg Temp: {avg_temp:.1f}°C" if avg_temp else ""
        success_msg = f"✓ Ninja data received! Wind: {wind_count} records | Temp: {solar_count} records{temp_info}"
        
        # Check if widgets still exist (page may have been destroyed)
        try:
            self._ninja_status_label.configure(text=success_msg, text_color="#10b981")
            self._ninja_viz_btn.configure(state="normal")
        except Exception:
            pass
        
        print(f"Renewable Ninja - Wind data: {wind_count} records, Solar/Temp data: {solar_count} records")
        
        # Auto-fill energy page with Ninja data (wind + temp)
        try:
            self._auto_fill_energy_data_ninja()
        except Exception as e:
            print(f"Error auto-filling energy data: {e}")

    # ── Visualization & Export Popup ──────────────────────────────────
    def _show_data_popup(self, source):
        """Open a popup window to visualize and export fetched data."""
        # Get the correct data based on source
        data_map = {"tmy": self._tmy_data, "hourly": self._hourly_data, "ninja": self._ninja_data}
        data = data_map.get(source)
        
        if data is None:
            # Show error if no data fetched yet
            import tkinter.messagebox as mb
            mb.showwarning("No Data", f"Please fetch {source.upper()} data first before opening visualization.")
            return
        
        # Create toplevel popup
        title_map = {"tmy": "PVGIS TMY", "hourly": "PVGIS Hourly", "ninja": "Renewable Ninja Wind"}
        popup = ctk.CTkToplevel(self)
        popup.title(f"Visualization & Export — {title_map[source]}")
        popup.geometry("900x600")
        popup.configure(fg_color=Colors.BG_PRIMARY)
        popup.transient(self.winfo_toplevel())
        popup.grab_set()
        popup.after(100, popup.focus_force)
        
        # Header
        header = ctk.CTkFrame(popup, fg_color=Colors.BG_ELEVATED, height=50, corner_radius=0)
        header.pack(fill="x")
        header.pack_propagate(False)
        ctk.CTkFrame(header, width=4, height=24, fg_color="#6366f1").pack(side="left", padx=(16, 8))
        ctk.CTkLabel(header, text=f"Visualization & Export — {title_map[source]}",
                     font=Fonts.H3, text_color=Colors.TEXT_DARK).pack(side="left")
        
        # Tab system
        tab_bar = ctk.CTkFrame(popup, fg_color=Colors.BG_ELEVATED, height=40, corner_radius=0)
        tab_bar.pack(fill="x")
        tab_bar.pack_propagate(False)
        
        content_frame = ctk.CTkFrame(popup, fg_color="transparent")
        content_frame.pack(fill="both", expand=True, padx=16, pady=12)
        
        tab_names = ["Metadata", "Data Table", "Export"]
        tab_buttons = []
        
        def switch_tab(selected):
            for child in content_frame.winfo_children():
                child.destroy()
            for i, btn in enumerate(tab_buttons):
                if tab_names[i] == selected:
                    btn.configure(fg_color="#6366f1", text_color="#ffffff")
                else:
                    btn.configure(fg_color="transparent", text_color=Colors.TEXT_SECONDARY)
            if selected == "Metadata":
                self._build_metadata_tab(content_frame, data, source)
            elif selected == "Data Table":
                self._build_table_tab(content_frame, data, source)
            elif selected == "Export":
                self._build_export_tab(content_frame, data, source)
        
        for name in tab_names:
            btn = ctk.CTkButton(
                tab_bar, text=name, font=Fonts.BODY_BOLD, height=32,
                fg_color="transparent", text_color=Colors.TEXT_SECONDARY,
                hover_color=Colors.BG_CARD, corner_radius=6, width=120,
                command=lambda n=name: switch_tab(n)
            )
            btn.pack(side="left", padx=4, pady=4)
            tab_buttons.append(btn)
        
        # Show first tab
        switch_tab("Metadata")
    
    def _build_metadata_tab(self, parent, data, source):
        """Build the Metadata tab content."""
        scroll = ctk.CTkScrollableFrame(parent, fg_color="transparent",
                    scrollbar_button_color=Colors.BORDER_DEFAULT)
        scroll.pack(fill="both", expand=True)
        
        # Units mapping for common metadata fields
        units_map = {
            # Location
            "latitude": "°",
            "longitude": "°",
            "elevation": "m",
            "altitude": "m",
            "height": "m",
            # Meteorological
            "year_min": "",
            "year_max": "",
            "use_horizon": "",
            "horizon_db": "",
            "horizon_data": "",
            "radiation_db": "",
            "meteo_db": "",
            # Mounting System
            "slope": "°",
            "azimuth": "°",
            "tilt": "°",
            "orientation": "°",
            "nominal_power": "kWp",
            "peakpower": "kWp",
            "installed_power": "kWp",
            "system_loss": "%",
            "loss": "%",
            # PV Module
            "technology": "",
            "efficiency": "%",
            "temp_coeff": "%/°C",
            # Ninja metadata
            "latitude_ninja": "°",
            "longitude_ninja": "°",
            "dataset": "",
            "year": "",
            "height": "m",
        }
        
        if source in ("tmy", "hourly"):
            inputs = data.get("inputs", {})
            location = inputs.get("location", {})
            meteo = inputs.get("meteo_data", {})
            mounting = inputs.get("mounting_system", {})
            pv_module = inputs.get("pv_module", {})
            
            sections = {
                "Location": location,
                "Meteorological Data": meteo,
                "Mounting System": mounting,
                "PV Module": pv_module,
            }
        else:
            metadata = data.get("metadata", {})
            sections = {"Metadata": metadata}
        
        for section_name, section_data in sections.items():
            if not section_data:
                continue
            # Section header
            ctk.CTkLabel(scroll, text=section_name, font=Fonts.H3,
                         text_color=Colors.TEXT_DARK).pack(anchor="w", pady=(12, 4))
            ctk.CTkFrame(scroll, height=1, fg_color=Colors.BORDER_DEFAULT).pack(fill="x", pady=(0, 8))
            
            # Flatten nested dicts
            items = self._flatten_dict(section_data)
            for key, value in items:
                # Format value with units
                unit = units_map.get(key, "")
                if unit and value not in ("", None, "None", "True", "False"):
                    display_value = f"{value} {unit}"
                else:
                    display_value = str(value)
                
                row = ctk.CTkFrame(scroll, fg_color="transparent")
                row.pack(fill="x", pady=1)
                ctk.CTkLabel(row, text=str(key), font=Fonts.BODY_BOLD,
                             text_color=Colors.TEXT_SECONDARY, width=200,
                             anchor="w").pack(side="left")
                ctk.CTkLabel(row, text=display_value, font=Fonts.BODY,
                             text_color=Colors.TEXT_PRIMARY,
                             anchor="w").pack(side="left", fill="x", expand=True)
        
        # Add Data Columns section with descriptions and units
        self._add_column_descriptions(scroll, source)
    
    def _add_column_descriptions(self, parent, source):
        """Add column descriptions and units section to metadata."""
        # Column descriptions and units mapping
        column_info = {}
        
        if source in ("tmy", "hourly"):
            # PVGIS column descriptions
            column_info = {
                "P": ("PV electricity", "W"),
                "Gb(i)": ("Beam (direct) irradiance on inclined plane", "W/m²"),
                "Gd(i)": ("Diffuse irradiance on inclined plane", "W/m²"),
                "Gr(i)": ("Reflected irradiance on inclined plane", "W/m²"),
                "H_sun": ("Sun height", "°"),
                "T2m": ("Air temperature at 2m", "°C"),
                "WS10m": ("Wind speed at 10m", "m/s"),
                "Int": ("Global irradiance on inclined plane", "W/m²"),
                "time": ("Timestamp", ""),
                "timestamp": ("Timestamp", ""),
                "E": ("Daily PV electricity (integrated)", "Wh"),
            }
        elif source == "ninja":
            # Renewables Ninja column descriptions
            column_info = {
                "electricity": ("Wind electricity", "kW"),
                "wind_speed": ("Wind speed", "m/s"),
                "temperature": ("Air temperature", "°C"),
                "pressure": ("Air pressure", "hPa"),
                "density": ("Air density", "kg/m³"),
                "radiation": ("Solar radiation", "W/m²"),
                "irradiance": ("Solar irradiance", "W/m²"),
                "timestamp": ("Timestamp", ""),
                "E": ("Daily wind electricity (integrated)", "kWh"),
            }
        
        if not column_info:
            return
        
        # Section header
        ctk.CTkLabel(parent, text="Data Columns", font=Fonts.H3,
                     text_color=Colors.TEXT_DARK).pack(anchor="w", pady=(16, 4))
        ctk.CTkFrame(parent, height=1, fg_color=Colors.BORDER_DEFAULT).pack(fill="x", pady=(0, 8))
        
        # Header row
        header_row = ctk.CTkFrame(parent, fg_color=Colors.BG_ELEVATED, height=28)
        header_row.pack(fill="x", pady=(0, 4))
        header_row.pack_propagate(False)
        ctk.CTkLabel(header_row, text="Column", font=Fonts.BODY_BOLD,
                     text_color=Colors.TEXT_DARK, width=100, anchor="w").pack(side="left", padx=8)
        ctk.CTkLabel(header_row, text="Description", font=Fonts.BODY_BOLD,
                     text_color=Colors.TEXT_DARK, width=300, anchor="w").pack(side="left", padx=8)
        ctk.CTkLabel(header_row, text="Unit", font=Fonts.BODY_BOLD,
                     text_color=Colors.TEXT_DARK, width=80, anchor="w").pack(side="left", padx=8)
        
        # Column descriptions
        for col_name, (description, unit) in sorted(column_info.items()):
            row = ctk.CTkFrame(parent, fg_color="transparent")
            row.pack(fill="x", pady=1)
            ctk.CTkLabel(row, text=col_name, font=Fonts.BODY_BOLD,
                         text_color=Colors.TEXT_SECONDARY, width=100, anchor="w").pack(side="left", padx=8)
            ctk.CTkLabel(row, text=description, font=Fonts.BODY,
                         text_color=Colors.TEXT_PRIMARY, width=300, anchor="w").pack(side="left", padx=8)
            ctk.CTkLabel(row, text=unit, font=Fonts.BODY,
                         text_color=Colors.ACCENT, width=80, anchor="w").pack(side="left", padx=8)
    
    def _flatten_dict(self, d, prefix=""):
        """Flatten a nested dict into (key, value) pairs."""
        items = []
        if isinstance(d, dict):
            for k, v in d.items():
                full_key = f"{prefix}.{k}" if prefix else k
                if isinstance(v, dict):
                    items.extend(self._flatten_dict(v, full_key))
                else:
                    items.append((full_key, v))
        return items
    
    def _build_table_tab(self, parent, data, source):
        """Build the Data Table tab with scrollable table and aggregation options."""
        # Extract table data
        rows, columns = self._extract_table_data(data, source)
        
        if not rows:
            ctk.CTkLabel(parent, text="No tabular data available.",
                         font=Fonts.BODY, text_color=Colors.TEXT_MUTED).pack(pady=40)
            return
        
        # Aggregation mode selection frame
        agg_frame = ctk.CTkFrame(parent, fg_color=Colors.BG_CARD, corner_radius=8)
        agg_frame.pack(fill="x", pady=(0, 8))
        
        ctk.CTkLabel(agg_frame, text="View Mode:", font=Fonts.BODY_BOLD,
                     text_color=Colors.TEXT_DARK).pack(side="left", padx=(12, 8))
        
        # Variable to track aggregation mode
        agg_mode = ctk.StringVar(value="hourly")
        
        # Radio buttons for aggregation
        rb_frame = ctk.CTkFrame(agg_frame, fg_color="transparent")
        rb_frame.pack(side="left", fill="x", expand=True)
        
        ctk.CTkRadioButton(rb_frame, text="Hourly", variable=agg_mode, value="hourly",
                           font=Fonts.BODY, text_color=Colors.TEXT_PRIMARY,
                           command=lambda: self._refresh_table(table_container, rows, columns, agg_mode.get(), source)).pack(side="left", padx=8)
        ctk.CTkRadioButton(rb_frame, text="Daily", variable=agg_mode, value="daily",
                           font=Fonts.BODY, text_color=Colors.TEXT_PRIMARY,
                           command=lambda: self._refresh_table(table_container, rows, columns, agg_mode.get(), source)).pack(side="left", padx=8)
        ctk.CTkRadioButton(rb_frame, text="Monthly", variable=agg_mode, value="monthly",
                           font=Fonts.BODY, text_color=Colors.TEXT_PRIMARY,
                           command=lambda: self._refresh_table(table_container, rows, columns, agg_mode.get(), source)).pack(side="left", padx=8)
        ctk.CTkRadioButton(rb_frame, text="Yearly", variable=agg_mode, value="annual",
                           font=Fonts.BODY, text_color=Colors.TEXT_PRIMARY,
                           command=lambda: self._refresh_table(table_container, rows, columns, agg_mode.get(), source)).pack(side="left", padx=8)
        
        # Info bar
        info = ctk.CTkFrame(parent, fg_color="#1e1b4b", height=36, corner_radius=8)
        info.pack(fill="x", pady=(0, 8))
        info.pack_propagate(False)
        
        info_label = ctk.CTkLabel(info, text="",
                     font=Fonts.CAPTION, text_color="#c7d2fe")
        info_label.pack(side="left", padx=8)
        
        # Table container frame (to be refreshed)
        table_container = ctk.CTkFrame(parent, fg_color="transparent")
        table_container.pack(fill="both", expand=True)
        
        # Initial render
        self._refresh_table(table_container, rows, columns, "hourly", source, info_label)
    
    def _refresh_table(self, parent, rows, columns, mode, source, info_label=None):
        """Refresh the table with aggregated data."""
        # Clear existing content
        for child in parent.winfo_children():
            child.destroy()
        
        print(f"DEBUG _refresh_table: mode={mode}, source={source}, {len(rows)} input rows")
        
        # Aggregate data based on mode (hierarchical: hourly → daily → monthly → years)
        if mode == "hourly":
            display_rows = rows
            agg_columns = columns
        elif mode == "daily":
            display_rows, agg_columns = self._aggregate_daily(rows, columns, source)
        elif mode == "monthly":
            # First aggregate to daily, then to monthly
            daily_rows, daily_columns = self._aggregate_daily(rows, columns, source)
            display_rows, agg_columns = self._aggregate_monthly(daily_rows, daily_columns, source)
        else:  # annual (years)
            # First aggregate to daily, then monthly, then years
            daily_rows, daily_columns = self._aggregate_daily(rows, columns, source)
            monthly_rows, monthly_columns = self._aggregate_monthly(daily_rows, daily_columns, source)
            display_rows, agg_columns = self._aggregate_annual(monthly_rows, monthly_columns, source)
        
        # Handle empty aggregation results - fallback with warning
        if not display_rows and mode != "hourly":
            print(f"WARNING: Aggregation returned empty for mode={mode}, falling back to hourly")
            display_rows = rows
            agg_columns = columns
        
        # Update info bar
        print(f"DEBUG _refresh_table: {len(display_rows)} display rows, {len(agg_columns)} agg columns")
        if display_rows:
            print(f"DEBUG first display row: {display_rows[0]}")
        
        if info_label:
            showing = min(len(display_rows), 2000)
            mode_text = {"hourly": "Hourly", "daily": "Daily", "monthly": "Monthly", "annual": "Yearly"}[mode]
            info_text = f"  \u2139  {len(display_rows)} {mode_text.lower()} records  |  {len(agg_columns)} columns  |  Showing {showing} rows"
            info_label.configure(text=info_text)
        
        # Use ttk.Treeview for fast rendering
        import tkinter as tk
        from tkinter import ttk as ttk_mod
        
        # Table frame
        table_frame = ctk.CTkFrame(parent, fg_color="#f8fafc", corner_radius=8,
                                   border_width=1, border_color="#e2e8f0")
        table_frame.pack(fill="both", expand=True)
        
        # Column IDs: # + all data columns
        col_ids = ["#"] + list(agg_columns)
        
        tree = ttk_mod.Treeview(table_frame, columns=col_ids, show="headings",
                                style="DataTable.Treeview", selectmode="browse")
        
        # Configure columns (batch optimization)
        tree.heading("#", text="#", anchor="center")
        tree.column("#", width=60, minwidth=50, anchor="center", stretch=False)
        
        for col_name in agg_columns:
            tree.heading(col_name, text=str(col_name), anchor="w")
            tree.column(col_name, width=120, minwidth=80, anchor="w")
        
        # Alternating row colors
        tree.tag_configure("even", background="#ffffff")
        tree.tag_configure("odd", background="#f1f5f9")
        
        # Prepare data for faster insertion
        MAX_DISPLAY = 2000
        display_rows_limited = display_rows[:MAX_DISPLAY]
        
        # Pre-calculate formatted values
        formatted_values = []
        for idx, row_data in enumerate(display_rows_limited):
            vals = [idx + 1]
            for col_name in agg_columns:
                val = row_data.get(col_name, "")
                if isinstance(val, float):
                    val = f"{val:.4f}" if abs(val) < 100 else f"{val:.2f}"
                vals.append(val)
            formatted_values.append(vals)
        
        # Bulk insert
        for idx, vals in enumerate(formatted_values):
            tag = "even" if idx % 2 == 0 else "odd"
            tree.insert("", "end", iid=idx, values=vals, tags=(tag,))
        
        # Scrollbars
        v_scroll = ttk_mod.Scrollbar(table_frame, orient="vertical", command=tree.yview)
        h_scroll = ttk_mod.Scrollbar(table_frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=v_scroll.set, xscrollcommand=h_scroll.set)
        
        v_scroll.pack(side="right", fill="y")
        h_scroll.pack(side="bottom", fill="x")
        tree.pack(fill="both", expand=True)
        
        if len(display_rows) > MAX_DISPLAY:
            footer = ctk.CTkFrame(parent, fg_color="#fef3c7", height=28, corner_radius=6)
            footer.pack(fill="x", pady=(6, 0))
            footer.pack_propagate(False)
            ctk.CTkLabel(footer, text=f"  \u26a0  Showing {MAX_DISPLAY} of {len(display_rows)} records. Export to see all data.",
                         font=Fonts.CAPTION, text_color="#92400e").pack(side="left", padx=8)
    
    def _aggregate_monthly(self, rows, columns, source):
        """Aggregate daily data to monthly totals (power integrated from daily, others averaged)."""
        from collections import defaultdict
        from datetime import datetime
        
        # Check if input is already daily aggregated (has "Day" column)
        is_daily = len(rows) > 0 and "Day" in rows[0]
        
        # Group rows by month
        monthly_data = defaultdict(lambda: defaultdict(list))
        
        for row in rows:
            if is_daily:
                # Input is daily aggregated, extract month from Day (YYYY-MM-DD)
                day_key = row.get("Day", "")
                if day_key and len(day_key) >= 7:
                    month_key = day_key[:7]  # YYYY-MM
                else:
                    continue
            else:
                # Input is hourly, extract timestamp
                ts = self._extract_timestamp(row, source)
                if ts:
                    month_key = ts.strftime("%Y-%m")  # YYYY-MM format
                else:
                    continue
            
            for col, val in row.items():
                monthly_data[month_key][col].append(val)
        
        # Aggregate each month
        agg_rows = []
        power_cols = self._get_power_columns(columns)
        
        for month_key in sorted(monthly_data.keys()):
            month_data = monthly_data[month_key]
            agg_row = {"Month": month_key}
            
            for col in columns:
                if col == "time" or col == "timestamp" or col == "Day":
                    continue
                    
                values = month_data.get(col, [])
                numeric_values = [v for v in values if isinstance(v, (int, float))]
                
                if numeric_values:
                    if col in power_cols:
                        # Power: integrate (sum) from daily values
                        agg_row[col] = sum(numeric_values)
                    else:
                        # Other parameters: average
                        agg_row[col] = sum(numeric_values) / len(numeric_values)
                else:
                    agg_row[col] = values[0] if values else ""
            
            agg_rows.append(agg_row)
        
        # New columns with Month first
        agg_columns = ["Month"] + [c for c in columns if c not in ("time", "timestamp", "Day")]
        return agg_rows, agg_columns
    
    def _aggregate_daily(self, rows, columns, source):
        """Aggregate hourly data to daily totals (power integrated to energy 'E', others averaged)."""
        from collections import defaultdict
        
        # Reset debug counter
        self._debug_ts_count = 0
        
        # Debug: Check input data
        print(f"DEBUG _aggregate_daily: {len(rows)} rows, columns={columns}, source={source}")
        if rows:
            print(f"DEBUG first row: {rows[0]}")
        
        # Group rows by day
        daily_data = defaultdict(lambda: defaultdict(list))
        
        success_count = 0
        for row in rows:
            ts = self._extract_timestamp(row, source)
            if ts:
                day_key = ts.strftime("%Y-%m-%d")  # YYYY-MM-DD format
                for col, val in row.items():
                    daily_data[day_key][col].append(val)
                success_count += 1
        
        print(f"DEBUG _aggregate_daily: extracted {success_count} timestamps, {len(daily_data)} unique days")
        
        # Aggregate each day
        agg_rows = []
        power_cols = self._get_power_columns(columns)
        
        # Build mapping from power column names to energy column names (P -> E)
        energy_col_map = {}
        for col in power_cols:
            col_str = str(col)
            # Replace P/p with E/e for power-related columns
            if col_str.lower().startswith('p') and len(col_str) > 1 and col_str[1:].isdigit():
                # Pattern like P1, P2, p1, p2 -> E1, E2, e1, e2
                new_col = 'E' + col_str[1:] if col_str[0].isupper() else 'e' + col_str[1:]
            elif col_str == 'P' or col_str == 'p':
                new_col = 'E' if col_str == 'P' else 'e'
            else:
                # For other power columns, keep the name (energy integration keeps the concept)
                new_col = col_str
            energy_col_map[col] = new_col
        
        for day_key in sorted(daily_data.keys()):
            day_data = daily_data[day_key]
            agg_row = {"Day": day_key}
            
            for col in columns:
                if col == "time" or col == "timestamp":
                    continue
                    
                values = day_data.get(col, [])
                numeric_values = [v for v in values if isinstance(v, (int, float))]
                
                # Use energy column name for power columns
                output_col = energy_col_map.get(col, col)
                
                if numeric_values:
                    if col in power_cols:
                        # Power: integrate (sum of hourly P values = daily E energy)
                        agg_row[output_col] = sum(numeric_values)
                    else:
                        # Other parameters: average
                        agg_row[output_col] = sum(numeric_values) / len(numeric_values)
                else:
                    agg_row[output_col] = values[0] if values else ""
            
            agg_rows.append(agg_row)
        
        # New columns with Day first, power columns renamed to energy columns
        agg_columns = ["Day"]
        for c in columns:
            if c not in ("time", "timestamp"):
                agg_columns.append(energy_col_map.get(c, c))
        return agg_rows, agg_columns
    
    def _aggregate_annual(self, rows, columns, source):
        """Aggregate monthly data to annual totals (power integrated from monthly, others averaged)."""
        from collections import defaultdict
        
        # Check if input is already monthly aggregated (has "Month" column)
        is_monthly = len(rows) > 0 and "Month" in rows[0]
        
        # Group rows by year
        annual_data = defaultdict(lambda: defaultdict(list))
        
        for row in rows:
            if is_monthly:
                # Input is monthly aggregated, extract year from Month (YYYY-MM)
                month_key = row.get("Month", "")
                if month_key and len(month_key) >= 4:
                    year_key = month_key[:4]  # YYYY
                else:
                    continue
            else:
                # Input is hourly, extract timestamp
                ts = self._extract_timestamp(row, source)
                if ts:
                    year_key = ts.strftime("%Y")  # YYYY format
                else:
                    continue
            
            for col, val in row.items():
                annual_data[year_key][col].append(val)
        
        # Aggregate each year
        agg_rows = []
        power_cols = self._get_power_columns(columns)
        
        for year_key in sorted(annual_data.keys()):
            year_data = annual_data[year_key]
            agg_row = {"Year": year_key}
            
            for col in columns:
                if col == "time" or col == "timestamp" or col == "Month":
                    continue
                    
                values = year_data.get(col, [])
                numeric_values = [v for v in values if isinstance(v, (int, float))]
                
                if numeric_values:
                    if col in power_cols:
                        # Power: integrate (sum) from monthly values
                        agg_row[col] = sum(numeric_values)
                    else:
                        # Other parameters: average
                        agg_row[col] = sum(numeric_values) / len(numeric_values)
                else:
                    agg_row[col] = values[0] if values else ""
            
            agg_rows.append(agg_row)
        
        # New columns with Year first
        agg_columns = ["Year"] + [c for c in columns if c not in ("time", "timestamp", "Month")]
        return agg_rows, agg_columns
    
    def _extract_timestamp(self, row, source):
        """Extract datetime from row based on source type."""
        from datetime import datetime
        
        # Initialize debug counter if not exists
        if not hasattr(self, '_debug_ts_count'):
            self._debug_ts_count = 0
        
        # Try common timestamp column names
        ts_value = row.get("time") or row.get("timestamp") or row.get("datetime")
        
        if not ts_value:
            if self._debug_ts_count < 5:
                print(f"DEBUG _extract_timestamp: No timestamp found in row keys={list(row.keys())}")
                self._debug_ts_count += 1
            return None
        
        if self._debug_ts_count < 5:
            print(f"DEBUG _extract_timestamp: ts_value={ts_value} (type={type(ts_value).__name__})")
            self._debug_ts_count += 1
        
        try:
            # PVGIS format: "20190101:0009" (YYYYMMDD:HHMM)
            if isinstance(ts_value, str) and len(ts_value) >= 13 and ":" in ts_value:
                date_part = ts_value[:8]
                time_part = ts_value[9:13]
                return datetime.strptime(f"{date_part}{time_part}", "%Y%m%d%H%M")
            
            # Ninja format: milliseconds (string or int)
            # Handle both string digits and float strings
            if isinstance(ts_value, (int, float)):
                ts_ms = int(ts_value)
            elif isinstance(ts_value, str):
                # Try to parse as float first (handles "1234567890.0"), then as int
                try:
                    ts_ms = int(float(ts_value))
                except ValueError:
                    # Not a number, try ISO format below
                    return datetime.fromisoformat(ts_value.replace('Z', '+00:00'))
            else:
                return None
                
            # Check if milliseconds timestamp
            # Note: Unix timestamps in milliseconds for years 1970-2001 are between 0 and 1,000,000,000,000
            # So we need to check if the value would make sense as seconds vs milliseconds
            # A seconds timestamp > 1 billion would be year 2001+, which is also valid as milliseconds
            if ts_ms > 10000000000:  # > 10 billion = milliseconds (year 1970+)
                result = datetime.fromtimestamp(ts_ms / 1000.0)
                if self._debug_ts_count < 5:
                    print(f"DEBUG _extract_timestamp: Parsed as milliseconds -> {result}")
                    self._debug_ts_count += 1
                return result
            elif ts_ms > 1000000000:  # > 1 billion = seconds (year 2001+)
                result = datetime.fromtimestamp(ts_ms)
                if self._debug_ts_count < 5:
                    print(f"DEBUG _extract_timestamp: Parsed as seconds -> {result}")
                    self._debug_ts_count += 1
                return result
            
            # ISO format fallback for strings
            result = datetime.fromisoformat(str(ts_value).replace('Z', '+00:00'))
            if self._debug_ts_count < 5:
                print(f"DEBUG _extract_timestamp: Parsed as ISO -> {result}")
                self._debug_ts_count += 1
            return result
        except Exception as e:
            if self._debug_ts_count < 5:
                print(f"DEBUG _extract_timestamp: ERROR {e} for ts_value={ts_value}")
                self._debug_ts_count += 1
            return None
    
    def _get_power_columns(self, columns):
        """Identify power-related columns for integration."""
        power_keywords = ['power', 'energy', 'p_', 'e_', 'electricity', 'generation', 'output']
        power_cols = []
        
        for col in columns:
            col_lower = str(col).lower()
            is_power = False
            if any(kw in col_lower for kw in power_keywords):
                is_power = True
            # Check for single 'P' or 'p' column (PVGIS power column)
            if col_lower == 'p':
                is_power = True
            # Check for single 'E' or 'e' column, or E1, E2, etc. (energy from integrated power)
            if col_lower in ['e'] or (len(col_lower) > 1 and col_lower[0] == 'e' and col_lower[1:].isdigit()):
                is_power = True
            if is_power:
                power_cols.append(col)
        
        return power_cols
    
    def _build_export_tab(self, parent, data, source):
        """Build the Export tab with aggregation options."""
        rows, columns = self._extract_table_data(data, source)
        
        center = ctk.CTkFrame(parent, fg_color="transparent")
        center.pack(expand=True)
        
        # Icon / title
        ctk.CTkLabel(center, text="Export Data", font=Fonts.H2,
                     text_color=Colors.TEXT_DARK).pack(pady=(20, 4))
        
        # === CSV Export Section ===
        csv_frame = ctk.CTkFrame(center, fg_color=Colors.BG_CARD, corner_radius=8)
        csv_frame.pack(fill="x", pady=(12, 12), padx=40)
        
        ctk.CTkLabel(csv_frame, text="Export as CSV (.csv)", font=Fonts.BODY_BOLD,
                     text_color=Colors.TEXT_DARK).pack(anchor="w", padx=16, pady=(12, 8))
        
        # Checkboxes for CSV export modes - mutually exclusive (radio button behavior)
        checkbox_frame = ctk.CTkFrame(csv_frame, fg_color="transparent")
        checkbox_frame.pack(fill="x", padx=16, pady=(0, 12))
        
        # Variables for checkboxes - only one can be selected at a time
        export_hourly = ctk.BooleanVar(value=True)
        export_daily = ctk.BooleanVar(value=False)
        export_monthly = ctk.BooleanVar(value=False)
        export_years = ctk.BooleanVar(value=False)
        
        # Store all export variables for mutual exclusivity
        export_vars = [export_hourly, export_daily, export_monthly, export_years]
        
        def make_exclusive(selected_index):
            """Ensure only the selected checkbox is checked, others are unchecked."""
            for i, var in enumerate(export_vars):
                if i != selected_index:
                    var.set(False)
            # Ensure at least one is always selected
            if not any(var.get() for var in export_vars):
                export_vars[selected_index].set(True)
        
        ctk.CTkCheckBox(checkbox_frame, text="Hourly", variable=export_hourly,
                        command=lambda: make_exclusive(0),
                        font=Fonts.BODY, text_color=Colors.TEXT_PRIMARY).pack(side="left", padx=(0, 16))
        ctk.CTkCheckBox(checkbox_frame, text="Daily", variable=export_daily,
                        command=lambda: make_exclusive(1),
                        font=Fonts.BODY, text_color=Colors.TEXT_PRIMARY).pack(side="left", padx=(0, 16))
        ctk.CTkCheckBox(checkbox_frame, text="Monthly", variable=export_monthly,
                        command=lambda: make_exclusive(2),
                        font=Fonts.BODY, text_color=Colors.TEXT_PRIMARY).pack(side="left", padx=(0, 16))
        ctk.CTkCheckBox(checkbox_frame, text="Yearly", variable=export_years,
                        command=lambda: make_exclusive(3),
                        font=Fonts.BODY, text_color=Colors.TEXT_PRIMARY).pack(side="left", padx=(0, 16))
        
        # Export CSV button - store reference to prevent garbage collection issues
        self._csv_export_btn = ctk.CTkButton(
            csv_frame, text="Export Selected as CSV",
            font=Fonts.BODY_BOLD, height=42, width=250,
            fg_color="#10b981", hover_color="#059669",
            text_color="#ffffff", corner_radius=10
        )
        self._csv_export_btn.pack(pady=(0, 12))
        
        # Set command separately to avoid closure issues
        print("DEBUG: Defining export_csv_cmd function...")  # This should print once during setup
        def export_csv_cmd():
            print("="*50)
            print("DEBUG: CSV BUTTON WAS CLICKED - STARTING EXPORT")
            print("="*50)
            self._export_csv_multi(rows, columns, source, 
                                   export_hourly.get(), export_daily.get(), 
                                   export_monthly.get(), export_years.get())
        self._csv_export_btn.configure(command=export_csv_cmd)
        
        # === Excel Export Section ===
        excel_frame = ctk.CTkFrame(center, fg_color=Colors.BG_CARD, corner_radius=8)
        excel_frame.pack(fill="x", pady=(12, 12), padx=40)
        
        ctk.CTkLabel(excel_frame, text="Export as Excel (.xlsx)", font=Fonts.BODY_BOLD,
                     text_color=Colors.TEXT_DARK).pack(anchor="w", padx=16, pady=(12, 8))
        
        ctk.CTkLabel(excel_frame, text="Exports all 4 sheets: Hourly, Daily, Monthly, Yearly",
                     font=Fonts.CAPTION, text_color=Colors.TEXT_MUTED).pack(anchor="w", padx=16, pady=(0, 12))
        
        # Export Excel button - store reference
        self._excel_export_btn = ctk.CTkButton(
            excel_frame, text="Export All as Excel",
            font=Fonts.BODY_BOLD, height=42, width=250,
            fg_color="#6366f1", hover_color="#4f46e5",
            text_color="#ffffff", corner_radius=10
        )
        self._excel_export_btn.pack(pady=(0, 12))
        
        # Set command separately
        print("DEBUG: Defining export_excel_cmd function...")  # This should print once during setup
        def export_excel_cmd():
            print("="*50)
            print("DEBUG: EXCEL BUTTON WAS CLICKED - STARTING EXPORT")
            print("="*50)
            self._export_excel_multi(rows, columns, source)
        self._excel_export_btn.configure(command=export_excel_cmd)
        
        # === Metadata Export Section ===
        if source in ("tmy", "hourly"):
            meta = data.get("inputs", {})
        else:
            meta = data.get("metadata", {})
        
        ctk.CTkButton(
            center, text="Export Metadata as CSV",
            command=lambda: self._export_metadata(meta, source),
            font=Fonts.BODY_BOLD, height=48, width=280,
            fg_color=Colors.BG_ELEVATED, hover_color=Colors.BORDER_DEFAULT,
            text_color=Colors.TEXT_PRIMARY, corner_radius=10,
            border_width=1, border_color=Colors.BORDER_DEFAULT
        ).pack(pady=(12, 8))
    
    def _extract_table_data(self, data, source):
        """Extract rows and columns from API response data."""
        rows = []
        columns = []
        
        if source == "tmy":
            outputs = data.get("outputs", {})
            tmy_hourly = outputs.get("tmy_hourly", [])
            if tmy_hourly:
                rows = tmy_hourly
                columns = list(tmy_hourly[0].keys()) if tmy_hourly else []
        elif source == "hourly":
            outputs = data.get("outputs", {})
            hourly = outputs.get("hourly", [])
            if hourly:
                rows = hourly
                columns = list(hourly[0].keys()) if hourly else []
        elif source == "ninja":
            raw_data = data.get("data", {})
            if isinstance(raw_data, dict):
                # Ninja data is {timestamp: {field: value}, ...}
                for timestamp, values in raw_data.items():
                    if isinstance(values, dict):
                        row = {"timestamp": timestamp}
                        row.update(values)
                        rows.append(row)
                    else:
                        rows.append({"timestamp": timestamp, "value": values})
                if rows:
                    columns = list(rows[0].keys())
            elif isinstance(raw_data, list):
                rows = raw_data
                columns = list(raw_data[0].keys()) if raw_data else []
        
        return rows, columns
    
    def _export_data(self, rows, columns, fmt, source):
        """Export data as CSV or XLSX."""
        import tkinter.messagebox as mb
        
        if not rows:
            mb.showwarning("No Data", "No data to export.")
            return
        
        title_map = {"tmy": "pvgis_tmy", "hourly": "pvgis_hourly", "ninja": "ninja_wind"}
        default_name = title_map.get(source, "data")
        
        # Helper function to convert timestamp from milliseconds to datetime string (UTC)
        def format_timestamp(value):
            try:
                # Handle string timestamps (from JSON keys)
                if isinstance(value, str):
                    value = float(value)
                # Handle numeric timestamps (milliseconds)
                if isinstance(value, (int, float)) and value > 1000000000000:  # Likely milliseconds
                    return datetime.fromtimestamp(value / 1000.0, tz=timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
            except (ValueError, TypeError, OSError, OverflowError):
                pass
            return value
        
        # Prepare rows with formatted timestamps
        formatted_rows = []
        has_timestamp = "timestamp" in columns
        for row in rows:
            new_row = dict(row)
            if has_timestamp and "timestamp" in new_row:
                new_row["timestamp"] = format_timestamp(new_row["timestamp"])
            formatted_rows.append(new_row)
        
        if fmt == "csv":
            filepath = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
                initialfile=f"{default_name}_export.csv"
            )
            if not filepath:
                return
            try:
                with open(filepath, "w", newline="", encoding="utf-8") as f:
                    writer = csv.DictWriter(f, fieldnames=columns)
                    writer.writeheader()
                    writer.writerows(formatted_rows)
            except PermissionError:
                mb.showerror("Permission Denied", 
                             f"Cannot write to:\n{filepath}\n\nThe file may be open in another program (e.g., Excel).\nPlease close the file and try again.")
                return
        
        elif fmt == "xlsx":
            try:
                import openpyxl
            except ImportError:
                # Fallback: save as CSV if openpyxl not installed
                mb.showinfo("openpyxl not found",
                            "openpyxl is not installed. Saving as CSV instead.\nInstall with: pip install openpyxl")
                self._export_data(rows, columns, "csv", source)
                return
            
            filepath = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=f"{default_name}_export.xlsx"
            )
            if not filepath:
                return
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = default_name
            # Header
            for col_idx, col_name in enumerate(columns, start=1):
                cell = ws.cell(row=1, column=col_idx, value=str(col_name))
                cell.font = openpyxl.styles.Font(bold=True)
            # Data
            for row_idx, row_data in enumerate(formatted_rows, start=2):
                for col_idx, col_name in enumerate(columns, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=row_data.get(col_name, ""))
            try:
                wb.save(filepath)
            except PermissionError:
                mb.showerror("Permission Denied", 
                             f"Cannot write to:\n{filepath}\n\nThe file may be open in another program (e.g., Excel).\nPlease close the file and try again.")
                return
        
        mb.showinfo("Export Complete", f"Data exported successfully to:\n{filepath}")
    
    def _export_data_with_mode(self, rows, columns, fmt, source, mode):
        """Export data with aggregation mode (hourly/daily/monthly/annual)."""
        # Aggregate data if needed (hierarchical: hourly → daily → monthly → years)
        if mode == "daily":
            agg_rows, agg_columns = self._aggregate_daily(rows, columns, source)
        elif mode == "monthly":
            # First aggregate to daily, then to monthly
            daily_rows, daily_columns = self._aggregate_daily(rows, columns, source)
            agg_rows, agg_columns = self._aggregate_monthly(daily_rows, daily_columns, source)
        elif mode == "annual":
            # First aggregate to daily, then monthly, then years
            daily_rows, daily_columns = self._aggregate_daily(rows, columns, source)
            monthly_rows, monthly_columns = self._aggregate_monthly(daily_rows, daily_columns, source)
            agg_rows, agg_columns = self._aggregate_annual(monthly_rows, monthly_columns, source)
        else:
            agg_rows, agg_columns = rows, columns
        
        # Call regular export with aggregated data
        self._export_data(agg_rows, agg_columns, fmt, source)
    
    def _export_csv_multi(self, rows, columns, source, hourly, daily, monthly, years):
        """Export selected aggregation modes as separate CSV files."""
        import tkinter.messagebox as mb
        import os
        
        # DEBUG: Confirm CSV function is called
        print(f"DEBUG: _export_csv_multi called with hourly={hourly}, daily={daily}, monthly={monthly}, years={years}")
        
        # Check if at least one mode is selected
        if not any([hourly, daily, monthly, years]):
            mb.showwarning("No Selection", "Please select at least one export mode.")
            return
        
        title_map = {"tmy": "pvgis_tmy", "hourly": "pvgis_hourly", "ninja": "ninja_wind"}
        base_name = title_map.get(source, "data")
        
        # Ask for directory to save files
        directory = filedialog.askdirectory(title="Select folder to save CSV files")
        if not directory:
            return
        
        exported_files = []
        
        try:
            # Export Hourly
            if hourly:
                filepath = os.path.join(directory, f"{base_name}_hourly.csv")
                self._export_data_to_path(rows, columns, filepath)
                exported_files.append(f"{base_name}_hourly.csv")
            
            # Export Daily
            if daily:
                daily_rows, daily_columns = self._aggregate_daily(rows, columns, source)
                filepath = os.path.join(directory, f"{base_name}_daily.csv")
                self._export_data_to_path(daily_rows, daily_columns, filepath)
                exported_files.append(f"{base_name}_daily.csv")
            
            # Export Monthly
            if monthly:
                daily_rows, daily_columns = self._aggregate_daily(rows, columns, source)
                monthly_rows, monthly_columns = self._aggregate_monthly(daily_rows, daily_columns, source)
                filepath = os.path.join(directory, f"{base_name}_monthly.csv")
                self._export_data_to_path(monthly_rows, monthly_columns, filepath)
                exported_files.append(f"{base_name}_monthly.csv")
            
            # Export Years
            if years:
                daily_rows, daily_columns = self._aggregate_daily(rows, columns, source)
                monthly_rows, monthly_columns = self._aggregate_monthly(daily_rows, daily_columns, source)
                years_rows, years_columns = self._aggregate_annual(monthly_rows, monthly_columns, source)
                filepath = os.path.join(directory, f"{base_name}_years.csv")
                self._export_data_to_path(years_rows, years_columns, filepath)
                exported_files.append(f"{base_name}_years.csv")
            
            mb.showinfo("Export Complete", f"Exported {len(exported_files)} file(s):\n" + "\n".join(exported_files))
        
        except Exception as e:
            mb.showerror("Export Error", f"Error exporting CSV files:\n{str(e)}")
    
    def _export_excel_multi(self, rows, columns, source):
        """Export all aggregation modes as separate sheets in one Excel file."""
        import tkinter.messagebox as mb
        
        # DEBUG: Confirm Excel function is called
        print("DEBUG: _export_excel_multi called")
        
        title_map = {"tmy": "pvgis_tmy", "hourly": "pvgis_hourly", "ninja": "ninja_wind"}
        default_name = title_map.get(source, "data")
        
        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=f"{default_name}_all.xlsx"
        )
        if not filepath:
            return
        
        try:
            # Check if openpyxl is available
            try:
                from openpyxl import Workbook
                from openpyxl.utils.dataframe import dataframe_to_rows
                import pandas as pd
            except ImportError:
                mb.showerror("Excel Export - Missing Dependency", 
                    "openpyxl and/or pandas not installed.\n"
                    "Install with: pip install openpyxl pandas\n\n"
                    "Note: This error should only appear for Excel export, not CSV export.")
                return
            
            # Create workbook
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            # Helper function to add sheet
            def add_sheet(wb, sheet_name, data_rows, data_columns):
                ws = wb.create_sheet(title=sheet_name)
                
                # Write header
                ws.append(list(data_columns))
                
                # Write data rows
                for row_data in data_rows:
                    row_values = [row_data.get(col, "") for col in data_columns]
                    ws.append(row_values)
            
            # Add Hourly sheet
            add_sheet(wb, "Hourly", rows, columns)
            
            # Add Daily sheet
            daily_rows, daily_columns = self._aggregate_daily(rows, columns, source)
            add_sheet(wb, "Daily", daily_rows, daily_columns)
            
            # Add Monthly sheet
            monthly_rows, monthly_columns = self._aggregate_monthly(daily_rows, daily_columns, source)
            add_sheet(wb, "Monthly", monthly_rows, monthly_columns)
            
            # Add Years sheet
            years_rows, years_columns = self._aggregate_annual(monthly_rows, monthly_columns, source)
            add_sheet(wb, "Yearly", years_rows, years_columns)
            
            # Save workbook
            wb.save(filepath)
            mb.showinfo("Export Complete", f"Excel file exported to:\n{filepath}")
        
        except Exception as e:
            mb.showerror("Export Error", f"Error exporting Excel file:\n{str(e)}")
    
    def _export_data_to_path(self, rows, columns, filepath):
        """Export data to a specific file path (CSV format)."""
        with open(filepath, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(list(columns))
            for row in rows:
                row_values = [row.get(col, "") for col in columns]
                writer.writerow(row_values)
    
    def _export_metadata(self, meta, source):
        """Export metadata as CSV."""
        title_map = {"tmy": "pvgis_tmy", "hourly": "pvgis_hourly", "ninja": "ninja_wind"}
        default_name = title_map.get(source, "data")
        
        filepath = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
            initialfile=f"{default_name}_metadata.csv"
        )
        if not filepath:
            return
        
        items = self._flatten_dict(meta)
        with open(filepath, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["Parameter", "Value"])
            for key, value in items:
                writer.writerow([key, value])
        
        import tkinter.messagebox as mb
        mb.showinfo("Export Complete", f"Metadata exported to:\n{filepath}")

    # ── Graphs Tab and Dialog Methods ─────────────────────────────────────
    
    def _build_energy_graphs_tab(self, parent, hourly_results, selected_model, turbine_specs, Cp, A, n_hours):
        """Build the Graphs tab content with button to open graph selection dialogs."""
        import customtkinter as ctk
        import tkinter.messagebox as mb
        
        # Clear parent
        for child in parent.winfo_children():
            child.destroy()
        
        # Create scrollable frame to ensure all content is visible
        scroll_frame = ctk.CTkScrollableFrame(
            parent,
            fg_color="transparent",
            scrollbar_button_color=Colors.BORDER_DEFAULT,
            scrollbar_button_hover_color=Colors.BORDER_BRIGHT
        )
        scroll_frame.pack(fill="both", expand=True, padx=Spacing.PAD_MD, pady=Spacing.PAD_MD)
        
        # Main container inside scrollable frame
        main_frame = ctk.CTkFrame(scroll_frame, fg_color="transparent")
        main_frame.pack(fill="both", expand=True)
        
        # Title
        ctk.CTkLabel(main_frame, text="Data Visualization", font=Fonts.H2,
                     text_color=Colors.TEXT_DARK).pack(anchor="w", pady=(0, Spacing.PAD_SM))
        
        ctk.CTkLabel(main_frame, 
                     text="Generate graphs from the calculated energy data. Select time period, model, and graph type.",
                     font=Fonts.BODY, text_color=Colors.TEXT_SECONDARY,
                     wraplength=700).pack(anchor="w", pady=(0, Spacing.PAD_LG))
        
        # Info card showing current data status
        info_card = ctk.CTkFrame(main_frame, fg_color=Colors.BG_CARD,
                                 corner_radius=Spacing.RADIUS_MD,
                                 border_width=1, border_color=Colors.BORDER_SUBTLE)
        info_card.pack(fill="x", pady=(0, Spacing.PAD_LG))
        
        info_inner = ctk.CTkFrame(info_card, fg_color="transparent")
        info_inner.pack(fill="x", padx=Spacing.PAD_LG, pady=Spacing.PAD_LG)
        
        ctk.CTkLabel(info_inner, text="Current Calculation Data", font=Fonts.H3,
                     text_color=Colors.TEXT_DARK).pack(anchor="w", pady=(0, Spacing.PAD_SM))
        
        # Data summary
        data_info = [
            ("Selected Model", selected_model),
            ("Total Records", f"{n_hours} hours"),
            ("Turbine", turbine_specs.get("name", "Unknown")),
            ("Rated Power", f"{turbine_specs.get('rated_power_kw', 0)} kW"),
        ]
        
        for key, value in data_info:
            row = ctk.CTkFrame(info_inner, fg_color="transparent")
            row.pack(fill="x", pady=2)
            ctk.CTkLabel(row, text=key, font=Fonts.BODY_BOLD,
                         text_color=Colors.TEXT_SECONDARY, width=150,
                         anchor="w").pack(side="left")
            ctk.CTkLabel(row, text=str(value), font=Fonts.BODY,
                         text_color=Colors.TEXT_PRIMARY,
                         anchor="w").pack(side="left")
        
        # Center button to start graph selection
        button_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        button_frame.pack(fill="x", pady=(Spacing.PAD_LG, Spacing.PAD_LG))
        
        ctk.CTkLabel(button_frame, text="Click below to start graph creation",
                     font=Fonts.BODY, text_color=Colors.TEXT_MUTED).pack(pady=(0, Spacing.PAD_MD))
        
        open_dialogs_btn = ctk.CTkButton(
            button_frame,
            text="  📊 Select Graph Options  ",
            font=Fonts.BODY_BOLD,
            fg_color=Colors.ACCENT,
            hover_color=Colors.ACCENT_DIM,
            text_color=Colors.TEXT_ON_ACCENT,
            corner_radius=Spacing.RADIUS_MD,
            height=50,
            command=lambda: self._show_graph_time_selection_dialog(hourly_results, selected_model, turbine_specs)
        )
        open_dialogs_btn.pack()
    
    def _show_graph_time_selection_dialog(self, hourly_results, selected_model, turbine_specs, current_time_period=None):
        """Show dialog to select time period (Hourly/Daily/Monthly)."""
        import customtkinter as ctk
        import tkinter.messagebox as mb
        
        dialog = ctk.CTkToplevel(self)
        dialog.title("Select Time Period")
        dialog.geometry("400x350")
        dialog.transient(self.winfo_toplevel())
        dialog.grab_set()
        dialog.configure(fg_color=Colors.BG_CARD)
        
        # Center the dialog
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() - 400) // 2
        y = (dialog.winfo_screenheight() - 350) // 2
        dialog.geometry(f"400x350+{x}+{y}")
        
        # Header
        header = ctk.CTkFrame(dialog, fg_color=Colors.BG_ELEVATED, height=50, corner_radius=0)
        header.pack(fill="x")
        header.pack_propagate(False)
        ctk.CTkFrame(header, width=4, height=24, fg_color=Colors.ACCENT).pack(side="left", padx=(16, 8), pady=13)
        ctk.CTkLabel(header, text="Select Time Period", font=Fonts.H3,
                     text_color=Colors.TEXT_DARK).pack(side="left", pady=13)
        
        # Create scrollable frame for content
        scroll_frame = ctk.CTkScrollableFrame(
            dialog,
            fg_color="transparent",
            scrollbar_button_color=Colors.BORDER_DEFAULT,
            scrollbar_button_hover_color=Colors.BORDER_BRIGHT
        )
        scroll_frame.pack(fill="both", expand=True, padx=Spacing.PAD_LG, pady=Spacing.PAD_LG)
        
        ctk.CTkLabel(scroll_frame, text="Choose the time aggregation for your graph:",
                     font=Fonts.BODY, text_color=Colors.TEXT_SECONDARY).pack(anchor="w", pady=(0, Spacing.PAD_MD))
        
        # Time period variable - use current selection if provided
        default_time = current_time_period if current_time_period else "hourly"
        time_var = ctk.StringVar(value=default_time)
        
        # Checkboxes frame
        checkbox_frame = ctk.CTkFrame(scroll_frame, fg_color=Colors.BG_ELEVATED,
                                      corner_radius=Spacing.RADIUS_MD)
        checkbox_frame.pack(fill="x", pady=Spacing.PAD_MD)
        
        checkbox_inner = ctk.CTkFrame(checkbox_frame, fg_color="transparent")
        checkbox_inner.pack(fill="x", padx=Spacing.PAD_MD, pady=Spacing.PAD_MD)
        
        # Time period options
        time_options = [
            ("Hourly", "hourly", "High-resolution data for each hour"),
            ("Daily", "daily", "Daily aggregated values"),
            ("Monthly", "monthly", "Monthly aggregated values"),
        ]
        
        for label, value, description in time_options:
            row = ctk.CTkFrame(checkbox_inner, fg_color="transparent")
            row.pack(fill="x", pady=Spacing.PAD_SM)
            
            rb = ctk.CTkRadioButton(row, text=label, variable=time_var, value=value,
                                    font=Fonts.BODY_BOLD, text_color=Colors.TEXT_PRIMARY)
            rb.pack(side="left")
            
            ctk.CTkLabel(row, text=f"  ({description})", font=Fonts.CAPTION,
                         text_color=Colors.TEXT_MUTED).pack(side="left")
        
        # Buttons frame (outside scroll area, fixed at bottom)
        btn_frame = ctk.CTkFrame(dialog, fg_color="transparent")
        btn_frame.pack(fill="x", padx=Spacing.PAD_LG, pady=(0, Spacing.PAD_LG))
        
        def on_cancel():
            dialog.destroy()
        
        def on_next():
            selected_time = time_var.get()
            dialog.destroy()
            self._show_graph_model_selection_dialog(hourly_results, selected_model, 
                                                     turbine_specs, selected_time)
        
        ctk.CTkButton(btn_frame, text="Cancel", font=Fonts.BODY,
                      fg_color=Colors.BG_HOVER, hover_color=Colors.BORDER_DEFAULT,
                      text_color=Colors.TEXT_PRIMARY, width=100,
                      command=on_cancel).pack(side="left")
        
        ctk.CTkButton(btn_frame, text="Next →", font=Fonts.BODY_BOLD,
                      fg_color=Colors.ACCENT, hover_color=Colors.ACCENT_DIM,
                      text_color=Colors.TEXT_ON_ACCENT, width=100,
                      command=on_next).pack(side="right")
    
    def _show_graph_model_selection_dialog(self, hourly_results, selected_model, turbine_specs, time_period, current_graph_model=None):
        """Show dialog to select Model 1 or Model 2."""
        import customtkinter as ctk
        
        dialog = ctk.CTkToplevel(self)
        dialog.title("Select Model")
        dialog.geometry("400x400")
        dialog.transient(self.winfo_toplevel())
        dialog.grab_set()
        dialog.configure(fg_color=Colors.BG_CARD)
        
        # Center the dialog
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() - 400) // 2
        y = (dialog.winfo_screenheight() - 400) // 2
        dialog.geometry(f"400x400+{x}+{y}")
        
        # Header
        header = ctk.CTkFrame(dialog, fg_color=Colors.BG_ELEVATED, height=50, corner_radius=0)
        header.pack(fill="x")
        header.pack_propagate(False)
        ctk.CTkFrame(header, width=4, height=24, fg_color=Colors.ACCENT).pack(side="left", padx=(16, 8), pady=13)
        ctk.CTkLabel(header, text="Select Model", font=Fonts.H3,
                     text_color=Colors.TEXT_DARK).pack(side="left", pady=13)
        
        # Create scrollable frame for content
        scroll_frame = ctk.CTkScrollableFrame(
            dialog,
            fg_color="transparent",
            scrollbar_button_color=Colors.BORDER_DEFAULT,
            scrollbar_button_hover_color=Colors.BORDER_BRIGHT
        )
        scroll_frame.pack(fill="both", expand=True, padx=Spacing.PAD_LG, pady=Spacing.PAD_LG)
        
        ctk.CTkLabel(scroll_frame, text=f"Time Period: {time_period.capitalize()}",
                     font=Fonts.BODY_BOLD, text_color=Colors.ACCENT).pack(anchor="w", pady=(0, Spacing.PAD_SM))
        
        ctk.CTkLabel(scroll_frame, text="Choose the calculation model for graph generation:",
                     font=Fonts.BODY, text_color=Colors.TEXT_SECONDARY).pack(anchor="w", pady=(0, Spacing.PAD_MD))
        
        # Scrollable list of models
        list_frame = ctk.CTkScrollableFrame(scroll_frame, fg_color=Colors.BG_ELEVATED,
                                            corner_radius=Spacing.RADIUS_MD,
                                            height=150)
        list_frame.pack(fill="x", pady=Spacing.PAD_MD)
        
        # Model variable - use current selection if provided, otherwise fall back to original selected_model
        default_model = current_graph_model if current_graph_model else (selected_model if selected_model else "Model 1")
        model_var = ctk.StringVar(value=default_model)
        
        # Model options with descriptions
        models = [
            ("Model 1", "Standard wind power equation (P = ½·ρ·A·Cp·V³)"),
            ("Model 2", "Cubic power curve model with cut-in/rated speeds"),
        ]
        
        for model_name, description in models:
            row = ctk.CTkFrame(list_frame, fg_color="transparent")
            row.pack(fill="x", pady=Spacing.PAD_SM)
            
            rb = ctk.CTkRadioButton(row, text=model_name, variable=model_var, value=model_name,
                                    font=Fonts.BODY_BOLD, text_color=Colors.TEXT_PRIMARY)
            rb.pack(anchor="w")
            
            ctk.CTkLabel(row, text=f"  {description}", font=Fonts.CAPTION,
                         text_color=Colors.TEXT_MUTED).pack(anchor="w", padx=(28, 0))
        
        # Buttons frame (outside scroll area, fixed at bottom)
        btn_frame = ctk.CTkFrame(dialog, fg_color="transparent")
        btn_frame.pack(fill="x", padx=Spacing.PAD_LG, pady=(0, Spacing.PAD_LG))
        
        def on_cancel():
            dialog.destroy()
        
        def on_next():
            selected_model_for_graph = model_var.get()
            dialog.destroy()
            self._show_graph_type_selection_dialog(hourly_results, selected_model_for_graph, 
                                                    turbine_specs, time_period)
        
        ctk.CTkButton(btn_frame, text="← Back", font=Fonts.BODY,
                      fg_color=Colors.BG_HOVER, hover_color=Colors.BORDER_DEFAULT,
                      text_color=Colors.TEXT_PRIMARY, width=100,
                      command=lambda: [dialog.destroy(), 
                                       self._show_graph_time_selection_dialog(hourly_results, selected_model, turbine_specs, time_period)]
                      ).pack(side="left")
        
        ctk.CTkButton(btn_frame, text="Next →", font=Fonts.BODY_BOLD,
                      fg_color=Colors.ACCENT, hover_color=Colors.ACCENT_DIM,
                      text_color=Colors.TEXT_ON_ACCENT, width=100,
                      command=on_next).pack(side="right")
    
    def _show_graph_type_selection_dialog(self, hourly_results, selected_model, turbine_specs, time_period, current_graph_type=None):
        """Show professional dialog to select graph type with history panel."""
        import customtkinter as ctk
        
        # Initialize graph history if not exists
        if not hasattr(self, '_graph_history'):
            self._graph_history = []
        
        # Release any existing grabs to avoid input conflicts
        try:
            self.winfo_toplevel().grab_release()
        except:
            pass
        
        print(f"Opening graph type selection dialog (history: {len(self._graph_history)} items)")
        
        dialog = ctk.CTkToplevel(self)
        dialog.title("Graph Visualization Studio")
        dialog.geometry("950x750")  # Optimized height to show all elements
        dialog.transient(self.winfo_toplevel())
        dialog.grab_set()
        dialog.configure(fg_color=Colors.BG_ROOT)
        
        # Center the dialog
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() - 950) // 2
        y = (dialog.winfo_screenheight() - 750) // 2
        dialog.geometry(f"950x750+{x}+{y}")
        
        # ========== HEADER ==========
        header = ctk.CTkFrame(dialog, fg_color=Colors.BG_CARD, height=60, corner_radius=0)
        header.pack(fill="x")
        header.pack_propagate(False)
        
        header_inner = ctk.CTkFrame(header, fg_color="transparent")
        header_inner.pack(fill="both", expand=True, padx=Spacing.PAD_LG, pady=Spacing.PAD_MD)
        
        # Icon + Title
        title_row = ctk.CTkFrame(header_inner, fg_color="transparent")
        title_row.pack(side="left", fill="y")
        
        icon_frame = ctk.CTkFrame(title_row, width=36, height=36, fg_color=Colors.ACCENT,
                                  corner_radius=Spacing.RADIUS_SM)
        icon_frame.pack(side="left", padx=(0, Spacing.PAD_SM))
        icon_frame.pack_propagate(False)
        ctk.CTkLabel(icon_frame, text="📊", font=(Fonts.FAMILY, 16)).place(relx=0.5, rely=0.5, anchor="center")
        
        ctk.CTkLabel(title_row, text="Graph Visualization Studio", font=Fonts.H2,
                     text_color=Colors.TEXT_DARK).pack(side="left")
        
        # Current configuration badge
        config_text = f"{time_period.capitalize()} • {selected_model}"
        config_badge = ctk.CTkFrame(header_inner, fg_color=Colors.ACCENT_BG,
                                    corner_radius=Spacing.RADIUS_MD)
        config_badge.pack(side="right")
        ctk.CTkLabel(config_badge, text=config_text, font=Fonts.CAPTION,
                     text_color=Colors.ACCENT).pack(padx=Spacing.PAD_MD, pady=Spacing.PAD_XS)
        
        # ========== MAIN CONTENT ==========
        main_frame = ctk.CTkFrame(dialog, fg_color="transparent")
        main_frame.pack(fill="both", expand=True, padx=Spacing.PAD_LG, pady=Spacing.PAD_LG)
        main_frame.grid_rowconfigure(0, weight=1)  # Row expands vertically
        main_frame.grid_columnconfigure(0, weight=3)  # Left panel - graph selection
        main_frame.grid_columnconfigure(1, weight=2)  # Right panel - history
        
        # ========== LEFT PANEL: Graph Selection ==========
        left_panel = ctk.CTkFrame(main_frame, fg_color=Colors.BG_CARD,
                                  corner_radius=Spacing.RADIUS_LG,
                                  border_width=1, border_color=Colors.BORDER_SUBTLE)
        left_panel.grid(row=0, column=0, sticky="nsew", padx=(0, Spacing.PAD_MD))
        left_panel.grid_rowconfigure(0, weight=0)  # Header - fixed
        left_panel.grid_rowconfigure(1, weight=0)  # Scrollable area - fixed height
        left_panel.grid_rowconfigure(2, weight=0)  # Action buttons - fixed at bottom
        left_panel.grid_columnconfigure(0, weight=1)
        
        # Header section (fixed)
        left_header = ctk.CTkFrame(left_panel, fg_color="transparent")
        left_header.grid(row=0, column=0, sticky="ew", padx=Spacing.PAD_LG, pady=(Spacing.PAD_LG, Spacing.PAD_MD))
        
        # Section title
        ctk.CTkLabel(left_header, text="Select Graph Type", font=Fonts.H3,
                     text_color=Colors.TEXT_DARK).pack(anchor="w")
        ctk.CTkLabel(left_header, text="Choose a visualization to generate",
                     font=Fonts.CAPTION, text_color=Colors.TEXT_SECONDARY).pack(anchor="w")
        
        # Graph type variable
        default_graph = current_graph_type if current_graph_type else ""
        graph_var = ctk.StringVar(value=default_graph)
        
        # Scrollable graph options container - fixed height so buttons stay visible
        graph_scroll = ctk.CTkScrollableFrame(left_panel, fg_color="transparent",
                                              height=350,  # Fixed height for 4 cards
                                              scrollbar_button_color=Colors.BORDER_DEFAULT,
                                              scrollbar_button_hover_color=Colors.BORDER_BRIGHT)
        graph_scroll.grid(row=1, column=0, sticky="nsew", padx=Spacing.PAD_LG, pady=(0, Spacing.PAD_SM))
        
        # Define graph types with icons
        if selected_model == "Model 1":
            power_energy_label = "Power (kW)" if time_period == "hourly" else "Energy (kWh)"
            graph_types = [
                ("Temperature", "Air temperature over time", "🌡️", "°C"),
                ("Wind", "Wind speed over time", "💨", "m/s"),
                ("Air Density", "Air density over time", "🌊", "kg/m³"),
                ("ρ/ρ₀", "Air density ratio over time", "📊", "-"),
                (power_energy_label, f"{power_energy_label} over time", "⚡", "kW" if time_period == "hourly" else "kWh"),
            ]
        else:
            power_energy_label = "Power P (kW)" if time_period == "hourly" else "Energy E (kWh)"
            graph_types = [
                ("V (Wind Speed)", "Wind speed V over time", "💨", "m/s"),
                ("Air Density", "Air density over time", "🌊", "kg/m³"),
                ("ρ/ρ₀", "Air density ratio over time", "📊", "-"),
                (power_energy_label, f"{power_energy_label} over time", "⚡", "kW" if time_period == "hourly" else "kWh"),
            ]
        
        graph_buttons = {}  # Store buttons for styling updates
        
        def select_graph_type(graph_name):
            graph_var.set(graph_name)
            # Update button styles
            for name, btn in graph_buttons.items():
                if name == graph_name:
                    btn.configure(fg_color=Colors.ACCENT_BG, border_color=Colors.ACCENT)
                else:
                    btn.configure(fg_color=Colors.BG_ELEVATED, border_color=Colors.BORDER_SUBTLE)
        
        # Debug: Print how many graphs we're creating
        print(f"Creating {len(graph_types)} graph type buttons: {[g[0] for g in graph_types]}")
        
        for i, (graph_name, description, icon, unit) in enumerate(graph_types):
            # Card-style button
            card = ctk.CTkFrame(graph_scroll, fg_color=Colors.BG_ELEVATED,
                               corner_radius=Spacing.RADIUS_MD,
                               border_width=2, border_color=Colors.BORDER_SUBTLE,
                               height=75)  # Fixed height for each card
            card.pack(fill="x", pady=4)
            card.pack_propagate(False)  # Prevent card from shrinking
            card.bind("<Button-1>", lambda e, g=graph_name: select_graph_type(g))
            
            card_inner = ctk.CTkFrame(card, fg_color="transparent")
            card_inner.pack(fill="x", padx=Spacing.PAD_MD, pady=Spacing.PAD_MD)
            
            # Radio button
            rb = ctk.CTkRadioButton(card_inner, text="", variable=graph_var, value=graph_name,
                                    fg_color=Colors.ACCENT, hover_color=Colors.ACCENT_DIM,
                                    command=lambda g=graph_name: select_graph_type(g))
            rb.pack(side="left", padx=(0, Spacing.PAD_SM))
            
            # Icon
            icon_label = ctk.CTkLabel(card_inner, text=icon, font=(Fonts.FAMILY, 20))
            icon_label.pack(side="left", padx=(0, Spacing.PAD_SM))
            
            # Text content
            text_frame = ctk.CTkFrame(card_inner, fg_color="transparent")
            text_frame.pack(side="left", fill="y", expand=True)
            
            ctk.CTkLabel(text_frame, text=graph_name, font=Fonts.BODY_BOLD,
                        text_color=Colors.TEXT_PRIMARY).pack(anchor="w")
            ctk.CTkLabel(text_frame, text=f"{description} ({unit})", font=Fonts.CAPTION,
                        text_color=Colors.TEXT_MUTED).pack(anchor="w")
            
            graph_buttons[graph_name] = card
            
            # Set initial selection style
            if graph_name == default_graph:
                card.configure(fg_color=Colors.ACCENT_BG, border_color=Colors.ACCENT)
        
        # Force the scrollable frame to update its scroll region
        graph_scroll.update_idletasks()
        
        # ========== LEFT PANEL: ACTION BUTTONS (Generate button here!) ==========
        actions_card = ctk.CTkFrame(left_panel, fg_color=Colors.BG_CARD,
                                    corner_radius=Spacing.RADIUS_LG,
                                    border_width=1, border_color=Colors.BORDER_SUBTLE,
                                    height=130)  # Fixed height for buttons
        actions_card.grid(row=2, column=0, sticky="ew", padx=Spacing.PAD_LG, pady=(Spacing.PAD_SM, 0))
        actions_card.grid_propagate(False)  # Keep fixed height
        
        actions_inner = ctk.CTkFrame(actions_card, fg_color="transparent")
        actions_inner.pack(fill="x", padx=Spacing.PAD_MD, pady=Spacing.PAD_MD)
        
        # ========== RIGHT PANEL: History & Actions ==========
        right_panel = ctk.CTkFrame(main_frame, fg_color="transparent")
        right_panel.grid(row=0, column=1, sticky="nsew")
        right_panel.grid_rowconfigure(0, weight=1, minsize=350)  # History section - larger minimum
        right_panel.grid_columnconfigure(0, weight=1)  # Column expands horizontally
        
        # History section
        history_card = ctk.CTkFrame(right_panel, fg_color=Colors.BG_CARD,
                                    corner_radius=Spacing.RADIUS_LG,
                                    border_width=1, border_color=Colors.BORDER_SUBTLE)
        history_card.grid(row=0, column=0, sticky="nsew")
        
        history_inner = ctk.CTkFrame(history_card, fg_color="transparent")
        history_inner.pack(fill="both", expand=True, padx=Spacing.PAD_MD, pady=Spacing.PAD_MD)
        
        # History header
        history_header = ctk.CTkFrame(history_inner, fg_color="transparent")
        history_header.pack(fill="x", pady=(0, Spacing.PAD_SM))
        
        ctk.CTkLabel(history_header, text="Generated Graphs", font=Fonts.H3,
                     text_color=Colors.TEXT_DARK).pack(side="left")
        
        # Clear history button
        if self._graph_history:
            clear_btn = ctk.CTkButton(history_header, text="Clear", font=Fonts.MICRO,
                                     fg_color="transparent", hover_color=Colors.BG_HOVER,
                                     text_color=Colors.TEXT_MUTED, width=50, height=20,
                                     command=lambda: self._clear_graph_history(history_list))
            clear_btn.pack(side="right")
        
        # History list - scrollable with visible scrollbar and minimum height
        history_list = ctk.CTkScrollableFrame(history_inner, fg_color=Colors.BG_ELEVATED,
                                              corner_radius=Spacing.RADIUS_MD,
                                              height=300,  # Minimum height to show multiple graphs
                                              scrollbar_button_color=Colors.BORDER_DEFAULT,
                                              scrollbar_button_hover_color=Colors.BORDER_BRIGHT)
        history_list.pack(fill="both", expand=True)
        
        self._refresh_graph_history(history_list)
        
        # ========== BUTTONS ARE IN LEFT PANEL (actions_inner) ==========
        def on_generate():
            try:
                selected_graph = graph_var.get()
                if not selected_graph:
                    import tkinter.messagebox as mb
                    mb.showwarning("No Selection", "Please select a graph type.", parent=dialog)
                    return
                
                # Add to history before closing
                history_entry = {
                    'graph_type': selected_graph,
                    'time_period': time_period,
                    'model': selected_model,
                    'turbine': turbine_specs.get('name', 'Unknown'),
                }
                if not self._graph_history or self._graph_history[0] != history_entry:
                    self._graph_history.insert(0, history_entry)
                    self._graph_history = self._graph_history[:10]
                
                # Close the selection dialog first, then generate graph
                dialog.destroy()
                
                # Generate the graph with callback to reopen selector
                self._generate_graph_with_selector(hourly_results, selected_model, turbine_specs, 
                                                   time_period, selected_graph)
            except Exception as e:
                import traceback
                import tkinter.messagebox as mb
                error_msg = f"Error in on_generate: {str(e)}\n{traceback.format_exc()}"
                print(error_msg)
                mb.showerror("Error", f"Failed to generate graph:\n{str(e)}", parent=dialog)
        
        ctk.CTkButton(actions_inner, text="▶  Generate Graph", font=Fonts.BODY_BOLD,
                      fg_color=Colors.ACCENT, hover_color=Colors.ACCENT_DIM,
                      text_color=Colors.TEXT_ON_ACCENT, height=44,
                      corner_radius=Spacing.RADIUS_MD,
                      command=on_generate).pack(fill="x", pady=(0, Spacing.PAD_SM))
        
        # Secondary actions
        btn_row = ctk.CTkFrame(actions_inner, fg_color="transparent")
        btn_row.pack(fill="x")
        btn_row.grid_columnconfigure((0, 1), weight=1)
        
        ctk.CTkButton(btn_row, text="← Back", font=Fonts.BODY,
                      fg_color=Colors.BG_ELEVATED, hover_color=Colors.BG_HOVER,
                      text_color=Colors.TEXT_SECONDARY, height=36,
                      corner_radius=Spacing.RADIUS_MD,
                      command=lambda: [dialog.destroy(), 
                                       self._show_graph_model_selection_dialog(hourly_results, selected_model, 
                                                                                turbine_specs, time_period, selected_model)]
                      ).grid(row=0, column=0, sticky="ew", padx=(0, Spacing.PAD_XS))
        
        ctk.CTkButton(btn_row, text="Close", font=Fonts.BODY,
                      fg_color=Colors.BG_ELEVATED, hover_color="#ef4444",
                      text_color=Colors.TEXT_SECONDARY, height=36,
                      corner_radius=Spacing.RADIUS_MD,
                      command=dialog.destroy).grid(row=0, column=1, sticky="ew", padx=(Spacing.PAD_XS, 0))
    
    def _generate_graph_with_selector(self, hourly_results, selected_model, turbine_specs, time_period, graph_type):
        """Generate graph with ability to reopen selector for switching graphs."""
        self._generate_graph(hourly_results, selected_model, turbine_specs, time_period, graph_type, 
                            allow_switch=True)
    
    def _generate_graph(self, hourly_results, selected_model, turbine_specs, time_period, graph_type, allow_switch=False):
        """Generate and display the selected interactive graph embedded in the app.
        
        Args:
            allow_switch: If True, adds a "New Graph" button to reopen the selection dialog
        """
        import customtkinter as ctk
        import tkinter.messagebox as mb
        
        print(f"Generating graph: {graph_type}, Model: {selected_model}, Period: {time_period}")
        
        # Aggregate data based on time period
        if time_period == "daily":
            display_data = self._aggregate_energy_daily(hourly_results)
        elif time_period == "monthly":
            display_data = self._aggregate_energy_monthly(hourly_results)
        else:
            display_data = hourly_results
        
        if not display_data:
            mb.showerror("No Data", "No data available for graph generation.")
            return
        
        try:
            # Check if matplotlib is available
            try:
                import matplotlib
                matplotlib.use('TkAgg')
            except ImportError:
                mb.showerror("Missing Dependency", 
                    "matplotlib is required for embedded charts.\n\n"
                    "Install with:\n"
                    "pip install matplotlib")
                return
            
            # Close any existing chart window to avoid clutter
            if hasattr(self, '_current_chart_popup') and self._current_chart_popup:
                try:
                    if self._current_chart_popup.winfo_exists():
                        self._current_chart_popup.destroy()
                except:
                    pass
                self._current_chart_popup = None
            
            # Clean up any existing features popup
            if hasattr(self, '_current_features_popup') and self._current_features_popup:
                try:
                    if self._current_features_popup.winfo_exists():
                        self._current_features_popup.destroy()
                except:
                    pass
                self._current_features_popup = None
            
            # Create popup window for the chart
            chart_popup = ctk.CTkToplevel(self)
            chart_popup.title(f"{graph_type} - {selected_model} ({time_period.capitalize()})")
            chart_popup.geometry("1000x700")
            chart_popup.transient(self.winfo_toplevel())
            
            # Store reference for cleanup
            self._current_chart_popup = chart_popup
            
            # Header
            header = ctk.CTkFrame(chart_popup, fg_color=Colors.BG_ELEVATED, height=50, corner_radius=0)
            header.pack(fill="x")
            header.pack_propagate(False)
            ctk.CTkFrame(header, width=4, height=24, fg_color=Colors.ACCENT).pack(side="left", padx=(16, 8), pady=13)
            ctk.CTkLabel(header, text=f"{graph_type} - Interactive Chart", font=Fonts.H3,
                        text_color=Colors.TEXT_DARK).pack(side="left", pady=13)
            
            # Info label with feature highlights
            info_frame = ctk.CTkFrame(chart_popup, fg_color=Colors.BG_CARD)
            info_frame.pack(fill="x", padx=Spacing.PAD_MD, pady=(Spacing.PAD_SM, 0))
            
            info_text = (
                "🖱️ Hover for Tooltips  •  "
                "🔍 Mouse Wheel to Zoom  •  "
                "✋ Click & Drag to Pan  •  "
                "👁️ Click Legend to Toggle Visibility  •  "
                "📅 Use Toolbar or Buttons for Time Range"
            )
            
            ctk.CTkLabel(info_frame, 
                        text=info_text,
                        font=Fonts.CAPTION, text_color=Colors.TEXT_SECONDARY).pack(padx=Spacing.PAD_MD, pady=Spacing.PAD_SM)
            
            # Main content area - use grid to reserve space for buttons
            content_frame = ctk.CTkFrame(chart_popup, fg_color="transparent")
            content_frame.pack(fill="both", expand=True, padx=Spacing.PAD_MD, pady=Spacing.PAD_MD)
            content_frame.grid_rowconfigure(0, weight=1)
            content_frame.grid_columnconfigure(0, weight=1)
            
            # Chart container frame
            chart_container = ctk.CTkFrame(content_frame, fg_color=Colors.BG_CARD)
            chart_container.grid(row=0, column=0, sticky="nsew", pady=(0, Spacing.PAD_MD))
            
            # Create embedded chart with reduced height
            from embedded_chart import create_embedded_chart
            
            chart_obj, chart_widget = create_embedded_chart(
                chart_container,
                display_data,
                selected_model,
                graph_type,
                width=950,
                height=520,  # Height includes chart + navigation toolbar
                dark_mode=True,
                time_period=time_period  # Pass time_period for sequential x-axis (1 to N)
            )
            
            chart_widget.pack(fill="both", expand=True)
            
            # Store reference for cleanup
            self._current_chart = chart_obj
            self._current_features_popup = None
            
            # Show features notification with scrollbar
            features_popup = ctk.CTkToplevel(chart_popup)
            features_popup.title("Interactive Chart Ready")
            features_popup.geometry("450x250")
            features_popup.transient(chart_popup)
            # Note: No grab_set() here so the selection dialog remains interactive
            
            # Track features popup for cleanup
            self._current_features_popup = features_popup
            
            # Scrollable content frame
            scroll_frame = ctk.CTkScrollableFrame(
                features_popup,
                fg_color="transparent",
                scrollbar_button_color=Colors.BORDER_DEFAULT,
                scrollbar_button_hover_color=Colors.BORDER_BRIGHT
            )
            scroll_frame.pack(fill="both", expand=True, padx=Spacing.PAD_MD, pady=Spacing.PAD_MD)
            
            ctk.CTkLabel(scroll_frame, text="✅ Interactive Chart Generated!", 
                        font=Fonts.H3, text_color=Colors.SUCCESS).pack(pady=(10, 10))
            
            features_text = (
                "📊 The chart is now displayed. You can:\n\n"
                "• Click 'New Graph' to switch to a different visualization\n"
                "• Interact with the chart (zoom, pan, hover for tooltips)\n"
                "• Close this notification to view the chart"
            )
            
            ctk.CTkLabel(scroll_frame, text=features_text,
                        font=Fonts.BODY, text_color=Colors.TEXT_PRIMARY,
                        justify="left").pack(pady=10, padx=10)
            
            def close_features():
                if features_popup and features_popup.winfo_exists():
                    features_popup.destroy()
                self._current_features_popup = None
            
            ctk.CTkButton(scroll_frame, text="Got it!", 
                         command=close_features,
                         fg_color=Colors.ACCENT, hover_color=Colors.ACCENT_DIM,
                         text_color=Colors.TEXT_ON_ACCENT, width=100).pack(pady=10)
            
            # Auto-close after 5 seconds
            features_popup.after(5000, close_features)
            
            # Button frame - always visible at bottom with grid
            bottom_section = ctk.CTkFrame(content_frame, fg_color=Colors.BG_CARD,
                                          corner_radius=Spacing.RADIUS_LG,
                                          border_width=2, border_color=Colors.BORDER_DEFAULT)
            bottom_section.grid(row=1, column=0, sticky="ew")
            
            btn_frame = ctk.CTkFrame(bottom_section, fg_color="transparent")
            btn_frame.pack(fill="x", padx=Spacing.PAD_LG, pady=Spacing.PAD_LG)
            
            def close_chart():
                # Clean up features popup first
                if hasattr(self, '_current_features_popup') and self._current_features_popup:
                    try:
                        if self._current_features_popup.winfo_exists():
                            self._current_features_popup.destroy()
                    except:
                        pass
                    self._current_features_popup = None
                
                # Clean up chart object
                if hasattr(self, '_current_chart') and self._current_chart:
                    try:
                        self._current_chart.destroy()
                    except:
                        pass
                    self._current_chart = None
                
                # Destroy the popup window
                try:
                    chart_popup.destroy()
                except:
                    pass
                self._current_chart_popup = None
            
            def switch_graph():
                """Close current chart and reopen selection dialog."""
                print(f"Switching graph - closing current chart and reopening selector")
                close_chart()
                # Reopen the graph type selection dialog
                self._show_graph_type_selection_dialog(
                    hourly_results, selected_model, turbine_specs, time_period, graph_type
                )
            
            # Add "New Graph" button if switching is allowed
            if allow_switch:
                ctk.CTkButton(btn_frame, text="➕  New Graph", font=Fonts.BODY_BOLD,
                             fg_color=Colors.ACCENT, hover_color=Colors.ACCENT_DIM,
                             text_color=Colors.TEXT_ON_ACCENT, height=48,
                             corner_radius=Spacing.RADIUS_MD,
                             command=switch_graph).pack(side="left")
            
            ctk.CTkButton(btn_frame, text="✕  Close", font=Fonts.BODY_BOLD,
                         fg_color=Colors.BG_ELEVATED, hover_color="#ef4444",
                         text_color=Colors.TEXT_PRIMARY, height=48,
                         corner_radius=Spacing.RADIUS_MD,
                         command=close_chart).pack(side="right")
            
            # Handle window close
            chart_popup.protocol("WM_DELETE_WINDOW", close_chart)
            
            print(f"Graph '{graph_type}' created successfully")
            
        except Exception as e:
            import traceback
            error_detail = f"Error generating graph:\n{str(e)}\n\nDetails:\n{traceback.format_exc()}"
            print(error_detail)  # Print to console for debugging
            mb.showerror("Graph Error", f"Error generating graph:\n{str(e)}")
    
    def _refresh_graph_history(self, parent_frame):
        """Refresh the graph history list display."""
        import customtkinter as ctk
        
        # Clear existing content
        for child in parent_frame.winfo_children():
            child.destroy()
        
        if not self._graph_history:
            empty_frame = ctk.CTkFrame(parent_frame, fg_color="transparent")
            empty_frame.pack(fill="both", expand=True, pady=Spacing.PAD_LG)
            
            ctk.CTkLabel(empty_frame, text="📊", font=(Fonts.FAMILY, 32),
                        text_color=Colors.TEXT_MUTED).pack(pady=(Spacing.PAD_MD, 0))
            ctk.CTkLabel(empty_frame, text="No graphs generated yet",
                        font=Fonts.BODY, text_color=Colors.TEXT_SECONDARY).pack(pady=Spacing.PAD_SM)
            ctk.CTkLabel(empty_frame, text="Select a graph type and click Generate",
                        font=Fonts.CAPTION, text_color=Colors.TEXT_MUTED).pack()
            return
        
        for i, entry in enumerate(self._graph_history):
            # History item card
            if i == 0:
                # Latest item - with accent border
                item = ctk.CTkFrame(parent_frame, fg_color=Colors.BG_CARD,
                                   corner_radius=Spacing.RADIUS_MD,
                                   border_width=1,
                                   border_color=Colors.ACCENT)
            else:
                # Older items - no border
                item = ctk.CTkFrame(parent_frame, fg_color=Colors.BG_ELEVATED,
                                   corner_radius=Spacing.RADIUS_MD)
            item.pack(fill="x", pady=Spacing.PAD_XS)
            
            item_inner = ctk.CTkFrame(item, fg_color="transparent")
            item_inner.pack(fill="x", padx=Spacing.PAD_SM, pady=Spacing.PAD_SM)
            
            # Latest indicator
            if i == 0:
                latest_badge = ctk.CTkFrame(item_inner, fg_color=Colors.ACCENT,
                                           corner_radius=Spacing.RADIUS_SM, width=6, height=6)
                latest_badge.pack(side="left", padx=(0, Spacing.PAD_XS))
                latest_badge.pack_propagate(False)
            
            # Graph info
            info_text = f"{entry['graph_type']}"
            sub_text = f"{entry['time_period'].capitalize()} • {entry['model']}"
            
            text_frame = ctk.CTkFrame(item_inner, fg_color="transparent")
            text_frame.pack(side="left", fill="y", expand=True)
            
            ctk.CTkLabel(text_frame, text=info_text, font=Fonts.BODY_BOLD if i == 0 else Fonts.BODY,
                        text_color=Colors.TEXT_PRIMARY).pack(anchor="w")
            ctk.CTkLabel(text_frame, text=sub_text, font=Fonts.CAPTION,
                        text_color=Colors.TEXT_MUTED).pack(anchor="w")
    
    def _clear_graph_history(self, history_list):
        """Clear the graph history."""
        self._graph_history = []
        self._refresh_graph_history(history_list)
